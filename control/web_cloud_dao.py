#!loginusr/bin/python3
# -*- coding: UTF-8 -*-

# Author   : JasonHung
# Date     : 20221102
# Update   : 20230720 , 20240930
# Function : otsuka factory work time record

import pymysql , logging , time , re , requests , json , calendar , csv , json , smtplib
import openpyxl , pyodbc , sys , os , base64 , io , asyncio , aiomysql , aioodbc , shutil
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import mplcursors
import pandas as pd
import numpy as np

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from bs4 import BeautifulSoup
from io import BytesIO
from fpdf import *
from flask import jsonify
from decimal import Decimal
from control.config.config import parameter
from openpyxl.styles import Font , PatternFill , Alignment
from ldap3 import Server, Connection, ALL, NTLM, SAFE_SYNC, MODIFY_REPLACE
from ldap3.core.exceptions import LDAPException 
from datetime import datetime , date
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas

### config
from control.config.config import parameter

########
# log
########
log_format = "%(asctime)s %(message)s"
logging.basicConfig(format=log_format , level=logging.INFO , datefmt="%Y-%m-%d %H:%M:%S")

########################################################################################################################################
#
# async_web_cloud_dao
#
########################################################################################################################################
class async_web_cloud_dao:
   
    ######################
    # async_mysql_query
    ######################
    async def async_mysql_query(self , pool , query):
        try:
            async with pool.acquire() as conn:
                async with conn.cursor() as curr:
                    await curr.execute(f"{query}")
                    res = await curr.fetchall()
                    return res
        except Exception as e:
            logging.info(f"<Error> async mysql query : {str(e)}\n")
        finally:
            pass

    ############################
    # __async_connect_mysql__ 
    ############################
    async def __async_connect_mysql__(self):
        try:
            conn_mysql = f"host='{parameter.otsuka_factory8['host']}',port='{parameter.otsuka_factory8['port']}',user='{parameter.otsuka_factory8['user']}',password='{parameter.otsuka_factory8['password']}',db='{parameter.otsuka_factory8['db']}',charset='{parameter.otsuka_factory8['charset']}',minsize=5,maxsize=10"
            pool = await aiomysql.create_pool(host=parameter.otsuka_factory['host'],port=parameter.otsuka_factory['port'],user=parameter.otsuka_factory['user'],password=parameter.otsuka_factory['pwd'],db=parameter.otsuka_factory['db'],charset=parameter.otsuka_factory['charset'])
            return pool
        except Exception as e:
            logging.info(f"<Error> async connect mysql : {str(e)}\n")
        finally:
            pass

    ######################
    # async_mssql_query
    ######################
    async def async_mssql_query(self, pool , query):
        try:
            async with pool.acquire() as conn:
                async with conn.cursor() as cur:
                    await cur.execute(f"{query}")
                    rows = await cur.fetchall()
                    return rows
        except Exception as e:
            logging.info(f"<Error> async query_db : {str(e)}\n")
        finally:
            pass

    ############################
    # __async_connect_mssql__ 
    ############################
    async def __async_connect_mssql__(self):
        
        try:
            ########################################
            # async ODBC Driver 18 for SQL Server 
            ########################################
            conn_str = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={parameter.parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            pool = await aioodbc.create_pool(conn_str)
            return pool
            
        except Exception as e:
            logging.info("\n<Error> __async_connect_mssql__ " + str(e))

        finally:
            pass
            #self.curr_mssql.close()
            #self.conn_mssql.close()

########################################################################################################################################
#
# web_cloud_dao
#
########################################################################################################################################
class web_cloud_dao:

    def generate_plot():
        # 创建一个简单的 Matplotlib 图表
        fig, ax = Figure(), FigureCanvas(Figure())
        ax.plot([1, 2, 3, 4, 5], [2, 4, 6, 8, 10])
        return fig

    ##########################
    # show_day_money_detail
    ##########################
    def show_day_money_detail(self , year , month):
        
        self.__connect__()
        
        try:
            month = '0' + month if int(month) < 10 else month

            year_month_sql = f"SELECT day_r_date , day_r_month , day_r_day FROM `day_money` WHERE day_r_year='2023' and day_r_month='09' group by day_r_day order by day_r_day desc"
            name_sql       = f"select a.a_name , b.employee_eng_name , a.day_r_date , a.day_t_money from `day_money` a left join hr_a b on a.a_name = b.employee_name  WHERE  day_r_year='2023' and day_r_month='09' and day_r_month != '9/' order by day_r_day desc"

            self.sql = f"select a_name , day_r_month from `day_money` WHERE  day_r_year='{str(year)}' and day_r_month='{str(month)}' and day_r_month != '9/' order by day_r_day desc" 
            self.curr.execute(self.sql)
            self.res = self.curr.fetchall()
            month    = []

            for val in self.res:
                month.append(val[0]) 

            return month

        except Exception as e:
            logging.error('\n<Error> show_day_money_detail : ' + str(e))

        finally:
            self.__disconnect__()

    ################################
    # show_day_money_detail_money
    ################################
    '''
    def show_day_money_detail_money(self , year , month):
        
        self.__connect__()
        
        try:
            month = '0' + month if int(month) < 10 else month

            name_sql = f"select a_name , e_name from day_money where day_r_year='{year}' and day_r_month='{month}' group by a_name  order by day_r_day asc"
            self.curr.execute(name_sql)
            name_res = self.curr.fetchall()

            for name_val in name_res:
             
                money_sql = f"select day_t_money from day_money where day_r_year='{year}' and day_r_month='{month}' and a_name='{name}' order by day_r_day asc"
                self.curr.execute(money_sql)
                money_res = self.curr.fetchall()

                return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_detail_money : ' + str(e))

        finally:
            self.__disconnect__()
    '''

    ################################################
    # show_day_money_parking_fee_detail_day_total
    ################################################
    def show_day_money_parking_fee_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_parking_fee where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_parking_fee_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # show_day_money_over_traffic_detail_day_total
    ############################################
    def show_day_money_over_traffic_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_over_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_over_traffic_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # show_day_money_traffic_detail_day_total
    ############################################
    def show_day_money_traffic_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_traffic_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################################
    # show_day_money_tolls_detail_day_total
    ##########################################
    def show_day_money_tolls_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_tolls where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_tolls_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################################
    # show_day_money_trick_detail_day_total
    ##########################################
    def show_day_money_trick_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_trick where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_trick_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #########################################
    # show_day_money_taxi_detail_day_total
    #########################################
    def show_day_money_taxi_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_taxi where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_taxi_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #########################################
    # show_day_money_stay_detail_day_total
    #########################################
    def show_day_money_stay_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_stay where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_stay_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #########################################
    # show_day_money_other_detail_day_total
    ##########################################
    def show_day_money_other_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_other_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    #######################################
    # show_day_money_oil_detail_day_total
    ########################################
    def show_day_money_oil_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money_oil where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_oil_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()

    ####################################
    # show_day_money_detail_day_total
    ####################################
    def show_day_money_detail_day_total(self , year , month):
        
        self.__connect__()
        
        try:
            #month = '0' + month if int(month) < 10 else month

            money_sql  = f"select " 
            money_sql += f"format(sum(day_t_money1),0) , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            money_sql += f"format(sum(day_t_money7),0) , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            money_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            money_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            money_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            money_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            money_sql += f"from day_money where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_detail_day_total : ' + str(e))

        finally:
            self.__disconnect__()
    
    ###########################################
    # show_day_money_parking_fee_detail_name
    ###########################################
    def show_day_money_parking_fee_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_parking_fee where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/parking_fee_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/parking_fee_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/parking_fee_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_parking_fee where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_parking_fee_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ######################################
    # show_day_money_over_traffic_detail_name
    ######################################
    def show_day_money_over_traffic_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_over_traffic where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/over_traffic_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/over_traffic_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/over_traffic_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_over_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_over_traffic_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ######################################
    # show_day_money_traffic_detail_name
    ######################################
    def show_day_money_traffic_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_traffic where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/traffic_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/traffic_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/traffic_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_traffic where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_traffic_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # show_day_money_tolls_detail_name
    #####################################
    def show_day_money_tolls_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_tolls where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/tolls_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/tolls_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/tolls_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_tolls where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_tolls_detail_name (日當 過路費) : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # show_day_money_trick_detail_name
    #####################################
    def show_day_money_trick_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_trick where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/trick_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/trick_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/trick_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_trick where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_trick_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ####################################
    # show_day_money_taxi_detail_name
    ####################################
    def show_day_money_taxi_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_taxi where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/taxi_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/taxi_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/taxi_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_taxi where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_taxi_detail_name : (日當  計程車)' + str(e))

        finally:
            self.__disconnect__()

    ####################################
    # show_day_money_stay_detail_name
    ####################################
    def show_day_money_stay_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_stay where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/stay_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/stay_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/stay_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_stay where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_stay_detail_name : (日當 住宿)' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # show_day_money_other_detail_name
    #####################################
    def show_day_money_other_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/other_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/other_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/other_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_other_detail_name : (日當  其他)' + str(e))

        finally:
            self.__disconnect__()

    ###################################
    # show_day_money_oil_detail_name
    ###################################
    def show_day_money_oil_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money_oil where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/oil_'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/oil_'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/oil_'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_oil where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_oil_detail_name : (日當 油票)' + str(e))

        finally:
            self.__disconnect__()

    ##############################
    # show_day_money_detail_name
    ##############################
    def show_day_money_detail_name(self , year , month):
        
        self.__connect__()
        
        try:
            money_sql  = f"select a_name , e_name , d_name , " 
            money_sql += f"day_t_money1 , day_t_money2 , day_t_money3 , day_t_money4  , day_t_money5  , day_t_money6  , day_t_money7  , day_t_money8  , day_t_money9  , day_t_money10 , " 
            money_sql += f"day_t_money11 , day_t_money12 , day_t_money13 , day_t_money14  , day_t_money15  , day_t_money16  , day_t_money17  , day_t_money18  , day_t_money19  , day_t_money20 , "
            money_sql += f"day_t_money21 , day_t_money22 , day_t_money23 , day_t_money24  , day_t_money25  , day_t_money26  , day_t_money27  , day_t_money28  , day_t_money29  , day_t_money30 , day_t_money31 , day_t_total "
            money_sql += f"from day_money where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/'+ year + '_' + month + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"中文,英文,部門,{month}/1,{month}/2,{month}/3,{month}/4,{month}/5,{month}/6,{month}/7,{month}/8,{month}/9,{month}/10,{month}/11,{month}/12,{month}/13,{month}/14,{month}/15,{month}/16,{month}/17,{month}/18,{month}/19,{month}/20,{month}/21,{month}/22,{month}/23,{month}/24,{month}/25,{month}/26,{month}/27,{month}/28,{month}/29,{month}/30,{month}/31,總計"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/'+ year + '_' + month + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[1]).encode('utf8').decode('latin1')+' , '+str(val[5]).encode('utf8').decode('latin1')+' , '+str(val[13]).encode('utf-8').decode('latin1')+' , '+str(val[6]).encode('utf8').decode('latin1')+' , '+str(val[7]).encode('utf8').decode('latin1')+' , '+str(val[8]).encode('utf8').decode('latin1')+' , '+str(val[9]).encode('utf8').decode('latin1')+' , '+str(val[10]).encode('utf8').decode('latin1')+' , '+str(val[11]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/'+ year + month + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['中文','英文','部門',f'{month}/1',f'{month}/2',f'{month}/3',f'{month}/4',f'{month}/5',f'{month}/6',f'{month}/7',f'{month}/8',f'{month}/9',f'{month}/10'
                     ,f'{month}/11',f'{month}/12',f'{month}/13',f'{month}/14',f'{month}/15',f'{month}/16',f'{month}/17',f'{month}/18',f'{month}/19',f'{month}/20'
                     ,f'{month}/21',f'{month}/22',f'{month}/23',f'{month}/24',f'{month}/25',f'{month}/26',f'{month}/27',f'{month}/28',f'{month}/29',f'{month}/30',f'{month}/31','總計']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money where day_r_year='{year}' and day_r_month='{month}' order by d_name asc"
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)

            workbook.save(excel_file)

            return money_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_detail_name : ' + str(e))

        finally:
            self.__disconnect__()

    ##############################
    # show_day_money_detail_day
    ##############################
    def show_day_money_detail_day(self , year , month):
        
        self.__connect__()
        
        try:
            month = '0' + month if int(month) < 10 else month

            # all day by month 
            day_sql = f"select day_r_month , day_r_day from day_money where day_r_year='{year}' and day_r_month='{month}' group by day_r_day order by day_r_day asc"
            self.curr.execute(day_sql)
            day_res = self.curr.fetchall() 

            return day_res

        except Exception as e:
            logging.error('\n<Error> show_day_money_detail_day : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # show_computer_chi_name
    ###########################
    def show_computer_chi_name(self , d_name):
        
        self.__connect__()
        
        try:
            # device chinese name
            d_sql = f"select employee_name from hr_a where d_name='{d_name}'"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchone() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_computer_chi_name : ' + str(e))

        finally:
            self.__disconnect__()

    ##############################
    # sensor_position_detail
    ##############################
    def sensor_position_detail(self , sensor):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user by s_number
            d_sql  = f"SELECT r_time , val_1 , val_2 from {r_date} where s_kind='{sensor}' order by no desc limit 0,1"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> sensor_position_detail : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################
    # computer_s_number_detail
    ##############################
    def computer_s_number_detail(self):
        
        self.__connect__()
        
        try:
            # all device user by s_number
            d_sql  = f"SELECT s_number , d_name , d_status , r_date FROM (SELECT s_number , d_name , d_status , r_date, ROW_NUMBER() OVER (PARTITION BY s_number ORDER BY r_date DESC) AS rn FROM device_list) AS subquery WHERE rn = 1 ORDER BY r_date DESC"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> computer_s_number_detail : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # search_show_computer_user_detail
    #####################################
    def search_show_computer_user_detail(self , s_number):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"select distinct d_name from device_list where s_number='{s_number}' order by l_activity desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_computer_user_detail : ' + str(e))

        finally:
            self.__disconnect__()

    ######################################
    # show_factory_monitor_position_img
    ######################################
    def show_factory_monitor_position_img(self):
        
        self.__connect4__()
        try:
            # record time
            now_month = time.strftime("%Y_%m" , time.localtime()) 
            
            # all device position
            d_sql = f"select distinct b.d_name , b.d_c_name from {now_month} a left join monitor_device b on a.s_kind=b.d_name where a.s_kind != 'I6-1' and a.s_kind !='I6-2' order by b.d_name asc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x = [row[0] for row in d_res]  
            y1 = [row[1] for row in d_res]  
            y2 = [row[2] for row in d_res]  

            # 生成多条线图
            plt.plot(x, y1, label='Line 1')
            plt.plot(x, y2, label='Line 2')
            plt.xlabel('X Label')
            plt.ylabel('Y Label')
            plt.title('MySQL Data Multiple Line Chart')
            plt.legend()

            # 保存图像到内存中
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)

            # 将图像转换为base64编码
            plot_url = base64.b64encode(img.getvalue()).decode()


            return plot_url

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_position_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ##################################
    # show_factory_monitor_position
    ##################################
    def show_factory_monitor_position(self):
        
        self.__connect4__()
        try:
            # record time
            now_month = time.strftime("%Y_%m" , time.localtime()) 
            
            # all device position
            d_sql = f"select distinct b.d_name , b.d_c_name from {now_month} a left join monitor_device b on a.s_kind=b.d_name where a.s_kind != 'I6-1' and a.s_kind !='I6-2' order by b.d_name asc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_position : ' + str(e))

        finally:
            self.__disconnect4__()

    ######################################
    # show_factory_monitor_detail_chart
    ######################################
    def show_factory_monitor_detail_chart(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,20"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_rh_img
    #########################################
    def show_factory_monitor_detail_rh_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[4]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='RH' ,  marker='o', markersize=4)
            axis.set_title('RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')
            
            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_rh_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_rh_val
    #########################################
    def show_factory_monitor_detail_rh_val(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select min(val_2) , max(val_2) , ROUND(AVG(val_2),2) from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_rh_val : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_temp_val
    #########################################
    def show_factory_monitor_detail_temp_val(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select min(val_1) , max(val_1) , ROUND(AVG(val_1),2) from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_temp_val : ' + str(e))

        finally:
            self.__disconnect4__()

    #############################################
    # show_factory_monitor_detail_rh_pie_img
    #############################################
    def show_factory_monitor_detail_rh_pie_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select val_2 , count(*) from {r_date} where s_kind='{s_kind}' group by val_2 order by val_2 desc limit 0,8"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            val = [row[0] for row in d_res]  # x轴数据
            count = [float(row[1]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            # 创建圆饼图
            
            axis.pie(count, labels=val, autopct='%1.1f%%', startangle=90)
            axis.axis('equal')  # 保证饼图是圆形的
            axis.set_title('RH')

            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_rh_pie_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #############################################
    # show_factory_monitor_detail_temp_pie_img
    #############################################
    def show_factory_monitor_detail_temp_pie_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select val_1 , count(*) from {r_date} where s_kind='{s_kind}' group by val_1 order by val_1 desc limit 0,8"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            val = [row[0] for row in d_res]  # x轴数据
            count = [float(row[1]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            # 创建圆饼图
            
            axis.pie(count, labels=val, autopct='%1.1f%%', startangle=90)
            axis.axis('equal')  # 保证饼图是圆形的
            axis.set_title('Temp')

            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_temp_pie_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################
    # show_factory_monitor_detail_temp_img
    #########################################
    def show_factory_monitor_detail_temp_img(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[3]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='Temp' , marker='o', markersize=4)
            axis.set_title('Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')

            # 启用鼠标悬停提示显示数值
            mplcursors.cursor(fig).connect("add", lambda axis: axis.annotation.set_text(f"{axis.target[0]:.2f},{axis.target[1]:.2f}"))
            # 添加提示文本（手动方式）
            #for i, txt in enumerate(y1):
            #    axis.annotate(f'{txt:.2f}', (x[i], y1[i]), textcoords="offset points", xytext=(0,10), ha='center')

            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_temp_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################################
    # show_factory_monitor_detail_temp_rh_img_2
    ##############################################
    def show_factory_monitor_detail_temp_rh_img_2(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[4]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='RH' ,  marker='o', markersize=4)
            axis.set_title(s_kind + ' RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')

            # 启用鼠标悬停提示显示数值
            mplcursors.cursor(fig).connect("add", lambda axis: axis.annotation.set_text(f"{axis.target[0]:.2f},{axis.target[1]:.2f}"))
            # 添加提示文本（手动方式）
            #for i, txt in enumerate(y1):
            #    axis.annotate(f'{txt:.2f}', (x[i], y1[i]), textcoords="offset points", xytext=(0,10), ha='center')
            
            axis.legend()
            fig.tight_layout()
    
            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_temp_rh_img_2 : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################
    # show_check_vm_state
    #########################
    def show_check_vm_state(self , position , state):
        
        self.__connect__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## vm_os_state
            vm_os_state_sql = f"SELECT c_date , c_time , vm_name , vm_os_state FROM `otsuka_vmware` where vm_position='{position}' and vm_os_state='{state}' group by vm_name order by c_d_time desc"
            self.curr.execute(vm_os_state_sql)
            vm_os_state_res = self.curr.fetchall() 
            
            return vm_os_state_res

        except Exception as e:
            logging.error(f"\n<Error> show_check_vm_state : {str(e)}\n")

        finally:
            self.__disconnect__()

    ##########################################
    # show_taipei_monitor_detail_vmware_img
    ##########################################
    def show_taipei_monitor_detail_vmware_img(self , vm_name):
        
        self.__connect__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## vm_os_state
            vm_os_state_sql = f"select c_date , c_time , vm_name , vm_os_state from otsuka_vmware where vm_name='{vm_name}' order by c_date , c_time desc limit 0,36"
            self.curr.execute(vm_os_state_sql)
            vm_os_state_res = self.curr.fetchall() 

            x  = [val[1] for val in vm_os_state_res]  # x轴数据
            y1 = [val[3] for val in vm_os_state_res]  # vm_os_status
            
            ### 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            
            ### 設定 y 軸標籤為英文
            axis.yaxis.set_major_formatter(plt.NullFormatter())
            axis.plot(x, y1 , label=vm_name)

            axis.set_title(f"Taipei VMware {vm_name} status")
            axis.set_xlabel('Time')
            axis.set_ylabel('Status')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            ### Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_taipei_monitor_detail_vmware_img ({vm_name}) : {str(e)}\n")

        finally:
            self.__disconnect__()

    #########################################################
    # show_factory_monitor_detail_warehouse_rh_img
    #########################################################
    def show_factory_monitor_detail_warehouse_rh_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-1
            s_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-1' order by r_time desc limit 0,40"
            self.curr.execute(s_1_sql)
            s_1_res = self.curr.fetchall() 
            ### S-2
            s_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-2' order by r_time desc limit 0,40"
            self.curr.execute(s_2_sql)
            s_2_res = self.curr.fetchall() 
            ### S-3
            s_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-3' order by r_time desc limit 0,40"
            self.curr.execute(s_3_sql)
            s_3_res = self.curr.fetchall() 
            ### S-4
            s_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-4' order by r_time desc limit 0,40"
            self.curr.execute(s_4_sql)
            s_4_res = self.curr.fetchall() 
            ### S-5
            s_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-5' order by r_time desc limit 0,40"
            self.curr.execute(s_5_sql)
            s_5_res = self.curr.fetchall() 
            ### S-6
            s_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-6' order by r_time desc limit 0,40"
            self.curr.execute(s_6_sql)
            s_6_res = self.curr.fetchall() 
            ### S-7
            s_7_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-7' order by r_time desc limit 0,40"
            self.curr.execute(s_7_sql)
            s_7_res = self.curr.fetchall() 
            ### S-8
            s_8_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-8' order by r_time desc limit 0,40"
            self.curr.execute(s_8_sql)
            s_8_res = self.curr.fetchall() 
            ### S-9
            s_9_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-9' order by r_time desc limit 0,40"
            self.curr.execute(s_9_sql)
            s_9_res = self.curr.fetchall() 
            ### S-10
            s_10_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-10' order by r_time desc limit 0,40"
            self.curr.execute(s_10_sql)
            s_10_res = self.curr.fetchall() 
            ### S-14
            s_14_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-14' order by r_time desc limit 0,40"
            self.curr.execute(s_14_sql)
            s_14_res = self.curr.fetchall() 
            ### S-17
            s_17_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-17' order by r_time desc limit 0,40"
            self.curr.execute(s_17_sql)
            s_17_res = self.curr.fetchall() 
            ### S-18
            s_18_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-18' order by r_time desc limit 0,40"
            self.curr.execute(s_18_sql)
            s_18_res = self.curr.fetchall() 

            x  = [row[0] for row in s_1_res]  # x轴数据
            y1 = [float(row[4]) for row in s_1_res]  # S-1
            y2 = [float(row[4]) for row in s_2_res]  # S-2
            y3 = [float(row[4]) for row in s_3_res]  # S-3
            y4 = [float(row[4]) for row in s_4_res]  # S-4
            y5 = [float(row[4]) for row in s_5_res]  # S-5
            y6 = [float(row[4]) for row in s_6_res]  # S-6
            y7 = [float(row[4]) for row in s_7_res]  # S-7
            y8 = [float(row[4]) for row in s_8_res]  # S-8
            y9 = [float(row[4]) for row in s_9_res]  # S-9
            y10 = [float(row[4]) for row in s_10_res]  # S-10
            y14 = [float(row[4]) for row in s_14_res]  # S-14
            y17 = [float(row[4]) for row in s_17_res]  # S-17
            y18 = [float(row[4]) for row in s_18_res]  # S-14
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='S-1')
            axis.plot(x, y2 , label='S-2')
            axis.plot(x, y3 , label='S-3')
            axis.plot(x, y4 , label='S-4')
            axis.plot(x, y5 , label='S-5')
            axis.plot(x, y6 , label='S-6')
            axis.plot(x, y7 , label='S-7')
            axis.plot(x, y8 , label='S-8')
            axis.plot(x, y9 , label='S-9')
            axis.plot(x, y10 , label='S-10')
            axis.plot(x, y14 , label='S-14')
            axis.plot(x, y17 , label='S-17')
            axis.plot(x, y18 , label='S-18')

            axis.set_title('Warehouse Sensor Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_warehouse_rh_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ###############################
    # hr_data_3
    ###############################
    def hr_data_3(self , user , sheet_page):
        
        try:
            ### variable
            r_date      = time.strftime("%Y_%m" , time.localtime())
            r_time      = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            excel_file  = parameter.hr_data['hr_file_1'] # HR 員工明細資料 
            excel_file2 = parameter.hr_data['hr_file_2'] # HR 人事異動單201701起
            excel_file3 = parameter.hr_data['hr_file_3'] # HR 2023-2024訓練記錄-Amy

            # 使用 pandas 讀取 Excel 文件
            df  = pd.read_excel(excel_file)
            df3 = pd.read_excel(excel_file3 , sheet_name=f"{sheet_page}")

            hr_file3 = []
            
            for i in range(len(df)):
                row = df.iloc[i]
                if str(row['登入帳號']).lower() == user.lower():
                    chi_name = row['中文姓名']
                    break 

            if chi_name:
                for i in range(len(df3)):
                    row = df3.iloc[i]
                    if row['姓名'] == chi_name:
                        hr_file3.append((row)) 

            '''
            for index , row in df.iterrows():
                if row['登入帳號'] == user:
                    chi_name = row['中文姓名']
            
            for index , row in df3.iterrows():
                if row['姓名'] == chi_name:
                    hr_file3.append((row))
            '''

            return hr_file3

        except Exception as e:
            logging.error(f"\n<Error> hr_data_3 : {str(e)}\n")

        finally:
            pass
    
    ###############################
    # hr_data_2
    ###############################
    def hr_data_2(self , user):
        
        try:
            ### variable
            r_date      = time.strftime("%Y_%m" , time.localtime())
            r_time      = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            excel_file  = parameter.hr_data['hr_file_1'] # HR 員工明細資料 
            excel_file2 = parameter.hr_data['hr_file_2'] # HR 人事異動單201701起
            excel_file3 = parameter.hr_data['hr_file_3'] # HR 2023-2024訓練記錄-Amy

            # 使用 pandas 讀取 Excel 文件
            df  = pd.read_excel(excel_file)
            df2 = pd.read_excel(excel_file2)

            hr_file2 = []
            
            for i in range(len(df)):
                row = df.iloc[i]
                if str(row['登入帳號']).lower() == user.lower():
                    chi_name = row['中文姓名']
                    break  
            
            
            if chi_name:
                for i in range(len(df2)):
                    row = df2.iloc[i]
                    if row['員工姓名'] == chi_name:
                        hr_file2.append((row)) 

            '''
            for index, row in df.iterrows():
                if row['登入帳號'] == user:
                    chi_name = row['中文姓名']
                
            for index , row in df2.iterrows():
                if row['員工姓名'] == chi_name:
                    hr_file2.append((row))
            '''
            
            return hr_file2

        except Exception as e:
            logging.error(f"\n<Error> hr_data_2 : {str(e)}\n")

        finally:
            pass
    
    ###############################
    # hr_data
    ###############################
    def hr_data(self , user):
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## HR 員工明細資料.xls file
            excel_file = parameter.hr_data['hr_file_1']

            # 使用 pandas 讀取 Excel 文件
            df = pd.read_excel(excel_file)

            hr_file = []

            # 遍歷 DataFrame 的每一行
            for i in range(len(df)):
                row = df.iloc[i]
                if str(row['登入帳號']).lower() == user.lower():
                    hr_file.append((row))  

            return hr_file

        except Exception as e:
            logging.error(f"\n<Error> hr_data : {str(e)}\n")

        finally:
            pass

    ###############################
    # factory_erp_realtime_query
    ###############################
    def factory_erp_realtime_query(self):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## chart 
            s_1_sql  = f"SELECT format(sum(ITEM98) , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchone() 

            return s_1_res[0]

        except Exception as e:
            logging.error(f"\n<Error> factory_erp_realtime_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    ####################################################
    # show_bpm_information_by_dep_kind_by_account_sum
    ####################################################
    def show_bpm_information_by_dep_kind_by_account_sum(self , dep , kind , user):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            if kind == '硬體,':
                label_title = 'Hardware'
                title = f"BPM - information form {label_title}"
            else:
                label_title = 'Software'
                title = f"BPM - information form {label_title}"

            ## chart 
            s_1_sql  = f"SELECT format(sum(ITEM98) , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchone() 

            return s_1_res[0]

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_by_account_sum : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #########################################
    # show_bpm_information_by_dep_kind_sum
    #########################################
    def show_bpm_information_by_dep_kind_sum(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            if kind == '硬體,':
                label_title = 'Hardware'
                title = f"BPM - information form {label_title}"
            else:
                label_title = 'Software'
                title = f"BPM - information form {label_title}"

            ## chart 
            s_1_sql  = f"SELECT format(sum(ITEM98) , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchone() 

            return s_1_res[0]

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_sum : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    ################################################################
    # show_bpm_information_by_dep_kind_statistics_by_account_list
    ################################################################
    def show_bpm_information_by_dep_kind_statistics_by_account_list(self , dep , account):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            res = []

            ## chart 
            statistics_sql  = f"SELECT ITEM147 , count(*)  FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{account}' and ITEM35='true' group by ITEM147"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            for val in statistics_res:
                res.append((str(val[0]).replace(",","").replace("_"," - ") , val[1]))
            
            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_statistics_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #####################################################
    # show_bpm_information_by_each_dep_statistics_list
    #####################################################
    def show_bpm_information_by_each_dep_statistics_list(self):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            res = []

            ## chart 
            statistics_sql  = f"SELECT ITEM19 , count(*) FROM ART01231708483412487_INS where ITEM35='true' group by ITEM19"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            for val in statistics_res:
                res.append((str(val[0]).replace(",","").replace("_"," - ") , val[1]))
            
            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_each_dep_statistics_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #####################################################
    # show_bpm_expenditure_form_search_money_sum
    #####################################################
    def show_bpm_expenditure_form_search_money_sum(self , q_s_date , q_e_date , q_b_e_dep , q_b_e_d_member , q_b_e_status , q_b_b_s_b_budget):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            q_s_date = str(q_s_date).replace("-","/")
            q_e_date = str(q_e_date).replace("-","/")

            ## chart 
            if q_b_e_status == '已結案':
                if q_b_e_d_member == '':
                    if q_b_b_s_b_budget == '':
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else:
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                else: 
                    if q_b_b_s_b_budget == '':
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' "
                    else:
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' "
            elif q_b_e_status == '未結案':
                if q_b_e_d_member == '':
                    if q_b_b_s_b_budget == '':
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else:
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                else:
                    if q_b_b_s_b_budget == '':
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else:
                        statistics_sql  = f"select format(sum(ITEM15),'N0','en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    
                
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchone() 

            return statistics_res[0]
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_form_search_money_sum : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #####################################################
    # show_erp_realtime_query_search3
    #####################################################
    def show_erp_realtime_query_search3(self , q_e_p_num3 , q_e_p_a_num3):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            sql  = f"SELECT "
            sql += f"A.TE004 AS '原物料代號', "
            sql += f"LEFT(A.TE004,1) AS '原＆物', "
            sql += f"A.TE010 AS '原物料批號', "
            sql += f"A.TE001 AS '領退料單別', "
            sql += f"A.TE002 AS '領退料單號', "
            sql += f"A.TE011 AS '製令單別', "
            sql += f"A.TE012 AS '製令單號', "
            sql += f"FORMAT(A.TE005 , 'N0' , 'en-US' ) AS '領退料數量', "
            sql += f"A.TE006 AS '單位', "
            sql += f"B.TA063 AS '產品批號', "
            sql += f"B.TA015 AS '預計產量', "
            sql += f"FORMAT(B.TA017 , 'N0' , 'en-US') AS '已生產量', "
            sql += f"B.TA006 AS '產品品號', "
            sql += f"C.MB002 AS '產品名稱', "
            sql += f"D.ME003 AS '最早入庫日', "
            sql += f"D.ME009 AS '有效日期', "
            sql += f"D.ME010 AS '複檢日期', "
            sql += f"E.MB002 AS '原物料品名', "
            sql += f"E.MB032 AS '供應商代號', "
            sql += f"F.MA002 AS '供應商名稱' "
            sql += f"FROM OtsukaDB.dbo.MOCTE A "
            sql += f"JOIN OtsukaDB.dbo.MOCTA B ON A.TE011 = B.TA001 AND A.TE012 = B.TA002 "
            sql += f"JOIN OtsukaDB.dbo.INVMB C ON B.TA006 = C.MB001 "
            sql += f"LEFT JOIN OtsukaDB.dbo.INVME D ON A.TE004 = D.ME001 AND A.TE010 = D.ME002 "
            sql += f"LEFT JOIN OtsukaDB.dbo.INVMB E ON D.ME001 = E.MB001 "
            sql += f"LEFT JOIN OtsukaDB.dbo.PURMA F ON E.MB032 = F.MA001 "
            sql += f"WHERE B.TA009 >= '20240101' AND D.ME003 >= '20200101' AND B.TA006='{q_e_p_num3}' AND B.TA063='{q_e_p_a_num3}'"
 
                
            self.curr_mssql_erp.execute(sql)
            res = self.curr_mssql_erp.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_erp_realtime_query_search3 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()


    #####################################################
    # show_erp_realtime_query_search2
    #####################################################
    def show_erp_realtime_query_search2(self , q_e_p_i_y_month):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            sql  = f"SELECT "
            sql += f"'生產' AS '生產', "
            sql += f"F.TF001 AS '入庫單別', "
            sql += f"F.TF002 AS '入庫單號', "
            sql += f"F.TF003 AS '入庫日期', "
            sql += f"CONVERT(VARCHAR(8), DATEADD(DAY, 14, CONVERT(DATE, F.TF003, 112)), 112) AS '預計合格日期', "
            sql += f"LEFT(F.TF003,4)+SUBSTRING(F.TF003,5,2) AS '入庫年月', "
            sql += f"G.TG003 AS '序號', "
            sql += f"G.TG004 AS '產品品號', "
            sql += f"G.TG005 AS '品名', "
            sql += f"G.TG006 AS '規格', "
            sql += f"G.TG007 AS '單位', "
            sql += f"G.TG009 AS '入/出別', "
            sql += f"G.TG010 AS '庫別', "
            sql += f"C.MC002 AS '庫別名稱', "
            sql += f"FORMAT(G.TG011 , 'N0' , 'en-US') AS '入庫數量', "
            sql += f"G.TG014 AS '製令單別', "
            sql += f"G.TG015 AS '製令單號', "
            sql += f"G.TG017 AS '批號', "
            sql += f"G.TG018 AS '有效日期' "
            sql += f"FROM OtsukaDB.dbo.MOCTF F "
            sql += f"JOIN OtsukaDB.dbo.MOCTG G ON F.TF001 = G.TG001 AND F.TF002 = G.TG002 "
            sql += f"JOIN OtsukaDB.dbo.CMSMC C ON G.TG010 = C.MC001 "
            sql += f"WHERE " 
            sql += f"F.TF003 LIKE '{q_e_p_i_y_month}%' "
            sql += f"AND F.TF001 IN ('580', '583', '584', '585', '590') "
            sql += f"AND G.TG018 > CONVERT(VARCHAR, GETDATE(), 112) "
            sql += f"AND LEFT(F.TF003, 6) BETWEEN FORMAT(DATEADD(MONTH, -3, GETDATE()), 'yyyyMM') AND FORMAT(DATEADD(MONTH, 0, GETDATE()), 'yyyyMM') "
            sql += f"ORDER BY F.TF003, G.TG003"
                
            self.curr_mssql_erp.execute(sql)
            res = self.curr_mssql_erp.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_erp_realtime_query_search2 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()
    
    #####################################################
    # show_login_out_record_query_operation_search
    #####################################################
    def show_login_out_record_query_operation_search(self , l_o_a_query , q_s_date , q_e_date):
        
        self.__connect__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            sql  = f"select item  , r_time from operation_record where a_user='{l_o_a_query}' and "
            sql += f"r_time >= '{q_s_date}%' and r_time <= '{q_e_date}%' order by r_time desc"
                
            self.curr.execute(sql)
            res = self.curr.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_login_out_record_query_operation_search : {str(e)}\n")

        finally:
            self.__disconnect__()

    #####################################################
    # show_login_out_record_query_search
    #####################################################
    def show_login_out_record_query_search(self , l_o_a_query , q_s_date , q_e_date):
        
        self.__connect__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            sql  = f"select login_ip , login_time , logout_time from login_out_record where a_user='{l_o_a_query}' and "
            sql += f"login_time >= '{q_s_date}%' and logout_time <= '{q_e_date}%' order by login_time desc"
                
            self.curr.execute(sql)
            res = self.curr.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_login_out_record_query_search : {str(e)}\n")

        finally:
            self.__disconnect__()

    #####################################################
    # submit_hr_360_content_data_member
    #####################################################
    def submit_hr_360_content_data_member(self , check_hr_360_date , check_hr_360_user , check_hr_360_lv , check_hr_360_name , hr_360_total_1_1 , hr_360_total_1_2 , hr_360_total_1_3 , hr_360_total_1_4 , hr_360_total_1_5 ,hr_360_total_2_1 , hr_360_total_2_2 , hr_360_total_2_3 , hr_360_total_2_4 , hr_360_total_2_5 , hr_360_total_3_1 , hr_360_total_3_2 , hr_360_total_3_3 , hr_360_total_3_4 , hr_360_total_3_5 ,hr_360_total_4_1 , hr_360_total_4_2 , hr_360_total_4_3 , hr_360_total_4_4 , hr_360_total_4_5 , hr_360_total_5_1 , hr_360_total_5_2 , hr_360_total_5_3 , hr_360_total_5_4 , hr_360_total_5_5 , hr_360_total_5_6 ,total_1 , total_1_avg ,total_2 , total_2_avg ,total_3 , total_3_avg ,total_4 , total_4_avg ,total_5 , total_5_avg):
        
        # login_id -> chinese name
        #c_name = web_cloud_dao.bpm_account_data(check_hr_360_user , 'EmployeeName')
        
        self.__connect7_1_38__()
        self.__connect_mssql__()

        try:

            sql        = f"SELECT employeename FROM T_HR_Employee the where loginID='{check_hr_360_user}'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchone()
            c_name     = res[0]
            
            sql  = f"SELECT s_user , s_name from hr_360_submit_member_content "
            sql += f"where s_user='{check_hr_360_user}' and s_name='{check_hr_360_name}'"
                
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchone()
            r_year  = check_hr_360_date.split('-')[0]
            r_month = check_hr_360_date.split('-')[1]
            r_day   = check_hr_360_date.split('-')[2]

            if res is None:

                ###############
                #
                # 新增紀錄
                #
                ###############
                sql2  = f"insert into hr_360_submit_member_content "
                sql2 += f"("
                sql2 += f"c_year , c_month , c_day , "
                sql2 += f"s_date , s_user , s_lv , s_name , "
                sql2 += f"s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , " 
                sql2 += f"s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , " 
                sql2 += f"s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , " 
                sql2 += f"s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , s_hr_360_4_5 , " 
                sql2 += f"s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , s_hr_360_5_4 , s_hr_360_5_5 , s_hr_360_5_6 , " 
                sql2 += f"s_hr_360_total_1 , s_hr_360_total_1_avg , " 
                sql2 += f"s_hr_360_total_2 , s_hr_360_total_2_avg , " 
                sql2 += f"s_hr_360_total_3 , s_hr_360_total_3_avg , " 
                sql2 += f"s_hr_360_total_4 , s_hr_360_total_4_avg , " 
                sql2 += f"s_hr_360_total_5 , s_hr_360_total_5_avg" 
                sql2 += f") " 
                sql2 += f"values("
                sql2 += f"'{r_year}' , '{r_month}' , '{r_day}' , "
                sql2 += f"'{check_hr_360_date}' , '{c_name}' , '{check_hr_360_lv}' , '{check_hr_360_name}' , "
                sql2 += f"'{hr_360_total_1_1}' , '{hr_360_total_1_2}' , '{hr_360_total_1_3}' , '{hr_360_total_1_4}' , '{hr_360_total_1_5}' , "
                sql2 += f"'{hr_360_total_2_1}' , '{hr_360_total_2_2}' , '{hr_360_total_2_3}' , '{hr_360_total_2_4}' , '{hr_360_total_2_5}' , "
                sql2 += f"'{hr_360_total_3_1}' , '{hr_360_total_3_2}' , '{hr_360_total_3_3}' , '{hr_360_total_3_4}' , '{hr_360_total_3_5}' , "
                sql2 += f"'{hr_360_total_4_1}' , '{hr_360_total_4_2}' , '{hr_360_total_4_3}' , '{hr_360_total_4_4}' , '{hr_360_total_4_5}' , "
                sql2 += f"'{hr_360_total_5_1}' , '{hr_360_total_5_2}' , '{hr_360_total_5_3}' , '{hr_360_total_5_4}' , '{hr_360_total_5_5}' , '{hr_360_total_5_6}' , "
                sql2 += f"'{total_1}' , '{total_1_avg}' , "
                sql2 += f"'{total_2}' , '{total_2_avg}' , "
                sql2 += f"'{total_3}' , '{total_3_avg}' , "
                sql2 += f"'{total_4}' , '{total_4_avg}' , "
                sql2 += f"'{total_5}' , '{total_5_avg}'"
                sql2 += f") "

                print(f"\n {sql2}")

                ###############
                #
                # 更新紀錄
                #
                ###############
                sql3 = (
                        f"UPDATE `hr_360_submit_manager_content_person` SET `s_status` = 'done' WHERE s_user=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(sql3 , (c_name , check_hr_360_name,))

                if self.curr_7_1_38.execute(sql2):
                    res = 'ok'
                    return res

            else:
                res = 'no'
                return res

        except Exception as e:
            logging.error(f"\n<Error> submit_hr_360_content_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
            self.__disconnect_mssql__()

    #####################################################
    # submit_hr_360_content_data
    #####################################################
    def submit_hr_360_content_data(self , check_hr_360_date , check_hr_360_user , check_hr_360_lv , check_hr_360_name , hr_360_total_1_1 , hr_360_total_1_2 , hr_360_total_1_3 , hr_360_total_1_4 , hr_360_total_1_5 ,hr_360_total_2_1 , hr_360_total_2_2 , hr_360_total_2_3 , hr_360_total_2_4 , hr_360_total_2_5 , hr_360_total_2_6 ,hr_360_total_3_1 , hr_360_total_3_2 , hr_360_total_3_3 , hr_360_total_3_4 , hr_360_total_3_5 ,hr_360_total_4_1 , hr_360_total_4_2 , hr_360_total_4_3 , hr_360_total_4_4 , hr_360_total_5_1 , hr_360_total_5_2 , hr_360_total_5_3 ,hr_360_total_6_1 , hr_360_total_6_2 , hr_360_total_6_3 , hr_360_total_6_4 , hr_360_total_6_5 ,total_1 , total_1_avg , total_2 , total_2_avg , total_3 , total_3_avg , total_4 , total_4_avg , total_5 , total_5_avg , total_6 , total_6_avg):
        
        # login_id -> chinese name
        #c_name = web_cloud_dao.bpm_account_data(check_hr_360_user , 'EmployeeName')
        
        self.__connect7_1_38__()
        self.__connect_mssql__()

        try:

            sql        = f"SELECT employeename FROM T_HR_Employee the where loginID='{check_hr_360_user}'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchone()
            c_name     = res[0]
            
            sql  = f"SELECT s_user , s_name from hr_360_submit_manager_content "
            sql += f"where s_user='{check_hr_360_user}' and s_name='{check_hr_360_name}'"
                
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchone()
            r_year  = check_hr_360_date.split('-')[0]
            r_month = check_hr_360_date.split('-')[1]
            r_day   = check_hr_360_date.split('-')[2]

            if res is None:

                ###############
                #
                # 新增紀錄
                #
                ###############
                sql2  = f"insert into hr_360_submit_manager_content "
                sql2 += f"("
                sql2 += f"c_year , c_month , c_day , "
                sql2 += f"s_date , s_user , s_lv , s_name , "
                sql2 += f"s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , " 
                sql2 += f"s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , s_hr_360_2_6 , " 
                sql2 += f"s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , " 
                sql2 += f"s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , " 
                sql2 += f"s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , " 
                sql2 += f"s_hr_360_6_1 , s_hr_360_6_2 , s_hr_360_6_3 , s_hr_360_6_4 , s_hr_360_6_5 , " 
                sql2 += f"s_hr_360_total_1 , s_hr_360_total_1_avg , " 
                sql2 += f"s_hr_360_total_2 , s_hr_360_total_2_avg , " 
                sql2 += f"s_hr_360_total_3 , s_hr_360_total_3_avg , " 
                sql2 += f"s_hr_360_total_4 , s_hr_360_total_4_avg , " 
                sql2 += f"s_hr_360_total_5 , s_hr_360_total_5_avg , " 
                sql2 += f"s_hr_360_total_6 , s_hr_360_total_6_avg" 
                sql2 += f") " 
                sql2 += f"values("
                sql2 += f"'{r_year}' , '{r_month}' , '{r_day}' , "
                sql2 += f"'{check_hr_360_date}' , '{c_name}' , '{check_hr_360_lv}' , '{check_hr_360_name}' , "
                sql2 += f"'{hr_360_total_1_1}' , '{hr_360_total_1_2}' , '{hr_360_total_1_3}' , '{hr_360_total_1_4}' , '{hr_360_total_1_5}' , "
                sql2 += f"'{hr_360_total_2_1}' , '{hr_360_total_2_2}' , '{hr_360_total_2_3}' , '{hr_360_total_2_4}' , '{hr_360_total_2_5}' , '{hr_360_total_2_6}' , "
                sql2 += f"'{hr_360_total_3_1}' , '{hr_360_total_3_2}' , '{hr_360_total_3_3}' , '{hr_360_total_3_4}' , '{hr_360_total_3_5}' , "
                sql2 += f"'{hr_360_total_4_1}' , '{hr_360_total_4_2}' , '{hr_360_total_4_3}' , '{hr_360_total_4_4}' , "
                sql2 += f"'{hr_360_total_5_1}' , '{hr_360_total_5_2}' , '{hr_360_total_5_3}' , "
                sql2 += f"'{hr_360_total_6_1}' , '{hr_360_total_6_2}' , '{hr_360_total_6_3}' , '{hr_360_total_6_4}' , '{hr_360_total_6_5}' , "
                sql2 += f"'{total_1}' , '{total_1_avg}' , "
                sql2 += f"'{total_2}' , '{total_2_avg}' , "
                sql2 += f"'{total_3}' , '{total_3_avg}' , "
                sql2 += f"'{total_4}' , '{total_4_avg}' , "
                sql2 += f"'{total_5}' , '{total_5_avg}' , "
                sql2 += f"'{total_6}' , '{total_6_avg}'"
                sql2 += f") "


                ###############
                #
                # 更新紀錄
                #
                ###############
                sql3 = (
                        f"UPDATE `hr_360_submit_manager_content_person` SET `s_status` = 'done' WHERE s_user=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(sql3 , (c_name , check_hr_360_name,))

                if self.curr_7_1_38.execute(sql2):
                    res = 'ok'
                    return res

            else:
                res = 'no'
                return res

        except Exception as e:
            logging.error(f"\n<Error> submit_hr_360_content_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
            self.__disconnect_mssql__()

    #####################################################
    # alter_hr_360_person_data
    #####################################################
    def alter_hr_360_person_data(self , d_dep , d_name):
        
        self.__connect7_1_38__()

        try:
            
            sql3  = f"SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup "
            sql3 += f"where c_dep='{d_dep}' and c_name='{d_name}'"
                
            self.curr_7_1_38.execute(sql3)
            res3 = self.curr_7_1_38.fetchall()

            return res3

        except Exception as e:
            logging.error(f"\n<Error> alter_hr_360_person_1_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()


    #####################################################
    # search_hr_360_person_name
    #####################################################
    def search_hr_360_person_name(self , s_name , item):
        
        self.__connect7_1_38__()

        try:
            
            sql = (
                    f"select {item} from hr_360_person_setup where c_name='{s_name}'"
            )
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchone()

            if res is not None:
                return res[0]

        except Exception as e:
            logging.error(f"\n<Error> search_hr_360_person_name : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_process_total_data3_2
    #####################################################
    def show_hr_360_person_process_total_data3_2(self , s_user , s_name , item , range , item2):
        
        self.__connect7_1_38__()

        try:
            
            if s_user == s_name:
                
                sql = (f"select {item} from hr_360_submit_member_content where s_user='{s_user}' and s_name='{s_name}'")
                
                self.curr_7_1_38.execute(sql)
                res = self.curr_7_1_38.fetchone()
                
                if res is None:

                    print(f"No data returned from the database query.")
                    pic = None  # Or set to a default value or error image

                else:
                    ###################
                    # transfer to pie pic
                    ########################
                    if item2 == 'none':
                        pic = self.show_pie_picture(range , res[0])
                    else:
                        pic = res[0]

                    return pic
            
            else:
                
                sql = (f"select {item} from hr_360_submit_member_content where s_user='{s_name}' and s_name='{s_user}'")
                
                self.curr_7_1_38.execute(sql)
                res = self.curr_7_1_38.fetchone()

                if res is None:
                    
                    print(f"No data returned from the database query.")
                    pic = None  # Or set to a default value or error image

                else:
                    ########################
                    # transfer to pie pic
                    ########################
                    pic = self.show_pie_picture(range , res[0])

                    if item2 == 'none':
                        pic = self.show_pie_picture(range , res[0])
                    else:
                        pic = res[0]

                    return pic

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_process_total_data3_2 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
    
    #####################################################
    # show_hr_360_person_process_total_data3
    #####################################################
    def show_hr_360_person_process_total_data3(self , s_user , s_name , item , range , item2):
        
        self.__connect7_1_38__()

        try:
            
            if s_user == s_name:
                
                sql = (f"select {item} from hr_360_submit_manager_content where s_user='{s_user}' and s_name='{s_name}'")
                
                self.curr_7_1_38.execute(sql)
                res = self.curr_7_1_38.fetchone()
                
                if res is None:

                    print(f"No data returned from the database query.")
                    pic = None  # Or set to a default value or error image

                else:
                    ###################
                    # transfer to pie pic
                    ########################
                    if item2 == 'none':
                        pic = self.show_pie_picture(range , res[0])
                    else:
                        pic = res[0]

                    return pic
            
            else:
                
                sql = (f"select {item} from hr_360_submit_manager_content where s_user='{s_name}' and s_name='{s_user}'")
                
                self.curr_7_1_38.execute(sql)
                res = self.curr_7_1_38.fetchone()

                if res is None:
                    
                    print(f"No data returned from the database query.")
                    pic = None  # Or set to a default value or error image

                else:
                    ########################
                    # transfer to pie pic
                    ########################
                    pic = self.show_pie_picture(range , res[0])

                    if item2 == 'none':
                        pic = self.show_pie_picture(range , res[0])
                    else:
                        pic = res[0]

                    return pic

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_process_total_data3 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_process_total_data2_2
    #####################################################
    def show_hr_360_person_process_total_data2_2(self , s_name):
        
        self.__connect7_1_38__()

        try:
            
            sql = (
                    "select "
                    "s_hr_360_total_1 , s_hr_360_total_1_avg , "
                    "s_hr_360_total_2 , s_hr_360_total_2_avg , "
                    "s_hr_360_total_3 , s_hr_360_total_3_avg , "
                    "s_hr_360_total_4 , s_hr_360_total_4_avg , "
                    "s_hr_360_total_5 , s_hr_360_total_5_avg , "
                    "s_date , s_user , s_name , "
                    "s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , "
                    "s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , "
                    "s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , "
                    "s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , s_hr_360_4_5 , "
                    "s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , s_hr_360_5_4 , s_hr_360_5_5 , s_hr_360_5_6 , s_lv "
                    "from hr_360_submit_member_content where s_name=%s"
            )
            self.curr_7_1_38.execute(sql , (s_name ,))
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_process_total_data2_2 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
    
    #####################################################
    # show_hr_360_person_process_total_data2
    #####################################################
    def show_hr_360_person_process_total_data2(self , s_name):
        
        self.__connect7_1_38__()

        try:
            
            sql = (
                    "select "
                    "s_hr_360_total_1 , s_hr_360_total_1_avg , "
                    "s_hr_360_total_2 , s_hr_360_total_2_avg , "
                    "s_hr_360_total_3 , s_hr_360_total_3_avg , "
                    "s_hr_360_total_4 , s_hr_360_total_4_avg , "
                    "s_hr_360_total_5 , s_hr_360_total_5_avg , "
                    "s_hr_360_total_6 , s_hr_360_total_6_avg , "
                    "s_date , s_user , s_name , "
                    "s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , "
                    "s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , s_hr_360_2_6 , "
                    "s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , "
                    "s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , "
                    "s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , "
                    "s_hr_360_6_1 , s_hr_360_6_2 , s_hr_360_6_3 , s_hr_360_6_4 , s_hr_360_6_5 , s_lv "
                    "from hr_360_submit_manager_content where s_name=%s"
            )
            self.curr_7_1_38.execute(sql , (s_name ,))
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_process_total_data2 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
    
    #####################################################
    # show_hr_360_person_process_total_data
    #####################################################
    def show_hr_360_person_process_total_data(self , s_user , s_name):
        
        self.__connect7_1_38__()

        try:
            
            sql = (
                    "select "
                    "s_hr_360_total_1 , s_hr_360_total_1_avg , "
                    "s_hr_360_total_2 , s_hr_360_total_2_avg , "
                    "s_hr_360_total_3 , s_hr_360_total_3_avg , "
                    "s_hr_360_total_4 , s_hr_360_total_4_avg , "
                    "s_hr_360_total_5 , s_hr_360_total_5_avg , "
                    "s_hr_360_total_6 , s_hr_360_total_6_avg , "
                    "s_date , s_user , s_name , "
                    "s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , "
                    "s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , s_hr_360_2_6 , "
                    "s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , "
                    "s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , "
                    "s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , "
                    "s_hr_360_6_1 , s_hr_360_6_2 , s_hr_360_6_3 , s_hr_360_6_4 , s_hr_360_6_5 "
                    "from hr_360_submit_manager_content where s_name=%s order by s_lv asc"
            )
            self.curr_7_1_38.execute(sql , (s_name ,))
            res = self.curr_7_1_38.fetchall()

            return res
            
            

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_process_total_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_process_data
    #####################################################
    def show_hr_360_person_process_data(self , s_user , s_name):
        
        self.__connect7_1_38__()

        try:
            
            sql = (
                    "select "
                    "s_hr_360_total_1 , s_hr_360_total_1_avg , "
                    "s_hr_360_total_2 , s_hr_360_total_2_avg , "
                    "s_hr_360_total_3 , s_hr_360_total_3_avg , "
                    "s_hr_360_total_4 , s_hr_360_total_4_avg , "
                    "s_hr_360_total_5 , s_hr_360_total_5_avg , "
                    "s_hr_360_total_6 , s_hr_360_total_6_avg , "
                    "s_date , s_user , s_name , "
                    "s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , "
                    "s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , s_hr_360_2_6 , "
                    "s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , "
                    "s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , "
                    "s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , "
                    "s_hr_360_6_1 , s_hr_360_6_2 , s_hr_360_6_3 , s_hr_360_6_4 , s_hr_360_6_5 "
                    "from hr_360_submit_manager_content where s_user=%s and s_name=%s"
            )
            self.curr_7_1_38.execute(sql , (s_user , s_name ,))
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_process_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # search_hr_360_person_lv_member
    #####################################################
    def search_hr_360_person_lv_member(self , s_user , s_name):
        
        c_s_user = self.search_bpm_employee_id(s_user)
        
        self.__connect7_1_38__()

        try:

            sql = (
                    f"SELECT c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 , c_name from hr_360_person_setup where c_name=%s and lv='member'"
            )
                
            self.curr_7_1_38.execute(sql , (s_name,))
            res = self.curr_7_1_38.fetchall()

            for val in res:
                if val[0] == c_s_user:
                    lv = '主管' 
                    return lv
                
                elif val[1] == c_s_user:
                    lv = '同輩1'
                    return lv

                elif val[2] == c_s_user:
                    lv = '同輩2'
                    return lv
                
                elif val[3] == c_s_user:
                    lv = '下屬1'
                    return lv
                
                elif val[4] == c_s_user:
                    lv = '下屬2'
                    return lv

                elif val[5] == c_s_user:
                    lv = '自己'
                    return lv


        except Exception as e:
            logging.error(f"\n<Error> search_hr_360_person_lv : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # search_hr_360_person_lv
    #####################################################
    def search_hr_360_person_lv(self , s_user , s_name):
        
        c_s_user = self.search_bpm_employee_id(s_user)
        
        self.__connect7_1_38__()

        try:

            sql = (
                    f"SELECT c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 , c_name from hr_360_person_setup where c_name=%s and lv='manager'"
            )
                
            self.curr_7_1_38.execute(sql , (s_name,))
            res = self.curr_7_1_38.fetchall()

            for val in res:
                if val[0] == c_s_user:
                    lv = '主管' 
                    return lv
                
                elif val[1] == c_s_user:
                    lv = '同輩1'
                    return lv

                elif val[2] == c_s_user:
                    lv = '同輩2'
                    return lv
                
                elif val[3] == c_s_user:
                    lv = '下屬1'
                    return lv
                
                elif val[4] == c_s_user:
                    lv = '下屬2'
                    return lv

                elif val[5] == c_s_user:
                    lv = '自己'
                    return lv


        except Exception as e:
            logging.error(f"\n<Error> search_hr_360_person_lv : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # del_hr_360_person_process_data
    #####################################################
    def del_hr_360_person_process_data(self , s_user , s_name):
        
        self.__connect7_1_38__()

        try:
            
            #######################################
            # tb : hr_360_submit_manager_content
            #######################################
            sql = (
                    "delete from hr_360_submit_manager_content where s_user=%s and s_name=%s"
            )
            self.curr_7_1_38.execute(sql , (s_user , s_name ,))

            ##############################################
            # tb : hr_360_submit_manager_content_person
            ##############################################
            sql2 = (
                    "update hr_360_submit_manager_content_person set s_status=NULL where s_user=%s and s_name=%s"
            )
            self.curr_7_1_38.execute(sql2 , (s_user , s_name ,))
            
            #######################################
            # tb : hr_360_submit_manager_content
            #######################################
            sql3 = (
                    "SELECT s_date , s_user , s_name from hr_360_submit_manager_content order by s_user desc"
            )
                
            self.curr_7_1_38.execute(sql3)
            res3 = self.curr_7_1_38.fetchall()

            return res3

        except Exception as e:
            logging.error(f"\n<Error> del_hr_360_person_process_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # del_hr_360_person_data
    #####################################################
    def del_hr_360_person_data(self , d_dep , d_name):
        
        self.__connect7_1_38__()

        try:
            
            sql = (f"delete from hr_360_person_setup where c_dep=%s and c_name=%s")
            self.curr_7_1_38.execute(sql , (d_dep , d_name,))

            d_sql = (f"delete from hr_360_submit_manager_content_person where s_dep=%s and s_name=%s")
            self.curr_7_1_38.execute(d_sql, (d_dep , d_name,))
            
            sql3 = (f"SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup order by c_dep")
                
            self.curr_7_1_38.execute(sql3)
            res3 = self.curr_7_1_38.fetchall()

            return res3

        except Exception as e:
            logging.error(f"\n<Error> del_hr_360_person_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_1_check_data2
    #####################################################
    def show_hr_360_person_1_check_data2(self):
        
        self.__connect7_1_38__()

        try:
                
            sql2 = (
                    f"select s_user , s_lv , s_name from hr_360_submit_manager_content_person where s_status is Null"
            )

            self.curr_7_1_38.execute(sql2)
            res2 = self.curr_7_1_38.fetchall()

            f_res = []

            for val in res2:
                
                email = self.search_bpm_account_item_list(str(val[0]))
                f_res.append((val[0] , val[1] , val[2] , email,))

            return f_res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_check_data2 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_1_check_data2_2
    #####################################################
    def show_hr_360_person_1_check_data2_2(self):
        
        self.__connect7_1_38__()

        try:
            sql = (
                    "select s_date , s_user , s_name "
                    "from hr_360_submit_manager_content order by s_user desc"
            )
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_check_data2_2 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_1_check_data2_3
    #####################################################
    def show_hr_360_person_1_check_data2_3(self):
        
        self.__connect7_1_38__()

        try:
            sql = (
                    "select s_date , s_user , s_name "
                    "from hr_360_submit_manager_content order by s_user desc"
            )
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_check_data2_3 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_1_check_data
    #####################################################
    def show_hr_360_person_1_check_data(self):
        
        self.__connect7_1_38__()

        try:
            sql = (
                    "select s_date , s_user , s_name "
                    "from hr_360_submit_manager_content order by s_user desc"
            )
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_check_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_1_data2
    #####################################################
    def show_hr_360_person_1_data2(self):
        
        self.__connect7_1_38__()

        try:
            sql = ("SELECT c_name from hr_360_person_setup group by c_name")
                
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            f_res = []

            for val in res:

                sql2 = (
                        f"select c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup where c_name=%s"
                )
                self.curr_7_1_38.execute(sql2 , (val[0],))
                res2 = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_hr_360_person_1_data_2
    #####################################################
    def show_hr_360_person_1_data_2(self):
        
        self.__connect7_1_38__()

        try:
            sql = (
                    "SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup where lv='member' order by c_dep"
            )
                
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_data_2 : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
    
    #####################################################
    # show_hr_360_person_1_data
    #####################################################
    def show_hr_360_person_1_data(self):
        
        self.__connect7_1_38__()

        try:
            sql = (
                    "SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup  where lv='manager' order by c_dep"
            )
                
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_hr_360_person_1_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
            

    #####################################################
    # add_hr_360_setup_data
    #####################################################
    def add_hr_360_setup_data(self , h_360_m_dep , h_360_m_name , h_360_m_lv , h_360_c_m_name , h_360_c_p_1_name , h_360_c_p_2_name , h_360_c_s_1_name , h_360_c_s_2_name):
        
        self.__connect7_1_38__()
        
        try:
            ### variable
            r_date   = time.strftime("%Y-%m-%d" , time.localtime())
            r_time   = time.strftime("%H:%M:%S" , time.localtime())
            r_d_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            r_year   = time.strftime("%Y" , time.localtime())
            r_month  = time.strftime("%m" , time.localtime())
            r_day    = time.strftime("%d" , time.localtime())

            sql  = f"SELECT * from hr_360_person_setup where "
            sql += f"c_dep = '{h_360_m_dep}' and "
            sql += f"c_name = '{h_360_m_name}'"
                
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchone()

            if res is None:
                
                ######################################
                #
                # 主管清單 - add hr_360_person_setup
                #
                ######################################
                sql2  = f"insert into hr_360_person_setup ( " 
                sql2 += f"c_date , c_time , c_d_time , c_year , c_month , c_day , c_dep , "
                sql2 += f"lv , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 ) "
                sql2 += f"value( "
                sql2 += f"'{r_date}','{r_time}','{r_d_time}','{r_year}','{r_month}','{r_day}', "
                sql2 += f"'{h_360_m_dep}','{h_360_m_lv}','{h_360_m_name}','{h_360_c_m_name}','{h_360_c_p_1_name}','{h_360_c_p_2_name}','{h_360_c_s_1_name}','{h_360_c_s_2_name}' "
                sql2 += f" )"

                self.curr_7_1_38.execute(sql2)

                ####################################################
                #
                # 未考評清單 - hr_360_submit_manager_content_person
                # (自己 , 主管 , 同輩1 , 同輩2 , 下屬1 , 下屬2)
                #
                ####################################################
                
                ########
                # 自己
                ########
                c_sql3_1 = (
                            f"select * from hr_360_submit_manager_content_person where "
                            "s_dep=%s  and s_user=%s and s_lv=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(c_sql3_1 , (h_360_m_dep , h_360_m_name , '自己' , h_360_m_name , ))
                c_res3_1 = self.curr_7_1_38.fetchone()

                if c_res3_1 is None:
                    sql3_1 = (
                            f"insert into hr_360_submit_manager_content_person (c_date , s_dep , s_user , s_lv , s_name) value( %s , %s , %s , %s , %s )"
                    )
                    self.curr_7_1_38.execute(sql3_1 , (r_date , h_360_m_dep , h_360_m_name , '自己' , h_360_m_name , ))
                
                #########
                # 主管
                #########
                c_sql3_2 = (
                            f"select * from hr_360_submit_manager_content_person where "
                            "s_dep=%s  and s_user=%s and s_lv=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(c_sql3_2 , (h_360_m_dep , h_360_c_m_name , '主管' , h_360_m_name , ))
                c_res3_2 = self.curr_7_1_38.fetchone()

                if c_res3_2 is None:
                    sql3_2 = (
                            f"insert into hr_360_submit_manager_content_person (c_date , s_dep , s_user , s_lv , s_name) value( %s , %s , %s , %s , %s )"
                    )
                    self.curr_7_1_38.execute(sql3_2 , (r_date , h_360_m_dep , h_360_c_m_name , '主管' , h_360_m_name , ))

                #########
                # 同輩1
                #########
                c_sql3_3 = (
                            f"select * from hr_360_submit_manager_content_person where "
                            "s_dep=%s  and s_user=%s and s_lv=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(c_sql3_3 , (h_360_m_dep , h_360_c_p_1_name , '同輩1' , h_360_m_name , ))
                c_res3_3 = self.curr_7_1_38.fetchone()

                if c_res3_3 is None:
                    sql3_3 = (
                            f"insert into hr_360_submit_manager_content_person (c_date , s_dep , s_user , s_lv , s_name) value( %s , %s , %s , %s , %s )"
                    )
                    self.curr_7_1_38.execute(sql3_3 , (r_date , h_360_m_dep , h_360_c_p_1_name , '同輩1' , h_360_m_name , ))
                
                #########
                # 同輩2
                #########
                c_sql3_4 = (
                            f"select * from hr_360_submit_manager_content_person where "
                            "s_dep=%s  and s_user=%s and s_lv=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(c_sql3_4 , (h_360_m_dep , h_360_c_p_2_name , '同輩2' , h_360_m_name , ))
                c_res3_4 = self.curr_7_1_38.fetchone()

                if c_res3_4 is None:
                    sql3_4 = (
                            f"insert into hr_360_submit_manager_content_person (c_date , s_dep , s_user , s_lv , s_name) value( %s , %s , %s , %s , %s )"
                    )
                    self.curr_7_1_38.execute(sql3_4 , (r_date , h_360_m_dep , h_360_c_p_2_name , '同輩2' , h_360_m_name , ))

                #########
                # 下屬1
                #########
                c_sql3_5 = (
                            f"select * from hr_360_submit_manager_content_person where "
                            "s_dep=%s  and s_user=%s and s_lv=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(c_sql3_5 , (h_360_m_dep , h_360_c_s_1_name , '下屬1' , h_360_m_name , ))
                c_res3_5 = self.curr_7_1_38.fetchone()

                if c_res3_5 is None:
                    sql3_5 = (
                            f"insert into hr_360_submit_manager_content_person (c_date , s_dep , s_user , s_lv , s_name) value( %s , %s , %s , %s , %s )"
                    )
                    self.curr_7_1_38.execute(sql3_5 , (r_date , h_360_m_dep , h_360_c_s_1_name , '下屬1' , h_360_m_name , ))

                
                #########
                # 下屬2
                #########
                c_sql3_6 = (
                            f"select * from hr_360_submit_manager_content_person where "
                            "s_dep=%s  and s_user=%s and s_lv=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(c_sql3_6 , (h_360_m_dep , h_360_c_s_2_name , '下屬2' , h_360_m_name , ))
                c_res3_6 = self.curr_7_1_38.fetchone()

                if c_res3_6 is None:
                    sql3_6 = (
                            f"insert into hr_360_submit_manager_content_person (c_date , s_dep , s_user , s_lv , s_name) value( %s , %s , %s , %s , %s )"
                    )
                    self.curr_7_1_38.execute(sql3_6 , (r_date , h_360_m_dep , h_360_c_s_2_name , '下屬2' , h_360_m_name , ))
                
                
                if str(h_360_m_lv) == 'member':
                    sql3 = f"SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup where lv='member' order by c_dep"
                    self.curr_7_1_38.execute(sql3)
                    res3 = self.curr_7_1_38.fetchall()

                    return res3

                else:
                    sql3 = f"SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup where lv='manager'  order by c_dep"
                    self.curr_7_1_38.execute(sql3)
                    res3 = self.curr_7_1_38.fetchall()

                    return res3
            
            else:

                ### variables
                r_d_update_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

                sql2  = f"update hr_360_person_setup set " 
                sql2 += f"c_d_time='{r_d_update_time}' , "
                sql2 += f"c_manager='{h_360_c_m_name}' , " 
                sql2 += f"c_peer1='{h_360_c_p_1_name}' , " 
                sql2 += f"c_peer2='{h_360_c_p_2_name}' , "
                sql2 += f"c_subordinate1='{h_360_c_s_1_name}' , " 
                sql2 += f"c_subordinate2='{h_360_c_s_2_name}' "
                sql2 += f"where c_dep='{h_360_m_dep}' and c_name='{h_360_m_name}'"

                self.curr_7_1_38.execute(sql2)


                ####################################################
                #
                # 未考評清單 - hr_360_submit_manager_content_person
                # (自己 , 主管 , 同輩1 , 同輩2 , 下屬1 , 下屬2)
                #
                ####################################################
                
                ########
                # 自己
                ########
                u_sql3_1 = (
                            f"update hr_360_submit_manager_content_person set "
                            "s_user=%s where s_lv=%s and s_dep=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(u_sql3_1 , (h_360_m_name , '自己' , h_360_m_dep , h_360_m_name , ))
                
                ########
                # 主管
                ########
                u_sql3_2 = (
                            f"update hr_360_submit_manager_content_person set "
                            "s_user=%s where s_lv=%s and s_dep=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(u_sql3_2 , (h_360_c_m_name , '主管' , h_360_m_dep , h_360_m_name , ))

                ########
                # 同輩1
                ########
                u_sql3_3 = (
                            f"update hr_360_submit_manager_content_person set "
                            "s_user=%s where s_lv=%s and s_dep=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(u_sql3_3 , (h_360_c_p_1_name , '同輩1' , h_360_m_dep , h_360_m_name , ))

                ########
                # 同輩2
                ########
                u_sql3_4 = (
                            f"update hr_360_submit_manager_content_person set "
                            "s_user=%s where s_lv=%s and s_dep=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(u_sql3_4 , (h_360_c_p_2_name , '同輩2' , h_360_m_dep , h_360_m_name , ))

                ########
                # 下屬1
                ########
                u_sql3_5 = (
                            f"update hr_360_submit_manager_content_person set "
                            "s_user=%s where s_lv=%s and s_dep=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(u_sql3_5 , (h_360_c_s_1_name , '下屬1' , h_360_m_dep , h_360_m_name , ))

                ########
                # 下屬2
                ########
                u_sql3_6 = (
                            f"update hr_360_submit_manager_content_person set "
                            "s_user=%s where s_lv=%s and s_dep=%s and s_name=%s"
                )
                self.curr_7_1_38.execute(u_sql3_6 , (h_360_c_s_2_name , '下屬2' , h_360_m_dep , h_360_m_name , ))

                if str(h_360_m_lv) == 'member':
                    sql3 = f"SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup where lv='member' order by c_dep"
                    self.curr_7_1_38.execute(sql3)
                    res3 = self.curr_7_1_38.fetchall()

                    return res3

                else:
                    sql3 = f"SELECT c_dep , c_name , c_manager , c_peer1 , c_peer2 , c_subordinate1 , c_subordinate2 from hr_360_person_setup where lv='manager' order by c_dep"
                    self.curr_7_1_38.execute(sql3)
                    res3 = self.curr_7_1_38.fetchall()
                    
                    return res3
        
        except Exception as e:
            logging.error(f"\n<Error> add_hr_360_setup_data : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # show_erp_realtime_query_search
    #####################################################
    def show_erp_realtime_query_search(self , q_e_p_num , q_e_p_name , q_e_p_a_num , q_e_p_l_date):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())


            sql  = f"SELECT "
            sql += f"D.MC002 AS '庫別名稱', "
            sql += f"A.MF007 AS '庫別', "
            sql += f"FORMAT(SUM(A.MF010 * A.MF008) , 'N0' , 'en-US') AS '總數量' "
            sql += f"FROM " 
            sql += f"OtsukaDB.dbo.INVMF A "
            sql += f"JOIN OtsukaDB.dbo.CMSMC D ON A.MF007 = D.MC001 "
            sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
            sql += f"WHERE "
            sql += f"A.MF003 >= '20190101' "
            sql += f"AND A.MF001 LIKE 'X%' "
            sql += f"AND EXISTS ( "
            sql += f"SELECT 1 "
            sql += f"FROM OtsukaDB.dbo.INVME E "
            sql += f"WHERE A.MF001 = E.ME001 "
            sql += f"AND A.MF002 = E.ME002 "
            sql += f"AND E.ME009 >= '20240101' "
            sql += f"AND E.ME003 >= '20200101' "
            sql += f"AND C.MB001 = '{q_e_p_num}' "
            sql += f"AND C.MB002 = '{q_e_p_name}' "
            sql += f"AND A.MF002 = '{q_e_p_a_num}' "
            sql += f"AND E.ME009 = '{q_e_p_l_date}' "
            sql += f") "
            sql += f"GROUP BY  C.MB001, C.MB002,A.MF007,D.MC002 "
            sql += f"ORDER BY C.MB001 , D.MC002"
                
            self.curr_mssql_erp.execute(sql)
            res = self.curr_mssql_erp.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_erp_realtime_query_search : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #####################################################
    # show_bpm_expenditure_form_search
    #####################################################
    def show_bpm_expenditure_form_search(self , q_s_date , q_e_date , q_b_e_dep , q_b_e_d_member , q_b_e_status , q_b_b_s_b_budget):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            q_s_date = str(q_s_date).replace("-","/")
            q_e_date = str(q_e_date).replace("-","/")

            ## chart 
            if q_b_e_status == '已結案':
                if q_b_e_d_member == '':
                    if q_b_b_s_b_budget == '':
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                    else:
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' and ITEM24='{q_b_b_s_b_budget}' order by ITEM12 desc"
                else: 
                    if q_b_b_s_b_budget == '': 
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                    else:
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
            elif q_b_e_status == '未結案':
                if q_b_e_d_member == '':
                    if q_b_b_s_b_budget == '': 
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                    else:
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                    
                else:
                    if q_b_b_s_b_budget == '': 
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                    else:
                        statistics_sql  = f"select ITEM11 , ITEM19 , ITEM24 , format(ITEM15 , 'N0' , 'en-US') , ITEM12 , ITEM50 from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                    
                
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            return statistics_res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_form_search : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()


    #####################################################
    # show_bpm_expenditure_budget_source_by_dep
    #####################################################
    def show_bpm_expenditure_budget_source_by_dep(self , dep):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            ## chart 
            statistics_sql  = f"select ITEM24 from ART00691673700397107_Ins where ITEM56='{dep}' group by ITEM24"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            return statistics_res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_budget_source_by_dep : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #####################################################
    # show_erp_realtime_query_data2
    #####################################################
    def show_erp_realtime_query_data2(self , item , p_num , p_name , p_a_num):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            
            
            if item == '有效日期':

                sql  = f"SELECT DISTINCT "
                sql += f"A.MF001 AS '產品品號' , A.MF002 AS '產品批號' , C.MB002 AS '產品名稱' , E.ME009 AS '有效日期' "
                sql += f"FROM OtsukaDB.dbo.INVMF A "
                sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                sql += f"JOIN OtsukaDB.dbo.INVME E ON A.MF001 = E.ME001 AND A.MF002 = E.ME002 "
                sql += f"WHERE A.MF003 >= '20190101' "
                sql += f"AND A.MF001 LIKE 'X%' "
                sql += f"AND A.MF001 = '{p_num}' "
                sql += f"AND C.MB002 = '{p_name}' "
                sql += f"AND A.MF002 = '{p_a_num}' "
                sql += f"ORDER BY E.ME009 DESC"
                
                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            elif item == '產品批號':

                sql  = f"SELECT DISTINCT "
                sql += f"A.MF001 AS '產品品號' , A.MF002 AS '產品批號' , C.MB002 AS '產品名稱' "
                sql += f"FROM OtsukaDB.dbo.INVMF A "
                sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                sql += f"WHERE A.MF003 >= '20190101' "
                sql += f"AND A.MF001 LIKE 'X%' "
                sql += f"AND A.MF001 = '{p_num}' "
                sql += f"AND C.MB002 = '{p_name}' "
                sql += f"ORDER BY A.MF002 DESC"

                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            elif item == '產品名稱':

                sql  = f"SELECT DISTINCT "
                sql += f"A.MF001 AS '產品品號' , C.MB002 AS '產品名稱' "
                sql += f"FROM OtsukaDB.dbo.INVMF A "
                sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                sql += f"WHERE A.MF003 >= '20190101' "
                sql += f"AND A.MF001 LIKE 'X%' "
                sql += f"AND A.MF001 = '{p_num}' "
                sql += f"ORDER BY A.MF001"

                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_erp_realtime_query_data : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #####################################################
    # show_erp_realtime_query_data3_1
    #####################################################
    def show_erp_realtime_query_data3_1(self , item , q_e_p_num):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            if item == '產品批號':

                sql  = f"SELECT DISTINCT "
                sql += f"B.TA063 AS '產品批號' "
                sql += f"FROM " 
                sql += f"OtsukaDB.dbo.MOCTE A "
                sql += f"JOIN " 
                sql += f"OtsukaDB.dbo.MOCTA B ON A.TE011 = B.TA001 AND A.TE012 = B.TA002 "
                sql += f"JOIN " 
                sql += f"OtsukaDB.dbo.INVMB C ON B.TA006 = C.MB001 "
                sql += f"LEFT JOIN " 
                sql += f"OtsukaDB.dbo.INVME D ON A.TE004 = D.ME001 AND A.TE010 = D.ME002 "
                sql += f"LEFT JOIN " 
                sql += f"OtsukaDB.dbo.INVMB E ON D.ME001 = E.MB001 "
                sql += f"LEFT JOIN " 
                sql += f"OtsukaDB.dbo.PURMA F ON E.MB032 = F.MA001 "
                sql += f"WHERE " 
                sql += f"B.TA009 >= '20240101' " 
                sql += f"AND D.ME003 >= '20200101' "
                sql += f"AND B.TA006 = '{q_e_p_num}' "
                sql += f"ORDER BY B.TA063 ASC"

                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_erp_realtime_query_data3_1 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()


    #####################################################
    # show_erp_realtime_query_data
    #####################################################
    def show_erp_realtime_query_data(self , item , p_num , p_name):
        
        self.__connect_mssql_erp__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            
            
            if item == '有效日期':

                sql  = f"SELECT DISTINCT "
                sql += f"A.MF001 AS '產品品號' , A.MF002 AS '產品批號' , C.MB002 AS '產品名稱' , E.ME009 AS '有效日期' "
                sql += f"FROM OtsukaDB.dbo.INVMF A "
                sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                sql += f"JOIN OtsukaDB.dbo.INVME E ON A.MF001 = E.ME001 AND A.MF002 = E.ME002 "
                sql += f"WHERE A.MF003 >= '20190101' "
                sql += f"AND A.MF001 LIKE 'X%' "
                sql += f"AND A.MF001 = '{p_num}' "
                sql += f"AND C.MB002 = '{p_name}' "
                sql += f"ORDER BY E.ME009 DESC"
                
                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            elif item == '產品批號':

                sql  = f"SELECT DISTINCT "
                sql += f"A.MF001 AS '產品品號' , A.MF002 AS '產品批號' , C.MB002 AS '產品名稱' "
                sql += f"FROM OtsukaDB.dbo.INVMF A "
                sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                sql += f"WHERE A.MF003 >= '20190101' "
                sql += f"AND A.MF001 LIKE 'X%' "
                sql += f"AND A.MF001 = '{p_num}' "
                sql += f"AND C.MB002 = '{p_name}' "
                sql += f"ORDER BY A.MF002 DESC"

                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            elif item == '產品名稱':

                sql  = f"SELECT DISTINCT "
                sql += f"A.MF001 AS '產品品號' , C.MB002 AS '產品名稱' "
                sql += f"FROM OtsukaDB.dbo.INVMF A "
                sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                sql += f"WHERE A.MF003 >= '20190101' "
                sql += f"AND A.MF001 LIKE 'X%' "
                sql += f"AND A.MF001 = '{p_num}' "
                sql += f"ORDER BY A.MF001"

                self.curr_mssql_erp.execute(sql)
                res = self.curr_mssql_erp.fetchall()

            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_erp_realtime_query_data : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #####################################################
    # show_bpm_expenditure_deplist_member
    #####################################################
    def show_bpm_expenditure_deplist_member(self , dep):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            ## chart 
            statistics_sql  = f"select ITEM14 from ART00691673700397107_Ins where ITEM56='{dep}' group by ITEM14"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            return statistics_res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_deplist_member : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #####################################################
    # show_bpm_expenditure_deplist
    #####################################################
    def show_bpm_expenditure_deplist(self):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            ## chart 
            statistics_sql  = f"select ITEM9 from ART00691673700397107_Ins order by ITEM12 desc"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            return statistics_res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_deplist : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    
    #####################################################
    # show_bpm_expenditure_by_dep_kind_statistics_list
    #####################################################
    def show_bpm_expenditure_by_dep_kind_statistics_list(self , dep):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            ## chart 
            statistics_sql  = f"select ITEM24 , count(*) from ART00691673700397107_Ins where ITEM49='true' and ITEM56='{dep}' group by ITEM24 order by ITEM24 desc"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            return statistics_res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_by_dep_kind_statistics_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #####################################################
    # show_bpm_information_by_dep_kind_statistics_list
    #####################################################
    def show_bpm_information_by_dep_kind_statistics_list(self , dep):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            res = []

            ## chart 
            statistics_sql  = f"SELECT ITEM147 , count(*)  FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM35='true' group by ITEM147;"
            self.curr_mssql.execute(statistics_sql)
            statistics_res = self.curr_mssql.fetchall() 

            for val in statistics_res:
                res.append((str(val[0]).replace(",","").replace("_"," - ") , val[1]))
            
            return res
        
        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_statistics_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #####################################################
    # show_bpm_information_by_dep_kind_by_account_list
    #####################################################
    def show_bpm_information_by_dep_kind_by_account_list(self , dep , kind , user):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            if kind == '硬體,':
                label_title = 'Hardware'
                title = f"BPM - information form {label_title}"
            else:
                label_title = 'Software'
                title = f"BPM - information form {label_title}"

            res = []

            ## chart 
            s_1_sql  = f"SELECT ITEM11 , FORMAT(ITEM98 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            for val in s_1_res:
                kind_sql = f"select ITEM147 from ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM11='{val[0]}' and ITEM125='{kind}' and ITEM35='true'"
                self.curr_mssql.execute(kind_sql)
                kind_res = self.curr_mssql.fetchall()

                for kind_val in kind_res:
                    res.append((val[0] , val[1] , str(kind_val[0]).replace(",","").replace("_"," - ")))

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_by_account_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #########################################
    # show_bpm_information_by_dep_kind_list
    #########################################
    def show_bpm_information_by_dep_kind_list(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            if kind == '硬體,':
                label_title = 'Hardware'
                title = f"BPM - information form {label_title}"
            else:
                label_title = 'Software'
                title = f"BPM - information form {label_title}"

            res = []

            ## chart 
            s_1_sql  = f"SELECT ITEM11 , FORMAT(ITEM98 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            for val in s_1_res:
                kind_sql = f"select ITEM147 from ART01231708483412487_INS where ITEM19='{dep}' and ITEM11='{val[0]}' and ITEM125='{kind}' and ITEM35='true'"
                self.curr_mssql.execute(kind_sql)
                kind_res = self.curr_mssql.fetchall()

                for kind_val in kind_res:
                    res.append((val[0] , val[1] , str(kind_val[0]).replace(",","").replace("_"," - ") ))

            return res

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    ###########################################################
    # show_bpm_information_by_dep_kind_detail_img
    ###########################################################
    def show_bpm_information_by_dep_kind_detail_img(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            if kind == '硬體,':
                label_title = '硬體'
                title = f"BPM - 電子表單 ({label_title})"
            else:
                label_title = '軟體'
                title = f"BPM - 電子表單 ({label_title})"

            kind = str(kind).replace(" - ","_") + ","

            ## chart 
            s_1_sql  = f"SELECT ITEM11 , FORMAT(ITEM98 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM147='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = []
            y1 = [] 
            y2 = []           

            for row in s_1_res:
                try:
                    date_str = row[0]  
                    date_obj = datetime.strptime(date_str, '%Y/%m/%d')  # 轉換日期格式
                    x.append(date_obj.strftime('%Y/%m/%d'))  # 格式化日期
                    num_str = row[1].replace(',', '')  # 移除千位分隔符
                    y1.append(float(num_str))  
                    y2.append(float(num_str))  
                except ValueError as e:
                    logging.error(f"Error processing row {row}: {e}")
                    continue
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(10, len(x) / 2)  
            fig = Figure(figsize=(width , 6))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            #axis.plot(x, y1 , color='skyblue')
            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)
            # 長條圖
            bars = axis.bar(x , y2 , label=label_title)

            for bar in bars:
                yval = bar.get_height()
                yval = int(yval)
                axis.text(bar.get_x() + bar.get_width()/2, yval + 3, f"{yval}",  # Adjust vertical offset to fit above the bar
                        ha='center', va='bottom', color='black' , fontproperties=prop)  # Center align text

            # x 軸 
            axis.set_xticks(x)
            axis.set_xticklabels(x , rotation=45 , ha='right')

            axis.set_title(title, fontproperties=prop)
            axis.set_xlabel('日期', fontproperties=prop)
            axis.set_ylabel('費用(元)', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_detail_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()


    ###########################################################
    # show_bpm_information_by_dep_kind_detail_by_account_img
    ###########################################################
    def show_bpm_information_by_dep_kind_detail_by_account_img(self , dep , kind , user):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            if kind == '硬體,':
                label_title = '硬體'
                title = f"BPM - 電子表單 ({label_title})"
            else:
                label_title = '軟體'
                title = f"BPM - 電子表單 ({label_title})"

            kind = str(kind).replace(" - ","_") + ","

            ## chart 
            s_1_sql  = f"SELECT ITEM11 , FORMAT(ITEM98 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM147='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = []
            y1 = [] 
            y2 = []           

            for row in s_1_res:
                try:
                    date_str = row[0]  
                    date_obj = datetime.strptime(date_str, '%Y/%m/%d')  # 轉換日期格式
                    x.append(date_obj.strftime('%Y/%m/%d'))  # 格式化日期
                    num_str = row[1].replace(',', '')  # 移除千位分隔符
                    y1.append(float(num_str))  
                    y2.append(float(num_str))  
                except ValueError as e:
                    logging.error(f"Error processing row {row}: {e}")
                    continue
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(10, len(x) / 2)  
            fig = Figure(figsize=(width , 6))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            #axis.plot(x, y1 , color='skyblue')
            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)
            # 長條圖
            bars = axis.bar(x , y2 , label=label_title)

            for bar in bars:
                yval = bar.get_height()
                yval = int(yval)
                axis.text(bar.get_x() + bar.get_width()/2, yval + 3, f"{yval}",  # Adjust vertical offset to fit above the bar
                        ha='center', va='bottom', color='black' , fontproperties=prop)  # Center align text

            # x 軸 
            axis.set_xticks(x)
            axis.set_xticklabels(x , rotation=45 , ha='right')

            axis.set_title(title, fontproperties=prop)
            axis.set_xlabel('日期', fontproperties=prop)
            axis.set_ylabel('費用(元)', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_detail_by_account_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    ####################################################
    # show_bpm_information_by_dep_kind_by_account_img
    ####################################################
    def show_bpm_information_by_dep_kind_by_account_img(self , dep , kind , user):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            if kind == '硬體,':
                label_title = '硬體'
                title = f"BPM - 電子表單 ({label_title})"
            else:
                label_title = '軟體'
                title = f"BPM - 電子表單 ({label_title})"

            ## chart 
            s_1_sql  = f"SELECT ITEM11 , FORMAT(ITEM98 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = []
            y1 = [] 
            y2 = []           

            for row in s_1_res:
                try:
                    date_str = row[0]  
                    date_obj = datetime.strptime(date_str, '%Y/%m/%d')  # 轉換日期格式
                    x.append(date_obj.strftime('%Y/%m/%d'))  # 格式化日期
                    num_str = row[1].replace(',', '')  # 移除千位分隔符
                    y1.append(float(num_str))  
                    y2.append(float(num_str))  
                except ValueError as e:
                    logging.error(f"Error processing row {row}: {e}")
                    continue
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(10, len(x) / 2)  
            fig = Figure(figsize=(width , 6))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            #axis.plot(x, y1 , color='skyblue')
            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)
            # 長條圖
            bars = axis.bar(x , y2 , label=label_title)

            for bar in bars:
                yval = bar.get_height()
                yval = int(yval)
                axis.text(bar.get_x() + bar.get_width()/2, yval + 3, f"{yval}",  # Adjust vertical offset to fit above the bar
                        ha='center', va='bottom', color='black' , fontproperties=prop)  # Center align text

            # x 軸 
            axis.set_xticks(x)
            axis.set_xticklabels(x , rotation=45 , ha='right')

            axis.set_title(title, fontproperties=prop)
            axis.set_xlabel('日期', fontproperties=prop)
            axis.set_ylabel('費用(元)', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_kind_by_account_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #########################################
    # show_bpm_expenditure_each_dep_list
    #########################################
    def show_bpm_expenditure_each_dep_list(self):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            label_title = "BPM 開支證明單 - 各部門使用統計清單"
            
            ## list
            s_1_sql = f"select ITEM56 , count(*) from ART00691673700397107_Ins where ITEM49='true' group by ITEM56 order by ITEM56 desc"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            return s_1_res

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_each_dep_list : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #########################################
    # show_bpm_expenditure_each_dep_img
    #########################################
    def show_bpm_expenditure_each_dep_img(self):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            label_title = "BPM 開支證明單 - 各部門使用統計圖"
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            ## chart 
            s_1_sql = f"select ITEM56 , count(*) from ART00691673700397107_Ins where ITEM49='true' group by ITEM56  order by ITEM56 desc"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = [int(row[1]) for row in s_1_res]  
            y1 = [str(row[0]) for row in s_1_res]  
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(len(x) / 2,10)
            height_per_category = 1  # 每個類別分配1英寸的高度
            fig = Figure(figsize=(10,height_per_category * len(y1) ))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            #axis.plot(x, y1 , color='skyblue')

            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)

            # 長條圖
            #bars = axis.bar(x , y1 , label=label_title)
            # 左長條圖
            bars = axis.barh(y1 , x , label=label_title)

            for bar in bars:
                width = bar.get_width()  # Get the width of the bar
                label_x_pos = width + 5  # Shift the text to the right side of the bar
                axis.text(label_x_pos, bar.get_y() + bar.get_height() / 2, str(width), va='center' , color='black' , fontproperties=prop)

            # x 軸 中文指定字體
            # Setting font properties in ticks
            axis.set_xticklabels([f'{int(x)}' for x in axis.get_xticks()], fontproperties=prop)  # Apply after setting ticks
            axis.set_yticklabels(y1, rotation=30, ha='right', fontproperties=prop)
            
            
            axis.set_title(label_title, fontproperties=prop)
            axis.set_xlabel('申請數', fontproperties=prop)
            axis.set_ylabel('部門', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_each_dep_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #########################################
    # show_bpm_information_each_dep_img
    #########################################
    def show_bpm_information_each_dep_img(self):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            label_title = "BPM 資訊需求單 - 各部門使用統計圖"
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            ## chart 
            s_1_sql = f"SELECT ITEM19 , count(*) FROM ART01231708483412487_INS where ITEM35='true' group by ITEM19"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = [row[0] for row in s_1_res]  
            y1 = [int(row[1]) for row in s_1_res]  
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(10, len(x) / 2)  
            fig = Figure(figsize=(width , 6))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            axis.plot(x, y1 , color='skyblue')

            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)

            # 長條圖
            bars = axis.bar(x , y1 , label=label_title)

            for bar in bars:
                yval = bar.get_height()
                yval = int(yval)
                axis.text(bar.get_x() + bar.get_width()/2, yval + 3, f"{yval}",  # Adjust vertical offset to fit above the bar
                        ha='center', va='bottom', color='black' , fontproperties=prop)  # Center align text

            # x 軸 中文指定字體
            axis.set_xticks(x)
            axis.set_xticklabels(x, fontproperties=prop , rotation=45, ha='right')

            axis.set_title(label_title, fontproperties=prop)
            axis.set_xlabel('部門', fontproperties=prop)
            axis.set_ylabel('申請數', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_each_dep_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #########################################
    # show_bpm_expenditure_by_dep_kind_img
    #########################################
    def show_bpm_expenditure_by_dep_kind_img(self , dep):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date      = time.strftime("%Y_%m" , time.localtime())
            r_time      = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            label_title = "BPM 開支證明單 - 已結案統計圖"
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            ## chart 
            s_1_sql  = f"select ITEM24 , count(*) from ART00691673700397107_Ins where ITEM49='true' and ITEM56='{dep}' group by ITEM24"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = [row[0] for row in s_1_res]  
            y1 = [int(row[1]) for row in s_1_res]  
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(10, len(x) / 2)  
            fig = Figure(figsize=(width , 6))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            #axis.plot(x, y1 , color='skyblue')
            
            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)
            
            # 長條圖
            bars = axis.bar(x , y1 , label=label_title)

            for bar in bars:
                yval = bar.get_height()
                yval = int(yval)
                axis.text(bar.get_x() + bar.get_width()/2, yval + 3, f"{yval}",  # Adjust vertical offset to fit above the bar
                        ha='center', va='bottom', color='black' , fontproperties=prop)  # Center align text


            # x 軸 
            axis.set_xticks(x)
            axis.set_xticklabels(x , rotation=45 , ha='right' , fontproperties=prop)

            axis.set_title(label_title , fontproperties=prop)
            axis.set_xlabel('費用種類', fontproperties=prop)
            axis.set_ylabel('數量統計', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_expenditure_by_dep_kind_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()
    
    #########################################
    # show_bpm_information_by_dep_kind_img
    #########################################
    def show_bpm_information_by_dep_kind_img(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            ### variable
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ### 設置全局字體 - 中文 
            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替換為你的字體文件路徑
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            if kind == '硬體,':
                label_title = '硬體'
                title = f"BPM - 電子表單 ({label_title})"
            else:
                label_title = '軟體'
                title = f"BPM - 電子表單 ({label_title})"

            ## chart 
            s_1_sql  = f"SELECT ITEM11 , FORMAT(ITEM98 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM125='{kind}' and ITEM35='true'"
            self.curr_mssql.execute(s_1_sql)
            s_1_res = self.curr_mssql.fetchall() 

            x  = []
            y1 = [] 
            y2 = []           

            for row in s_1_res:
                try:
                    date_str = row[0]  
                    date_obj = datetime.strptime(date_str, '%Y/%m/%d')  # 轉換日期格式
                    x.append(date_obj.strftime('%Y/%m/%d'))  # 格式化日期
                    num_str = row[1].replace(',', '')  # 移除千位分隔符
                    y1.append(float(num_str))  
                    y2.append(float(num_str))  
                except ValueError as e:
                    logging.error(f"Error processing row {row}: {e}")
                    continue
            
            # 生成多條線圖
            fig  = Figure()
            # img auto size
            width = max(10, len(x) / 2)  
            fig = Figure(figsize=(width , 6))
            axis = fig.add_subplot(1, 1, 1)
            
            # 折線圖
            #axis.plot(x, y1 , label=label_title , marker='o', markersize=5 , color='green')
            #axis.plot(x, y1 , color='skyblue')
            
            # 點圖
            #axis.scatter(x, y2, color='red', label=label_title)
            
            # 長條圖
            bars = axis.bar(x , y2 , label=label_title)

            for bar in bars:
                yval = bar.get_height()
                yval = int(yval)
                axis.text(bar.get_x() + bar.get_width()/2, yval + 3, f"{yval}",  # Adjust vertical offset to fit above the bar
                        ha='center', va='bottom', color='black' , fontproperties=prop)  # Center align text


            # x 軸 
            axis.set_xticks(x)
            axis.set_xticklabels(x , rotation=45 , ha='right')

            axis.set_title(title, fontproperties=prop)
            axis.set_xlabel('日期', fontproperties=prop)
            axis.set_ylabel('費用(元)', fontproperties=prop)
    
            axis.legend(prop=prop)
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error(f"\n<Error> show_bpm_information_by_dep_img : {str(e)}\n")

        finally:
            self.__disconnect_mssql__()

    #########################################################
    # show_factory_monitor_detail_warehouse_temp_img
    #########################################################
    def show_factory_monitor_detail_warehouse_temp_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-1
            s_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-1' order by r_time desc limit 0,40"
            self.curr.execute(s_1_sql)
            s_1_res = self.curr.fetchall() 
            ### S-2
            s_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-2' order by r_time desc limit 0,40"
            self.curr.execute(s_2_sql)
            s_2_res = self.curr.fetchall() 
            ### S-3
            s_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-3' order by r_time desc limit 0,40"
            self.curr.execute(s_3_sql)
            s_3_res = self.curr.fetchall() 
            ### S-4
            s_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-4' order by r_time desc limit 0,40"
            self.curr.execute(s_4_sql)
            s_4_res = self.curr.fetchall() 
            ### S-5
            s_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-5' order by r_time desc limit 0,40"
            self.curr.execute(s_5_sql)
            s_5_res = self.curr.fetchall() 
            ### S-6
            s_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-6' order by r_time desc limit 0,40"
            self.curr.execute(s_6_sql)
            s_6_res = self.curr.fetchall() 
            ### S-7
            s_7_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-7' order by r_time desc limit 0,40"
            self.curr.execute(s_7_sql)
            s_7_res = self.curr.fetchall() 
            ### S-8
            s_8_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-8' order by r_time desc limit 0,40"
            self.curr.execute(s_8_sql)
            s_8_res = self.curr.fetchall() 
            ### S-9
            s_9_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-9' order by r_time desc limit 0,40"
            self.curr.execute(s_9_sql)
            s_9_res = self.curr.fetchall() 
            ### S-10
            s_10_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-10' order by r_time desc limit 0,40"
            self.curr.execute(s_10_sql)
            s_10_res = self.curr.fetchall() 
            ### S-14
            s_14_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-14' order by r_time desc limit 0,40"
            self.curr.execute(s_14_sql)
            s_14_res = self.curr.fetchall() 
            ### S-17
            s_17_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-17' order by r_time desc limit 0,40"
            self.curr.execute(s_17_sql)
            s_17_res = self.curr.fetchall() 
            ### S-18
            s_18_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-18' order by r_time desc limit 0,40"
            self.curr.execute(s_18_sql)
            s_18_res = self.curr.fetchall() 

            x  = [row[0] for row in s_1_res]  # x轴数据
            y1 = [float(row[3]) for row in s_1_res]  # S-1
            y2 = [float(row[3]) for row in s_2_res]  # S-2
            y3 = [float(row[3]) for row in s_3_res]  # S-3
            y4 = [float(row[3]) for row in s_4_res]  # S-4
            y5 = [float(row[3]) for row in s_5_res]  # S-5
            y6 = [float(row[3]) for row in s_6_res]  # S-6
            y7 = [float(row[3]) for row in s_7_res]  # S-7
            y8 = [float(row[3]) for row in s_8_res]  # S-8
            y9 = [float(row[3]) for row in s_9_res]  # S-9
            y10 = [float(row[3]) for row in s_10_res]  # S-10
            y14 = [float(row[3]) for row in s_14_res]  # S-14
            y17 = [float(row[3]) for row in s_17_res]  # S-17
            y18 = [float(row[3]) for row in s_18_res]  # S-14
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='S-1')
            axis.plot(x, y2 , label='S-2')
            axis.plot(x, y3 , label='S-3')
            axis.plot(x, y4 , label='S-4')
            axis.plot(x, y5 , label='S-5')
            axis.plot(x, y6 , label='S-6')
            axis.plot(x, y7 , label='S-7')
            axis.plot(x, y8 , label='S-8')
            axis.plot(x, y9 , label='S-9')
            axis.plot(x, y10 , label='S-10')
            axis.plot(x, y14 , label='S-14')
            axis.plot(x, y17 , label='S-17')
            axis.plot(x, y18 , label='S-18')

            axis.set_title('Warehouse Sensor Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_warehouse_temp_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################################
    # show_factory_monitor_detail_quality_control_rh_img
    #########################################################
    def show_factory_monitor_detail_quality_control_rh_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-11-1
            s_11_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-1' order by r_time desc limit 0,40"
            self.curr.execute(s_11_1_sql)
            s_11_1_res = self.curr.fetchall() 
            ## S-11-2
            s_11_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-2' order by r_time desc limit 0,40"
            self.curr.execute(s_11_2_sql)
            s_11_2_res = self.curr.fetchall() 
            ### S-12
            s_12_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-12' order by r_time desc limit 0,40"
            self.curr.execute(s_12_sql)
            s_12_res = self.curr.fetchall() 
            ### S-13
            s_13_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-13' order by r_time desc limit 0,40"
            self.curr.execute(s_13_sql)
            s_13_res = self.curr.fetchall() 
            ### S-15-1
            s_15_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-1' order by r_time desc limit 0,40"
            self.curr.execute(s_15_1_sql)
            s_15_1_res = self.curr.fetchall()
            ### S-15-2
            s_15_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-2' order by r_time desc limit 0,40"
            self.curr.execute(s_15_2_sql)
            s_15_2_res = self.curr.fetchall() 
            ### S-15-3
            s_15_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-3' order by r_time desc limit 0,40"
            self.curr.execute(s_15_3_sql)
            s_15_3_res = self.curr.fetchall() 
            ### S-15-4
            s_15_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-4' order by r_time desc limit 0,40"
            self.curr.execute(s_15_4_sql)
            s_15_4_res = self.curr.fetchall() 
            ### S-15-5
            s_15_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-5' order by r_time desc limit 0,40"
            self.curr.execute(s_15_5_sql)
            s_15_5_res = self.curr.fetchall() 
            ### S-15-6
            s_15_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-6' order by r_time desc limit 0,40"
            self.curr.execute(s_15_6_sql)
            s_15_6_res = self.curr.fetchall() 
            ### S-16
            s_16_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-16' order by r_time desc limit 0,40"
            self.curr.execute(s_16_sql)
            s_16_res = self.curr.fetchall() 
            ### S-19
            s_19_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-19' order by r_time desc limit 0,40"
            self.curr.execute(s_19_sql)
            s_19_res = self.curr.fetchall() 

            x  = [row[0] for row in s_11_1_res]  # x轴数据
            y11_1 = [float(row[4]) for row in s_11_1_res]  # S-11-1
            y11_2 = [float(row[4]) for row in s_11_2_res]  # S-11-2
            y12   = [float(row[4]) for row in s_12_res]  # S-12
            y13   = [float(row[4]) for row in s_13_res]  # S-13
            y15_1 = [float(row[4]) for row in s_15_1_res]  # S-15-1
            y15_2 = [float(row[4]) for row in s_15_2_res]  # S-15-2
            y15_3 = [float(row[4]) for row in s_15_3_res]  # S-15-3
            y15_4 = [float(row[4]) for row in s_15_4_res]  # S-15-4
            y15_5 = [float(row[4]) for row in s_15_5_res]  # S-15-5
            y15_6 = [float(row[4]) for row in s_15_6_res]  # S-15-6
            y16   = [float(row[4]) for row in s_16_res]  # S-16
            y19   = [float(row[4]) for row in s_19_res]  # S-19
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y11_1 , label='S-11-1')
            axis.plot(x, y11_2 , label='S-11-2')
            axis.plot(x, y12 , label='S-12')
            axis.plot(x, y13 , label='S-13')
            axis.plot(x, y15_1 , label='S-15-1')
            axis.plot(x, y15_2 , label='S-15-2')
            axis.plot(x, y15_3 , label='S-15-3')
            axis.plot(x, y15_4 , label='S-15-4')
            axis.plot(x, y15_5 , label='S-15-5')
            axis.plot(x, y15_6 , label='S-15-6')
            axis.plot(x, y16 , label='S-16')
            axis.plot(x, y19 , label='S-19')

            axis.set_title('Quality Control Sensor RH')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (%)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_quality_control_rh_img : ' + str(e))

        finally:
            self.__disconnect4__()

    #########################################################
    # show_factory_monitor_detail_quality_control_temp_img
    #########################################################
    def show_factory_monitor_detail_quality_control_temp_img(self):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            
            ## S-11-1
            s_11_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-1' order by r_time desc limit 0,40"
            self.curr.execute(s_11_1_sql)
            s_11_1_res = self.curr.fetchall() 
            ## S-11-2
            s_11_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-11-2' order by r_time desc limit 0,40"
            self.curr.execute(s_11_2_sql)
            s_11_2_res = self.curr.fetchall() 
            ### S-12
            s_12_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-12' order by r_time desc limit 0,40"
            self.curr.execute(s_12_sql)
            s_12_res = self.curr.fetchall() 
            ### S-13
            s_13_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-13' order by r_time desc limit 0,40"
            self.curr.execute(s_13_sql)
            s_13_res = self.curr.fetchall() 
            ### S-15-1
            s_15_1_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-1' order by r_time desc limit 0,40"
            self.curr.execute(s_15_1_sql)
            s_15_1_res = self.curr.fetchall()
            ### S-15-2
            s_15_2_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-2' order by r_time desc limit 0,40"
            self.curr.execute(s_15_2_sql)
            s_15_2_res = self.curr.fetchall() 
            ### S-15-3
            s_15_3_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-3' order by r_time desc limit 0,40"
            self.curr.execute(s_15_3_sql)
            s_15_3_res = self.curr.fetchall() 
            ### S-15-4
            s_15_4_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-4' order by r_time desc limit 0,40"
            self.curr.execute(s_15_4_sql)
            s_15_4_res = self.curr.fetchall() 
            ### S-15-5
            s_15_5_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-5' order by r_time desc limit 0,40"
            self.curr.execute(s_15_5_sql)
            s_15_5_res = self.curr.fetchall() 
            ### S-15-6
            s_15_6_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-15-6' order by r_time desc limit 0,40"
            self.curr.execute(s_15_6_sql)
            s_15_6_res = self.curr.fetchall() 
            ### S-16
            s_16_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-16' order by r_time desc limit 0,40"
            self.curr.execute(s_16_sql)
            s_16_res = self.curr.fetchall() 
            ### S-19
            s_19_sql = f"select r_time , s_content , s_protocol , val_1 , val_2  from {r_date} where s_kind='S-19' order by r_time desc limit 0,40"
            self.curr.execute(s_19_sql)
            s_19_res = self.curr.fetchall() 

            x  = [row[0] for row in s_11_1_res]  # x轴数据
            y11_1 = [float(row[3]) for row in s_11_1_res]  # S-11-1
            y11_2 = [float(row[3]) for row in s_11_2_res]  # S-11-2
            y12 = [float(row[3]) for row in s_12_res]  # S-12
            y13 = [float(row[3]) for row in s_13_res]  # S-13
            y15_1 = [float(row[3]) for row in s_15_1_res]  # S-15-1
            y15_2 = [float(row[3]) for row in s_15_2_res]  # S-15-2
            y15_3 = [float(row[3]) for row in s_15_3_res]  # S-15-3
            y15_4 = [float(row[3]) for row in s_15_4_res]  # S-15-4
            y15_5 = [float(row[3]) for row in s_15_5_res]  # S-15-5
            y15_6 = [float(row[3]) for row in s_15_6_res]  # S-15-6
            y16 = [float(row[3]) for row in s_16_res]  # S-16
            y19 = [float(row[3]) for row in s_19_res]  # S-19
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y11_1 , label='S-11-1')
            axis.plot(x, y11_2 , label='S-11-2')
            axis.plot(x, y12 , label='S-12')
            axis.plot(x, y13 , label='S-13')
            axis.plot(x, y15_1 , label='S-15-1')
            axis.plot(x, y15_2 , label='S-15-2')
            axis.plot(x, y15_3 , label='S-15-3')
            axis.plot(x, y15_4 , label='S-15-4')
            axis.plot(x, y15_5 , label='S-15-5')
            axis.plot(x, y15_6 , label='S-15-6')
            axis.plot(x, y16 , label='S-16')
            axis.plot(x, y19 , label='S-19')

            axis.set_title('Quality Control Sensor Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_quality_control_temp_img : ' + str(e))

        finally:
            self.__disconnect4__()

    ##############################################
    # show_factory_monitor_detail_temp_rh_img_1
    ##############################################
    def show_factory_monitor_detail_temp_rh_img_1(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            x  = [row[0] for row in d_res]  # x轴数据
            y1 = [float(row[3]) for row in d_res]  # 第一条线的y轴数据
            
            # 生成多条线图
            fig  = Figure()
            axis = fig.add_subplot(1, 1, 1)
            axis.plot(x, y1 , label='Temp' ,  marker='o', markersize=4)
            
            axis.set_title(s_kind + ' Temp')
            axis.set_xlabel('date time')
            axis.set_ylabel('value (°C)')
    
            axis.legend()
            fig.tight_layout()

            img = io.BytesIO()
            FigureCanvas(fig).print_png(img)
            img.seek(0)

            # Convert the image to base64 for embedding in HTML
            img_data = base64.b64encode(img.getvalue()).decode('utf-8')

            return img_data

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail_temp_rh_img_1 : ' + str(e))

        finally:
            self.__disconnect4__()

    ################################
    # show_factory_monitor_detail
    ################################
    def show_factory_monitor_detail(self , s_kind):
        
        self.__connect4__()
        
        try:
            ### r_time
            r_date = time.strftime("%Y_%m" , time.localtime())
            r_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())

            # all device user 
            d_sql = f"select r_time , s_content , s_protocol , val_1 , val_2 , val_3 , val_4 , val_5   from {r_date} where s_kind='{s_kind}' order by r_time desc limit 0,40"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_factory_monitor_detail : ' + str(e))

        finally:
            self.__disconnect4__()



    #############################
    # show_vmware_item_detail
    #############################
    def show_vmware_item_detail(self , vm_name , vm_name_item):
        
        self.__connect__()
        
        try:
            # all device user
            if vm_name_item == 'vm_boot_time': 
                d_sql = f"SELECT DATE_FORMAT({vm_name_item} , '%Y-%m-%d %H:%i:%s') FROM `otsuka_vmware` WHERE vm_name='{vm_name}' group by {vm_name_item} order by c_d_time desc limit 0,1"
            else:
                d_sql = f"SELECT {vm_name_item} FROM `otsuka_vmware` WHERE vm_name='{vm_name}' group by {vm_name_item} order by c_d_time desc limit 0,1"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchone() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_item_detail : ' + str(e))

        finally:
            self.__disconnect__()
    
    ###################################
    # show_vmware_detail_item_record
    ###################################
    def show_vmware_detail_item_record(self , vm_name , vm_name_item):
        
        self.__connect__()
        
        try:
            # all device user
            if vm_name_item == 'vm_boot_time': 
                d_sql = f"SELECT c_d_time , DATE_FORMAT({vm_name_item} , '%Y-%m-%d %H:%i:%s') , count(*) FROM `otsuka_vmware` WHERE vm_name='{vm_name}' group by {vm_name_item} order by c_d_time desc limit 0,6"
            else:
                d_sql = f"SELECT c_d_time , {vm_name_item} , count(*) FROM `otsuka_vmware` WHERE vm_name='{vm_name}' group by {vm_name_item} order by c_d_time desc limit 0,6"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_detail_item_record : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##############################
    # show_vmware_detail_record
    ##############################
    def show_vmware_detail_record(self , vm_name):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"SELECT * FROM `otsuka_vmware` WHERE vm_name='{vm_name}' order by c_d_time desc limit 0,120"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_detail_record : ' + str(e))

        finally:
            self.__disconnect__()

    #############################
    # mrd_8_del_auto_email
    #############################
    def mrd_8_del_auto_email(self , email):
        
        self.__connect__()
        
        try:
            # check email
            c_sql = f"select email from mrd_8_auto_email_push where email='{email}'"
            self.curr.execute(c_sql)
            c_res = self.curr.fetchone() 

            if c_res is not None:
                a_sql = f"delete from mrd_8_auto_email_push where email='{email}'"
                self.curr.execute(a_sql)
                self.conn.commit()
                res = 'ok'

                return res

        except Exception as e:
            logging.error(f'\n<Error> mrd_8_del_auto_email : {str(e)}\n')

        finally:
            self.__disconnect__()

    #############################
    # mrd_8_add_auto_email
    #############################
    def mrd_8_add_auto_email(self , email):
        
        self.__connect__()
        
        try:
            # check email
            c_sql = f"select email from mrd_8_auto_email_push where email='{email}'"
            self.curr.execute(c_sql)
            c_res = self.curr.fetchone() 

            if c_res is None:
                a_sql = f"insert into mrd_8_auto_email_push(email) value('{email}')"
                self.curr.execute(a_sql)
                self.conn.commit()
                res = 'ok'

                return res

        except Exception as e:
            logging.error(f'\n<Error> mrd_8_add_auto_email : {str(e)}\n')

        finally:
            self.__disconnect__()
    
    #############################
    # show_vmware_host_factory
    #############################
    def show_vmware_host_factory(self):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"SELECT DISTINCT vm_position , vm_name , vm_os , vm_os_state FROM `otsuka_vmware` where vm_position='factory' order by vm_os_state desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_host_factory : ' + str(e))

        finally:
            self.__disconnect__()
    
    #############################
    # show_vmware_host_taipei
    #############################
    def show_vmware_host_taipei(self):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"SELECT DISTINCT vm_position , vm_name , vm_os , vm_os_state FROM `otsuka_vmware` where vm_position='taipei' order by vm_os_state desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_host_taipei : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##################################
    # show_vmware_statistics_factory
    ##################################
    def show_vmware_statistics_factory(self):
        
        self.__connect__()
        
        try:
            # all device user 
            #d_sql = f"SELECT DISTINCT vm_position , vm_name , vm_os_state , DATE_FORMAT(vm_boot_time, '%Y-%m-%d %H:%i:%s') FROM `otsuka_vmware` where vm_position='factory' order by vm_os_state desc"
            d_sql = f"SELECT DISTINCT vm_position , vm_name , vm_os_state , vm_ip FROM `otsuka_vmware` where vm_position='factory' order by vm_os_state desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_statistics_factory : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##################################
    # show_vmware_statistics_taipei
    ##################################
    def show_vmware_statistics_taipei(self):
        
        self.__connect__()
        
        try:
            # all device user 
            #d_sql = f"SELECT DISTINCT vm_position , vm_name , vm_os_state , DATE_FORMAT(vm_boot_time, '%Y-%m-%d %H:%i:%s') FROM `otsuka_vmware` where vm_position='taipei' order by vm_os_state desc"
            d_sql = f"SELECT DISTINCT vm_position , vm_name , vm_os_state , vm_ip FROM `otsuka_vmware` where vm_position='taipei' order by vm_os_state desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_vmware_statistics_taipei : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##############################
    # show_computer_user_detail
    ##############################
    def show_computer_user_detail(self , d_name):
        
        self.__connect__()
        
        try:
            # all device user 
            d_sql = f"select r_date , d_status , o_name , l_activity , e_ip , i_ip , s_number , s_model , s_manu , registered , cpu_usage , ram_usage , disk_usage from device_list where d_name='{d_name}' order by l_activity desc"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_computer_user_detail : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##########################
    # show_device_name_list
    ##########################
    def show_device_name_list(self):
        
        #self.__connect__()
        self.__connect_taipei_1_38__()
        
        try:
            # all device user 
            #d_sql = f"select r_date , d_name from device_list group by d_name order by r_date desc"
            d_sql = f"SELECT count(*) , name FROM `comodolog` group by name"
            self.curr.execute(d_sql)
            d_res = self.curr.fetchall() 

            return d_res

        except Exception as e:
            logging.error('\n<Error> show_device_name_list : ' + str(e))

        finally:
            #self.__disconnect__()
            self.__disconnect_taipei_1_38__()
            

    ###########################
    # bpm_day_money_by_month
    ###########################
    def bpm_day_money_by_month(self , year):
        
        self.__connect__()
        
        try:
            month_sql = f"select day_r_month from `day_money` WHERE  day_r_year='{str(year)}' group by day_r_month order by day_r_month desc" 
            self.curr.execute(month_sql)
            month_res = self.curr.fetchall()

            return month_res

        except Exception as e:
            logging.error('\n<Error> bpm_day_money_by_month : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################
    # bpm_day_money_by_year
    ##########################
    def bpm_day_money_by_year(self):
        
        self.__connect__()
        
        try:
            self.sql = f"select day_r_year from `day_money` WHERE day_r_year != '112/' group by day_r_year order by day_r_year desc" 
            self.curr.execute(self.sql)
            self.res = self.curr.fetchall()

            return self.res

        except Exception as e:
            logging.error('\n<Error> bpm_day_money_by_year : ' + str(e))

        finally:
            self.__disconnect__()

    ##################
    # bpm_day_money
    ##################
    def bpm_day_money(self):
        
        self.__connect__()
        
        try:

            self.sql = f"SELECT day_money.r_date , day_money.c_name , hr_a.employee_eng_name , day_money.t_money FROM `day_money` left join hr_a on day_money.c_name = hr_a.employee_name WHERE day_money.r_year='2023' and day_money.r_month='09' order by day_money.r_day desc" 
            self.curr_mssql.execute(self.sql)
            self.res           = self.curr_mssql.fetchall()

            for val in self.res:
                logging.info(f"{val[0]} , {val[1]} , {val[2]} , {val[3]}")


        except Exception as e:
            logging.info('\n<Error> bpm_day_money : ' + str(e))

        finally:
            self.__disconnect__()
    
    ###################################
    # HR_360_department_query_member
    ###################################
    def HR_360_department_query_member(self , dep):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"select DepartmentID from T_HR_Department where DepartmentName='{dep}'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchone()

            sql2 = f"select EmployeeName from T_HR_Employee the where DepartmentID ='{res[0]}'"
            self.curr_mssql.execute(sql2)
            res2 = self.curr_mssql.fetchall()

            return res2

        except Exception as e:
            logging.info('\n<Error> HR_360_department_query_member : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    ##############################
    # HR_360_employee_hr_name
    ##############################
    def HR_360_employee_hr_name(self , name):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"select EmployeeName from T_HR_Employee the where EmployeeName like '{name}%'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()

            return res

        except Exception as e:
            logging.info('\n<Error> HR_360_employee_hr_name : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    ##############################
    # HR_360_employee_hr_name
    ##############################
    def HR_360_employee_hr_name(self , name):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"select EmployeeName from T_HR_Employee the where EmployeeName like '{name}%'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()

            return res

        except Exception as e:
            logging.info('\n<Error> HR_360_employee_hr_name : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    ##############################
    # HR_360_employee
    ##############################
    def HR_360_employee(self):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"select EmployeeName from T_HR_Employee the" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()

            return res

        except Exception as e:
            logging.info('\n<Error> HR_360_HR_360_employee : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    ##############################
    # HR_360_department
    ##############################
    def HR_360_department(self):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"select DepartmentName from T_HR_Department group by DepartmentName" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()

            return res

        except Exception as e:
            logging.info('\n<Error> HR_360_department : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    ##############################
    # bpm_account_up_department
    ##############################
    def bpm_account_up_department(self , item):
        
        self.__connect_mssql__()

        try:
            
            self.sql        = f"select DepartmentName from T_HR_Department where DepartmentID = '{item}'" 
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]

        except Exception as e:
            logging.info('\n<Error> bpm_account_department : ' + str(e))
        finally:
            self.__disconnect_mssql__()
    
    #############################################
    # bpm_information_form_record_list_account
    #############################################
    def bpm_information_form_record_list_account(self , dep , account):
        
        self.__connect_mssql__()

        try:
            
            sql  = f"SELECT " 
            sql += f"ITEM10 , ITEM18 , ITEM19 , ITEM11 , ITEM125 , "
            sql += f"ITEM35 , ITEM36 , ITEM147 , ITEM75 , InsID , FORMAT(ITEM98 , 'N0' , 'en-US') , FORMAT(ITEM145 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{account}' and ITEM35='true' order by ITEM11 desc" 
            
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()

            if res is not None:
                return res

        except Exception as e:
            logging.info(f"\n<Error> bpm_information_form_record_list_account : {str(e)}\n")
        finally:
            self.__disconnect_mssql__()

    #####################################
    # bpm_expenditure_form_record_list
    #####################################
    def bpm_expenditure_form_record_list(self , dep , status):
        
        self.__connect_mssql__()

        try:
            
            ### 資訊需求單 
            '''
            sql  = f"SELECT " 
            sql += f"ITEM10 , ITEM18 , ITEM19 , ITEM11 , ITEM125 , "
            sql += f"ITEM35 , ITEM36 , ITEM147 , ITEM75 , InsID , FORMAT(ITEM98 , 'N0' , 'en-US') , FORMAT(ITEM145 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM35='{status}' order by ITEM11 desc" 
            '''

            ### 開支證明單
            sql  = f"select ITEM12 , ITEM11 , ITEM56 , ITEM50 , ITEM24 , ITEM25 , ITEM19 , FORMAT(ITEM15 , 'N0' , 'en-US') , ITEM49 "
            sql += f"from ART00691673700397107_Ins where ITEM49='{status}' and ITEM56='{dep}' order by ITEM12 DESC"

            res_all = []

            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()

            for val in res:
                res_all.append((val[0] , val[1] , val[2] , val[3] , val[4] , val[5] , val[6] , val[7] , val[8]))

            if res_all is not None:
                return res_all

        except Exception as e:
            logging.info(f"\n<Error> bpm_expenditure_form_record_list : {str(e)}\n")
        finally:
            self.__disconnect_mssql__()
    
    #####################################
    # bpm_information_form_record_list
    #####################################
    def bpm_information_form_record_list(self , dep , status):
        
        self.__connect_mssql__()

        try:
            
            sql  = f"SELECT " 
            sql += f"ITEM10 , ITEM18 , ITEM19 , ITEM11 , ITEM125 , "
            sql += f"ITEM35 , ITEM36 , ITEM147 , ITEM75 , InsID , FORMAT(ITEM98 , 'N0' , 'en-US') , FORMAT(ITEM145 , 'N0' , 'en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM35='{status}' order by ITEM11 desc" 
            
            res_all = []

            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()

            for val in res:
                res_all.append((val[0] , val[1] , val[2] , val[3] , str(val[4]).replace(",","") , val[5] , val[6] , str(val[7]).replace(",","").replace("_"," - ") , val[8] , val[9] , val[10] , val[11]))

            if res_all is not None:
                return res_all

        except Exception as e:
            logging.info(f"\n<Error> bpm_information_form_record_list : {str(e)}\n")
        finally:
            self.__disconnect_mssql__()
    
   ###############################
    # show_radar_picture2
    ###############################
    def show_radar_picture2(self , score , score2 , score3 , score4 , score5 , score6):
        try:
           
            # Example usage
            labels = ['對於客戶的服務\n(包含內部客戶或外部客戶)', '溝通與待人接物能力', '工作品質', '判斷力和決策能力', '其他能力']
            colors = ['blue', 'red', 'green', 'purple', 'orange' , 'pink']  # Add more colors if needed
            tick_labels = ['0', '1', '2', '3', '4' , '5']
            score1 = score
            score2 = score2
            score3 = score3
            score4 = score4
            score5 = score5
            score6 = score6

            values_list = [score1, score2, score3, score4, score5, score6]

             # Initialize the plot
            fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
            
            # Number of variables
            num_vars = len(labels)
            
            # Compute the angle for each label
            angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
            angles += angles[:1]  # Complete the loop
            
            # Plot each radar chart
            for i, values in enumerate(values_list):
                values = np.concatenate((values, [values[0]]))  # Close the loop
                color = colors[i % len(colors)] 
                ax.plot(angles, values, linewidth=2 , color=color)
                ax.fill(angles, values, alpha=0 , color=color)

            # Set y-ticks and their labels
            num_ticks = len(tick_labels)
            y_max = 5  # Example maximum value based on tick_labels, adjust as needed
            y_ticks = np.linspace(0, y_max, num_ticks)
            ax.set_yticks(y_ticks)  # Position of y-ticks
            ax.set_yticklabels(tick_labels)  # Labels for y-ticks

            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # Replace with your font file path
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(labels, fontproperties=prop)

            # Save the plot to a BytesIO object
            buf = io.BytesIO()
            plt.savefig(buf, format='png' , bbox_inches='tight', pad_inches=0, transparent=True)
            buf.seek(0)
            img = base64.b64encode(buf.getvalue()).decode('utf8')
            buf.close()
            plt.close(fig)  # Close the figure to free up memory

            radar_pic = f'<img class="img-fluid w-100" src="data:image/png;base64,{img}"/>'

            return radar_pic

        except Exception as e:
            logging.info(f"\n<Error> show_radar_picture : {str(e)}\n")

        finally:
            pass
   
    ###############################
    # show_rader_picture
    ###############################
    def show_radar_picture(self , score , score2 , score3 , score4 , score5 , score6):
        try:
           
            # Example usage
            labels = ['管理能力', '提供支援', '以身作則', '效率導向', '培育人才', '高效溝通']
            colors = ['blue', 'red', 'green', 'purple', 'orange', 'pink']  # Add more colors if needed
            tick_labels = ['0', '1', '2', '3', '4', '5']
            score1 = score
            score2 = score2
            score3 = score3
            score4 = score4
            score5 = score5
            score6 = score6

            values_list = [score1, score2, score3, score4, score5, score6]

             # Initialize the plot
            fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
            
            # Number of variables
            num_vars = len(labels)
            
            # Compute the angle for each label
            angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
            angles += angles[:1]  # Complete the loop
            
            # Plot each radar chart
            for i, values in enumerate(values_list):
                values = np.concatenate((values, [values[0]]))  # Close the loop
                color = colors[i % len(colors)] 
                ax.plot(angles, values, linewidth=2 , color=color)
                ax.fill(angles, values, alpha=0 , color=color)

            # Set y-ticks and their labels
            num_ticks = len(tick_labels)
            y_max = 5  # Example maximum value based on tick_labels, adjust as needed
            y_ticks = np.linspace(0, y_max, num_ticks)
            ax.set_yticks(y_ticks)  # Position of y-ticks
            ax.set_yticklabels(tick_labels)  # Labels for y-ticks

            font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # Replace with your font file path
            prop = fm.FontProperties(fname=font_path)
            plt.rcParams['font.sans-serif'] = [font_path]
            plt.rcParams['axes.unicode_minus'] = False
            
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(labels, fontproperties=prop)

            # Save the plot to a BytesIO object
            buf = io.BytesIO()
            plt.savefig(buf, format='png' , bbox_inches='tight', pad_inches=0, transparent=True)
            buf.seek(0)
            img = base64.b64encode(buf.getvalue()).decode('utf8')
            buf.close()
            plt.close(fig)  # Close the figure to free up memory

            radar_pic = f'<img class="img-fluid w-100" src="data:image/png;base64,{img}"/>'

            return radar_pic

        except Exception as e:
            logging.info(f"\n<Error> show_radar_picture : {str(e)}\n")

        finally:
            pass


    ###############################
    # show_pie_picture
    ###############################
    def show_pie_picture(self , range , score):
        try:
            # 数据
            total     = int(range)
            actual    = float(score)
            remaining = total - actual

            # 甜甜圈图的部分
            sizes = [actual, remaining]
            colors = ['blue', 'lightgray']  # 实际得分显示为蓝色，剩余部分显示为浅灰色
            labels = ['Actual Score', 'Remaining']

            # 绘制甜甜圈图
            fig, ax = plt.subplots()
            ax.pie(sizes, colors=colors, startangle=90, counterclock=False, 
                wedgeprops=dict(width=0.3))

            # 添加中心的圆来创建甜甜圈效果
            center_circle = plt.Circle((0,0), 0.7, color='white')
            fig.gca().add_artist(center_circle)

            # 在中心添加文本（显示为小数位数）
            actual_with_decimal = f'{actual:.1f}'  # 显示1位小数
            ax.text(0, 0, actual_with_decimal, horizontalalignment='center', 
                    verticalalignment='center', fontsize=60, fontweight='bold')

            # 将图表保存到内存中的一个字节流
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            img = base64.b64encode(buf.getvalue()).decode('utf8')
            buf.close()

            # 渲染图像到网页中
            pie_pic = f'<img class="img-fluid w-100" src="data:image/png;base64,{img}"/>'
        
            return pie_pic
        
        except Exception as e:
            logging.info(f"<Error> show_pie_pic : {str(e)}\n")
        finally:
            pass
    
    

    ###############################
    # hr_360_manager_question2
    ###############################
    def hr_360_manager_question2(self , c_name):
        
        self.__connect7_1_38__()

        try:

            sql2 = (
                    "select c_name from `hr_360_person_setup` where " 
                    "c_name=%s or "
                    "c_manager=%s or "
                    "c_peer1=%s or "
                    "c_peer2=%s or "
                    "c_subordinate1=%s or "
                    "c_subordinate2=%s"
            ) 

            self.curr_7_1_38.execute(sql2 , (c_name,c_name,c_name,c_name,c_name,c_name,))
            res2 = self.curr_7_1_38.fetchall()

            if res2 is not None:
                return res2
            
        except Exception as e:
            logging.info(f"\n<Error> hr_360_manager_question : {str(e)}\n")
        finally:
            self.__disconnect7_1_38__()
    
    ###################################
    # update_erp_bom
    ###################################
    def update_erp_bom(self, m_no, m_license, e_no, p_amount):
        
        self.__connect_mssql_erp__()
        
        try:
            r_year = time.strftime("%Y", time.localtime())
            d_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            r_day  = time.strftime("%Y-%m-%d", time.localtime())
            r_time = time.strftime("%H:%M:%S", time.localtime())

            m_no = m_no.strip()
            m_license = m_license.strip()
            e_no = e_no.strip()
            p_amount = p_amount.strip()

            update_sql = '''
                UPDATE OtsukaDB.dbo.BOMMD SET MD006 = ? WHERE MD001 = ? AND MD002 = ? AND MD003 = ?
            '''

            self.curr_mssql_erp.execute(update_sql, (p_amount, m_no, m_license, e_no,))
            self.conn_mssql_erp.commit()

            return 'ok'

        except Exception as e:
            logging.info(f"\n<Error> update_erp_bom : {str(e)}\n")
            self.conn_mssql_erp.rollback()  

        finally:
            self.__disconnect_mssql_erp__()

    ###################################
    # hr_360_member_question
    ###################################
    def hr_360_member_question_member(self , c_name , item):
        
        self.__connect7_1_38__()

        try:
            
            #########
            #
            # self
            #
            #########
            if item == 'self':
                
                r_year = time.strftime("%Y" , time.localtime())
                d_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
                r_day  = time.strftime("%Y-%m-%d" , time.localtime())
                r_time = time.strftime("%H:%M:%S" , time.localtime())

                ########
                # 員工
                ########
                sql = (
                        "SELECT DISTINCT ps.c_name "
                        "FROM hr_360_person_setup ps "
                        "LEFT JOIN hr_360_submit_member_content smc "
                        "ON ps.c_name = smc.s_name AND smc.s_user = %s AND smc.c_year=%s "
                        "WHERE %s IN (ps.c_name, ps.c_manager, ps.c_peer1, ps.c_peer2, ps.c_subordinate1, ps.c_subordinate2) "
                        "AND smc.s_name IS NULL AND ps.lv='member'"
                )

                self.curr_7_1_38.execute(sql , (c_name , r_year , c_name ,))
                res = self.curr_7_1_38.fetchall()

                # 提取结果
                f_res = [row for row in res]

                return f_res

        except Exception as e:
            logging.info(f"\n<Error> hr_360_manager_question : {str(e)}\n")
        finally:
            self.__disconnect7_1_38__()

    ###############################
    # hr_360_member_question
    ###############################
    def hr_360_member_question(self , c_name , item):
        
        self.__connect7_1_38__()

        try:
            
            #########
            #
            # self
            #
            #########
            if item == 'self':
                
                r_year = time.strftime("%Y" , time.localtime())
                d_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
                r_day  = time.strftime("%Y-%m-%d" , time.localtime())
                r_time = time.strftime("%H:%M:%S" , time.localtime())

                ########
                # 員工
                ########
                sql = (
                        "SELECT DISTINCT ps.c_name "
                        "FROM hr_360_person_setup ps "
                        "LEFT JOIN hr_360_submit_member_content smc "
                        "ON ps.c_name = smc.s_name AND smc.s_user = %s AND smc.c_year=%s "
                        "WHERE %s IN (ps.c_name, ps.c_manager, ps.c_peer1, ps.c_peer2, ps.c_subordinate1, ps.c_subordinate2) "
                        "AND smc.s_name IS NULL AND ps.lv='member'"
                )

                self.curr_7_1_38.execute(sql , (c_name , r_year , c_name ,))
                res = self.curr_7_1_38.fetchall()

                # 提取结果
                f_res = [row for row in res]

                return f_res

        except Exception as e:
            logging.info(f"\n<Error> hr_360_member_question : {str(e)}\n")
        finally:
            self.__disconnect7_1_38__()

    ###############################
    # hr_360_manager_question
    ###############################
    def hr_360_manager_question(self , c_name , item):
        
        self.__connect7_1_38__()

        try:
            
            #########
            #
            # self
            #
            #########
            if item == 'self':
                
                r_year = time.strftime("%Y" , time.localtime())
                d_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
                r_day  = time.strftime("%Y-%m-%d" , time.localtime())
                r_time = time.strftime("%H:%M:%S" , time.localtime())

                ########
                # 主管
                ########
                sql = (
                        "SELECT DISTINCT ps.c_name "
                        "FROM hr_360_person_setup ps "
                        "LEFT JOIN hr_360_submit_manager_content smc "
                        "ON ps.c_name = smc.s_name AND smc.s_user = %s AND smc.c_year=%s "
                        "WHERE %s IN (ps.c_name, ps.c_manager, ps.c_peer1, ps.c_peer2, ps.c_subordinate1, ps.c_subordinate2) "
                        "AND smc.s_name IS NULL AND ps.lv='manager'"
                )

                self.curr_7_1_38.execute(sql , (c_name , r_year , c_name ,))
                res = self.curr_7_1_38.fetchall()

                # 提取结果
                f_res = [row for row in res]

                return f_res

        except Exception as e:
            logging.info(f"\n<Error> hr_360_manager_question : {str(e)}\n")
        finally:
            self.__disconnect7_1_38__()

    ###############################
    # ss2_order_form_record_list
    ###############################
    def ss2_order_form_record_list(self):
        
        self.__connect7_1_38__()

        try:
            
            sql = f"SELECT ss2_form , count(*) FROM `ss2_export_record` group by ss2_form order by c_d_time desc" 
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchall()

            if res is not None:
                return res

        except Exception as e:
            logging.info(f"\n<Error> ss2_order_form_record_list : {str(e)}\n")
        finally:
            self.__disconnect7_1_38__()

    ###########################
    # erp_query_device
    ###########################
    def erp_query_device(self , barcode):
        
        self.__connect_mssql_erp__()

        try:
            
            self.sql        = f"SELECT A.MB001 AS '固資編號', A.MB002 AS '固資名稱', A.MB016 AS '取得日期',B.MC003 AS '工號', D.MV002 AS '保管人姓名', B.MC002 AS '部門代號',C.ME002 AS '部門名稱', B.MC006 AS '放置地點'FROM ASTMB A, ASTMC B, CMSME C, CMSMV D WHERE A.MB001=B.MC001 AND C.ME001=B.MC002 AND D.MV001=B.MC003 AND MB001='{barcode}'" 
            self.curr_mssql_erp.execute(self.sql)
            self.res        = self.curr_mssql_erp.fetchall()

            if self.res is not None:
                return self.res

        except Exception as e:
            logging.info('\n<Error> erp_query_device : ' + str(e))
        finally:
            self.__disconnect_mssql_erp__()

    ##################################
    # search_bpm_employee_id
    ##################################
    def search_bpm_employee_id(self , login_id):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"select EmployeeName , LoginID from T_HR_Employee" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()
            
            for val in res:
                if str(val[1]).lower() == str(login_id).lower():
                    return val[0]
            
        except Exception as e:
            logging.info(f'\n<Error> search_bpm_employee_id : {str(e)}')
        finally:
            self.__disconnect_mssql__()

    ##################################
    # bpm_load_account_list_by_dep2
    ##################################
    def bpm_load_account_list_by_dep2(self , upper_dep2):
        
        self.__connect_mssql__()

        try:
            ### by_upper_department
            sql        = f"select EmployeeID , EmployeeName , JobRank , JobTitleName , EnterDate , EmployeeEnglishName  , loginID , Email from T_HR_Employee the where DepartmentID ='{upper_dep2}' order by JobRank desc" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()

            return res
            
        except Exception as e:
            logging.info('\n<Error> bpm_load_account_list_by_dep2 : ' + str(e))
        finally:
            self.__disconnect_mssql__()
   
    #################################
    # bpm_load_account_list_by_dep
    #################################
    def bpm_load_account_list_by_dep(self , upper_dep):
        
        self.__connect_mssql__()

        try:
            ### by_upper_department
            sql        = f"select DepartmentID , DepartmentName  from T_HR_Department where UpperDepartmentID = '{upper_dep}'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchall()

            final_res = []

            for val in res:
                sql2 = f"select EmployeeName , EmployeeEnglishName  from T_HR_Employee where DepartmentID='{upper_dep}'"
                self.curr_mssql.execute(sql2)
                res2 = self.curr_mssql.fetchone()

                sql3 = f"select count(*) from T_HR_Employee where DepartmentID='{val[0]}'"
                self.curr_mssql.execute(sql3)
                res3 = self.curr_mssql.fetchone()

                if res2:
                    if res3:
                        final_res.append((val[0] , val[1] , res2[0] , res2[1] , res3[0]))

            return final_res
            
        except Exception as e:
            logging.info('\n<Error> bpm_load_account_list_by_dep : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    #############################################
    # modify_ad_pwd
    #############################################
    def modify_ad_pwd(self , account):

        ### variables
        
        
        r_year = time.strftime("%Y" , time.localtime())
        d_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
        r_day  = time.strftime("%Y-%m-%d" , time.localtime())
        r_time = time.strftime("%H:%M:%S" , time.localtime())

        self.__connect__()

        try:   
            # AD服务器的地址
            server_address = '192.168.1.10'
            # AD管理员账户的用户名和密码
            admin_username = 'CN=admin_username,CN=Users,DC=yourdomain,DC=com'
            admin_password = 'admin_password'
            # 用户的完整DN和新密码
            user_dn = 'CN=target_user,CN=Users,DC=yourdomain,DC=com'
            new_password = 'new_password'

            # 创建 Server 对象
            server = Server(server_address, use_ssl=True, get_info=ALL)

            try:
                # 连接到服务器
                with Connection(server, admin_username, admin_password, auto_bind=True) as conn:
                    # 更改密码
                    # 注意：新密码通常需要用引号括起来并且是UTF-16 LE编码
                    password_value = ('"%s"' % new_password).encode('utf-16-le')
                    conn.modify(user_dn, {'userPassword': [(MODIFY_REPLACE, [password_value])]})
                    
                    # 检查操作是否成功
                    if conn.result['result'] == 0:
                        print("Password changed successfully")
                    else:
                        print("Failed to change password:", conn.result['description'])
            except Exception as e:
                print("An error occurred:", str(e)) 
                
            
        except Exception as e:
            logging.info(f'\n<Error> modify_ad_pwd : {str(e)}\n')
        finally:
            self.__disconnect__()


    #############################################
    # mrd_8_query_announcement_auto_search_res
    #############################################
    def mrd_8_query_announcement_auto_search_res(self):

        ### variables
        url = parameter.mrd_8['衛生福利部食品藥物管理署公告']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}
        
        r_year = time.strftime("%Y" , time.localtime())
        d_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
        r_day  = time.strftime("%Y-%m-%d" , time.localtime())
        r_time = time.strftime("%H:%M:%S" , time.localtime())

        self.__connect__()

        try:    
                y = str(r_year).replace(" " , "+")

                url = f"{url}&y={y}&key=&scid="
                res = requests.get(url , headers=headers , allow_redirects=False)
            
                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')

                    table = soup.find('table', class_='listTable')
                    #print("Found table:", table)  

                    table_data = []

                    rows = table.find_all('tr') if table else []
                    #print(f"Found {len(rows)} rows in the table.")  

                    for index, row in enumerate(rows):
                        cells = row.find_all('td')
                        cells = cells[:3]  
                        #print(f"Row {index} has {len(cells)} cells.")  

                        if cells:
                            data_row = []
                            for cell in cells:
                                text = cell.text.strip()
                                link = cell.find('a')
                                if link and link.has_attr('href'):
                                    data_row.append((text, link['href']))
                                else:
                                    data_row.append((text))
                            
                            table_data.append(data_row)
                    
                    for data_row in table_data:

                        data      = str(data_row[1]).split(',')
                        data2     = str(data[0]).split('\'')
                        data3     = str(data[1]).split('\'')
                        b_from    = 'FDA'
                        b_title   = str(data2[1]).strip()
                        b_url     = str('https://www.fda.gov.tw/TC/' + data3[1]).strip()
                        b_release = str(data_row[2]).strip()
                        sum = 0
                        
                        s_sql = f"select b_title from mrd_8_government_bulletin_release where b_title='{b_title}'"
                        self.curr.execute(s_sql)
                        s_res = self.curr.fetchone()

                        if s_res is None:
                            sum+=1
                            sql  = f"INSERT INTO mrd_8_government_bulletin_release (c_date , c_time , c_d_time , b_from , b_title , b_url , b_release) "
                            sql += f"VALUES ('{str(r_day)}','{str(r_time)}','{str(d_time)}','{b_from}','{b_title}','{b_url}','{b_release}')"
                            self.curr.execute(sql)
                            self.conn.commit()

                            if sum > 0:
                                ### 顯示新增公告
                                print(f"\n********************************************************************************************************************************************")
                                logging.info(f"< FDA 衛生福利部食品藥物管理署公告 > \n\t 公告日期 : {b_release} \n\t 新增公告 : {b_title}\n\t 公告網址 : {b_url}\n")

                            #############################
                            #
                            # FDA 新公告自動寄 Email 通知
                            #
                            #############################
                            email_sql = f"select email from mrd_8_auto_email_push"
                            self.curr.execute(email_sql)
                            push_email = self.curr.fetchall()

                            for val in push_email:
                                ### send email parameter -> (subject , email , title , url , b_release)
                                self.send_email('< FDA 衛生福利部食品藥物管理署公告 >', str(val[0]) , b_title , b_url , b_release)

                            

                else:
                    tb_content = f'<Error> mrd_8_query_announcement_search_res'
            
        except Exception as e:
            logging.info(f'\n<Error> mrd_8_query_announcement_auto_search_res : {str(e)}\n')
        finally:
            self.__disconnect__()

    #######################
    # send_email
    #######################
    def send_email(self , subject , email , title , url , b_release):
        try:
            ### variable
            r_time           = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime())
            send_mail        = email
            send_msg_content = f"{subject} \n 公告日期 : {b_release} \n 新增公告 : {title} \n 公告網址 : {url} \n (系統自動發信，勿回覆)"
            
            ############
            # Email
            ############
            smtpobj = smtplib.SMTP('smtp.office365.com' , 587)
            ### start smtp 
            smtpobj.ehlo() 
            ### use TLS transmission   
            smtpobj.starttls()

            smtpobj.login('Jason_Hung@otsuka.com.tw','otsuka#123')
            
            msg = MIMEMultipart()
            msg["subject"] = subject
            msg["from"]    = 'Jason_Hung@otsuka.com.tw'
            msg["to"]      = send_mail
            msg.attach(MIMEText(send_msg_content))
            smtpobj.send_message(msg=msg)

            ### send msg     
            logging.info(f'{subject} 新公告 , 通知 : {send_mail} , 成功')
            
            smtpobj.quit()
            
        except smtplib.SMTPException as e:
            logging.info(f"\n<Error> send mail : {str(e)}\n")
        finally:
            pass
    
    ########################################
    # mrd_8_query_announcement_search_res
    ########################################
    def mrd_8_query_announcement_search_res(self , mrd_8_query_announcement_year , mrd_8_query_announcement_item , mrd_8_query_announcement_keyword):

        ### variables
        url = parameter.mrd_8['衛生福利部食品藥物管理署公告']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                y    = str(mrd_8_query_announcement_year).replace(" " , "+")
                scid = str(mrd_8_query_announcement_item).replace(" " , "+")
                key  = str(mrd_8_query_announcement_keyword).replace(" " , "+")

                url = f"{url}&y={y}&key={key}&scid={scid}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('table' , class_='listTable')

                else:
                    
                    tb_content = f'<Error> mrd_8_query_announcement_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_query_announcement_search_res : ' + str(e))
        finally:
            pass

    ###########################################################
    # mrd_8_cosmetic_query_ingredients_prohibited_search_res
    ###########################################################
    def mrd_8_cosmetic_query_ingredients_prohibited_search_res(self , mrd_8_cosmetic_query_ingredients_prohibited_k):

        ### variables
        url = parameter.mrd_8['化粧品禁限用成分管理規定']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k  = str(mrd_8_cosmetic_query_ingredients_prohibited_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_cosmetic_query_ingredients_prohibited_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_cosmetic_query_ingredients_prohibited_search_res : ' + str(e))
        finally:
            pass

    #####################################################
    # mrd_8_food_query_infant_formula_all_search_res
    #####################################################
    def mrd_8_food_query_infant_formula_all_search_res(self , mrd_8_food_query_infant_formula_k):

        ### variables
        url = parameter.mrd_8['嬰兒與較大嬰兒配方食品許可資料查詢']

        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k   = str(mrd_8_food_query_infant_formula_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_infant_formula_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_infant_formula_search_res : ' + str(e))
        finally:
            pass

    #####################################################
    # mrd_8_food_query_infant_formula_search_res
    #####################################################
    def mrd_8_food_query_infant_formula_search_res(self , mrd_8_food_query_infant_formula_ct , mrd_8_food_query_infant_formula_cn , mrd_8_food_query_infant_formula_en , mrd_8_food_query_infant_formula_cp , mrd_8_food_query_infant_formula_ph1 , mrd_8_food_query_infant_formula_ph2 , mrd_8_food_query_infant_formula_ph3 , mrd_8_food_query_infant_formula_k):

        ### variables
        url = parameter.mrd_8['嬰兒與較大嬰兒配方食品許可資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                ct  = str(mrd_8_food_query_infant_formula_ct).replace(" " , "+")
                cn  = str(mrd_8_food_query_infant_formula_cn).replace(" " , "+")
                en  = str(mrd_8_food_query_infant_formula_en).replace(" " , "+")
                cp  = str(mrd_8_food_query_infant_formula_cp).replace(" " , "+")
                ph1 = str(mrd_8_food_query_infant_formula_ph1).replace(" " , "+")
                ph2 = str(mrd_8_food_query_infant_formula_ph2).replace(" " , "+")
                ph3 = str(mrd_8_food_query_infant_formula_ph3).replace(" " , "+")
                k   = str(mrd_8_food_query_infant_formula_k).replace(" " , "+")

                url = f"{url}&ct={ct}&cn={cn}&en={en}&cp={cp}&ph1={ph1}&ph2={ph2}&ph3={ph3}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_infant_formula_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_infant_formula_search_res : ' + str(e))
        finally:
            pass

    #########################################################
    # mrd_8_food_query_genetic_modification_all_search_res
    #########################################################
    def mrd_8_food_query_genetic_modification_all_search_res(self , mrd_8_food_query_genetic_modification_k):

        ### variables
        url = parameter.mrd_8['衛生福利部審核通過之基因改造食品原料之查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k  = str(mrd_8_food_query_genetic_modification_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_genetic_modification_all_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_genetic_modification_all_search_res : ' + str(e))
        finally:
            pass


    #####################################################
    # mrd_8_food_query_genetic_modification_search_res
    #####################################################
    def mrd_8_food_query_genetic_modification_search_res(self , mrd_8_food_query_genetic_modification_t , mrd_8_food_query_genetic_modification_t2 , mrd_8_food_query_genetic_modification_pn , mrd_8_food_query_genetic_modification_an , mrd_8_food_query_genetic_modification_sd , mrd_8_food_query_genetic_modification_ed , mrd_8_food_query_genetic_modification_k):

        ### variables
        url = parameter.mrd_8['衛生福利部審核通過之基因改造食品原料之查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                t  = str(mrd_8_food_query_genetic_modification_t).replace(" " , "+")
                t2 = str(mrd_8_food_query_genetic_modification_t2).replace(" " , "+")
                pn = str(mrd_8_food_query_genetic_modification_pn).replace(" " , "+")
                an = str(mrd_8_food_query_genetic_modification_an).replace(" " , "+")
                sd = str(mrd_8_food_query_genetic_modification_sd).replace(" " , "+")
                ed = str(mrd_8_food_query_genetic_modification_ed).replace(" " , "+")
                k  = str(mrd_8_food_query_genetic_modification_k).replace(" " , "+")

                sd_year  = str(mrd_8_food_query_genetic_modification_sd)[0:4]
                sd_month = str(mrd_8_food_query_genetic_modification_sd)[5:7]
                sd_day   = str(mrd_8_food_query_genetic_modification_sd)[8:10]
                sd2 = sd_year + '年' + sd_month + '月' + sd_day + '日'

                ed_year  = str(mrd_8_food_query_genetic_modification_ed)[0:4]
                ed_month = str(mrd_8_food_query_genetic_modification_ed)[5:7]
                ed_day   = str(mrd_8_food_query_genetic_modification_ed)[8:10]
                ed2 = ed_year + '年' + ed_month + '月' + ed_day + '日'

                url = f"{url}&t={t}&t2={t2}&pn={pn}&an={an}&sd={sd2}&ed={ed2}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_genetic_modification_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_genetic_modification_search_res : ' + str(e))
        finally:
            pass

    ######################################################
    # mrd_8_food_query_domestic_vitamins_all_search_res
    ######################################################
    def mrd_8_food_query_domestic_vitamins_all_search_res(self , mrd_8_food_query_domestic_vitamins_k):

        ### variables
        url = parameter.mrd_8['國產維生素類錠狀膠囊狀食品查驗登記證資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k   = str(mrd_8_food_query_domestic_vitamins_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_domestic_vitamins_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_domestic_vitamins_search_res : ' + str(e))
        finally:
            pass

    ##################################################
    # mrd_8_food_query_domestic_vitamins_search_res
    ##################################################
    def mrd_8_food_query_domestic_vitamins_search_res(self , mrd_8_food_query_domestic_vitamins_ct , mrd_8_food_query_domestic_vitamins_cn , mrd_8_food_query_domestic_vitamins_en , mrd_8_food_query_domestic_vitamins_cp , mrd_8_food_query_domestic_vitamins_ph1 , mrd_8_food_query_domestic_vitamins_ph2 , mrd_8_food_query_domestic_vitamins_ph3 , mrd_8_food_query_domestic_vitamins_k):

        ### variables
        url = parameter.mrd_8['國產維生素類錠狀膠囊狀食品查驗登記證資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                ct  = str(mrd_8_food_query_domestic_vitamins_ct).replace(" " , "+")
                cn  = str(mrd_8_food_query_domestic_vitamins_cn).replace(" " , "+")
                en  = str(mrd_8_food_query_domestic_vitamins_en).replace(" " , "+")
                cp  = str(mrd_8_food_query_domestic_vitamins_cp).replace(" " , "+")
                ph1 = str(mrd_8_food_query_domestic_vitamins_ph1).replace(" " , "+")
                ph2 = str(mrd_8_food_query_domestic_vitamins_ph2).replace(" " , "+")
                ph3 = str(mrd_8_food_query_domestic_vitamins_ph3).replace(" " , "+")
                k   = str(mrd_8_food_query_domestic_vitamins_k).replace(" " , "+")

                url = f"{url}&ct={ct}&cn={cn}&en={en}&cp={cp}&ph1={ph1}&ph2={ph2}&ph3={ph3}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_domestic_vitamins_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_domestic_vitamins_search_res : ' + str(e))
        finally:
            pass

    ##################################################
    # mrd_8_food_query_enter_capsule_all_search_res
    ##################################################
    def mrd_8_food_query_enter_capsule_all_search_res(self , mrd_8_food_query_enter_capsule_k):

        ### variables
        url = parameter.mrd_8['輸入膠囊錠狀食品核備查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k   = str(mrd_8_food_query_enter_capsule_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_enter_capsule_all_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_enter_capsule_all_search_res : ' + str(e))
        finally:
            pass

    ###############################################
    # mrd_8_food_query_enter_capsule_search_res
    ###############################################
    def mrd_8_food_query_enter_capsule_search_res(self , mrd_8_food_query_enter_capsule_ct , mrd_8_food_query_enter_capsule_cn , mrd_8_food_query_enter_capsule_en , mrd_8_food_query_enter_capsule_cp , mrd_8_food_query_enter_capsule_ph1 , mrd_8_food_query_enter_capsule_ph2 , mrd_8_food_query_enter_capsule_ph3 , mrd_8_food_query_enter_capsule_k):

        ### variables
        url = parameter.mrd_8['輸入膠囊錠狀食品核備查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                ct  = str(mrd_8_food_query_enter_capsule_ct).replace(" " , "+")
                cn  = str(mrd_8_food_query_enter_capsule_cn).replace(" " , "+")
                en  = str(mrd_8_food_query_enter_capsule_en).replace(" " , "+")
                cp  = str(mrd_8_food_query_enter_capsule_cp).replace(" " , "+")
                ph1 = str(mrd_8_food_query_enter_capsule_ph1).replace(" " , "+")
                ph2 = str(mrd_8_food_query_enter_capsule_ph2).replace(" " , "+")
                ph3 = str(mrd_8_food_query_enter_capsule_ph3).replace(" " , "+")
                k   = str(mrd_8_food_query_enter_capsule_k).replace(" " , "+")

                url = f"{url}&ct={ct}&cn={cn}&en={en}&cp={cp}&ph1={ph1}&ph2={ph2}&ph3={ph3}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_enter_capsule_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_enter_capsule_search_res : ' + str(e))
        finally:
            pass

    ###################################################
    # mrd_8_food_query_disease_recipe_all_search_res
    ###################################################
    def mrd_8_food_query_disease_recipe_all_search_res(self , mrd_8_food_query_disease_recipe_k):

        ### variables
        url = parameter.mrd_8['特定疾病配方食品']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k   = str(mrd_8_food_query_disease_recipe_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_disease_recipe_all_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_disease_recipe_all_search_res : ' + str(e))
        finally:
            pass

    ###############################################
    # mrd_8_food_query_disease_recipe_search_res
    ###############################################
    def mrd_8_food_query_disease_recipe_search_res(self , mrd_8_food_query_disease_recipe_t , mrd_8_food_query_disease_recipe_ct , mrd_8_food_query_disease_recipe_cn , mrd_8_food_query_disease_recipe_en , mrd_8_food_query_disease_recipe_cp , mrd_8_food_query_disease_recipe_ph1 , mrd_8_food_query_disease_recipe_ph2 , mrd_8_food_query_disease_recipe_ph3 , mrd_8_food_query_disease_recipe_k):

        ### variables
        url = parameter.mrd_8['特定疾病配方食品']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                t   = str(mrd_8_food_query_disease_recipe_t).replace(" " , "+")
                ct  = str(mrd_8_food_query_disease_recipe_ct).replace(" " , "+")
                cn  = str(mrd_8_food_query_disease_recipe_cn).replace(" " , "+")
                en  = str(mrd_8_food_query_disease_recipe_en).replace(" " , "+")
                cp  = str(mrd_8_food_query_disease_recipe_cp).replace(" " , "+")
                ph1 = str(mrd_8_food_query_disease_recipe_ph1).replace(" " , "+")
                ph2 = str(mrd_8_food_query_disease_recipe_ph2).replace(" " , "+")
                ph3 = str(mrd_8_food_query_disease_recipe_ph3).replace(" " , "+")
                k   = str(mrd_8_food_query_disease_recipe_k).replace(" " , "+")

                url = f"{url}&t={t}&ct={ct}&cn={cn}&en={en}&cp={cp}&ph1={ph1}&ph2={ph2}&ph3={ph3}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_disease_recipe_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_disease_recipe_search_res : ' + str(e))
        finally:
            pass

    #########################################
    # mrd_8_food_query_pass_all_search_res
    #########################################
    def mrd_8_food_query_pass_all_search_res(self , mrd_8_food_query_pass_k):

        ### variables
        url = parameter.mrd_8['衛生福利部審核通過之健康食品資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k   = str(mrd_8_food_query_pass_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_pass_all_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_pass_all_search_res : ' + str(e))
        finally:
            pass

    #########################################
    # mrd_8_food_query_pass_search_res
    #########################################
    def mrd_8_food_query_pass_search_res(self , mrd_8_food_query_pass_t , mrd_8_food_query_pass_Tid , mrd_8_food_query_pass_Cop , mrd_8_food_query_pass_Cna , mrd_8_food_query_pass_t2 , mrd_8_food_query_pass_k):

        ### variables
        url = parameter.mrd_8['衛生福利部審核通過之健康食品資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                t   = str(mrd_8_food_query_pass_t).replace(" " , "+")
                Tid = str(mrd_8_food_query_pass_Tid).replace(" " , "+")
                Cop = str(mrd_8_food_query_pass_Cop).replace(" " , "+")
                Cna = str(mrd_8_food_query_pass_Cna).replace(" " , "+")
                t2  = str(mrd_8_food_query_pass_t2).replace(" " , "+")
                k   = str(mrd_8_food_query_pass_k).replace(" " , "+")

                url = f"{url}&t={t}&Tid={Tid}&Cop={Cop}&Cna={Cna}&t2={t2}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_pass_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_pass_search_res : ' + str(e))
        finally:
            pass

    #############################################
    # mrd_8_food_query_platform_all_search_res
    #############################################
    def mrd_8_food_query_platform_all_search_res(self , mrd_8_food_query_platform_k):

        ### variables
        url = parameter.mrd_8['食品原料整合查詢平臺']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k  = str(mrd_8_food_query_platform_k).replace(" " , "+")

                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_platform_all_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_platform_all_search_res : ' + str(e))
        finally:
            pass

    #########################################
    # mrd_8_food_query_platform_search_res
    #########################################
    def mrd_8_food_query_platform_search_res(self , mrd_8_food_query_platform_c , mrd_8_food_query_platform_t , mrd_8_food_query_platform_k):

        ### variables
        url = parameter.mrd_8['食品原料整合查詢平臺']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                c  = str(mrd_8_food_query_platform_c).replace(" " , "+")
                t  = str(mrd_8_food_query_platform_t).replace(" " , "+")
                k  = str(mrd_8_food_query_platform_k).replace(" " , "+")

                url = f"{url}&c={c}&t={t}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_food_query_platform_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_query_platform_search_res : ' + str(e))
        finally:
            pass


    #################################
    # mrd_8_license_all_search_res
    #################################
    def mrd_8_license_all_search_res(self , mrd_8_license_k):

        ### variables
        url = parameter.mrd_8['食品添加物許可證資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                k   = str(mrd_8_license_k).replace(" " , "+")


                url = f"{url}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_license_all_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_license_all_search_res : ' + str(e))
        finally:
            pass

    ############################
    # mrd_8_license_search_res
    ############################
    def mrd_8_license_search_res(self , mrd_8_license_ct , mrd_8_license_cn , mrd_8_license_en , mrd_8_license_cp , mrd_8_license_ph1 , mrd_8_license_ph2 , mrd_8_license_ph3 , mrd_8_license_k):

        ### variables
        url = parameter.mrd_8['食品添加物許可證資料查詢']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
                ct  = str(mrd_8_license_ct).replace(" " , "+")
                cn  = str(mrd_8_license_cn).replace(" " , "+")
                en  = str(mrd_8_license_en).replace(" " , "+")
                cp  = str(mrd_8_license_cp).replace(" " , "+")
                ph1 = str(mrd_8_license_ph1).replace(" " , "+")
                ph2 = str(mrd_8_license_ph2).replace(" " , "+")
                ph3 = str(mrd_8_license_ph3).replace(" " , "+")
                k   = str(mrd_8_license_k).replace(" " , "+")


                url = f"{url}&ct={ct}&cn={cn}&en={en}&cp={cp}&ph1={ph1}&ph2={ph2}&ph3={ph3}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

                if res.status_code == 200:

                    soup = BeautifulSoup(res.text, 'html.parser')
                    tb_content = soup.find_all('div' , class_='tabs')

                else:
                    
                    tb_content = f'<Error> mrd_8_license_search_res'
                
                return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_license_search_res : ' + str(e))
        finally:
            pass


    ##############################
    # mrd_8_food_all_search_res
    ##############################
    def mrd_8_food_all_search_res(self , mrd_8_food_k):

        ### variables
        url = parameter.mrd_8['食品添加物使用範圍及限量暨規格標準']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
            
            k   = str(mrd_8_food_k).replace(" " , "+")
            url = f"{url}&k={k}"
            res = requests.get(url , headers=headers , allow_redirects=False)

            if res.status_code == 200:

                soup = BeautifulSoup(res.text, 'html.parser')
                tb_content = soup.find_all('div' , class_='tabs')

            else:
                
                tb_content = f'<Error> mrd_8_food_all_search_res'
            
            return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_all_search_res : ' + str(e))
        finally:
            pass

    ############################
    # mrd_8_food_search_res
    ############################
    def mrd_8_food_search_res(self , mrd_8_food_t , mrd_8_food_k):

        ### variables
        url = parameter.mrd_8['食品添加物使用範圍及限量暨規格標準']
        
        headers={"User-Agent" : "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.6) ",
                "Accept" : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language" : "en-us",
                "Connection" : "keep-alive",
                "Accept-Charset" : "GB2312,utf-8;q=0.7,*;q=0.7"}

        try:
            if len(str(mrd_8_food_t)) == 0 and len(str(mrd_8_food_k)) == 0:
                
                res    = requests.get(url , headers=headers , allow_redirects=False)

            elif len(str(mrd_8_food_t)) == 0 and len(str(mrd_8_food_k)) != 0:
                
                k = str(mrd_8_food_k).replace(" " , "+")
                url = f"{url}&k={k}"
                res = requests.get( url , headers=headers , allow_redirects=False)

            else:
                
                t   = str(mrd_8_food_t).replace(" " , "+")
                k   = str(mrd_8_food_k).replace(" " , "+")
                url = f"{url}&t={t}&k={k}"
                res = requests.get(url , headers=headers , allow_redirects=False)

            if res.status_code == 200:

                soup = BeautifulSoup(res.text, 'html.parser')
                tb_content = soup.find_all('div' , class_='tabs')

            else:
                
                tb_content = f'<Error> mrd_8_food_search_res'
            
            return tb_content
            
        except Exception as e:
            logging.info('\n<Error> mrd_8_food_search_res : ' + str(e))
        finally:
            pass

    ###################################
    # bpm_account_list_by_dep_search
    ###################################
    def bpm_account_list_by_dep_search(self , query_account):
        
        self.__connect_mssql__()

        try:
            ### by_upper_department
            sql = f"SELECT DepartmentID from T_HR_Employee where loginID='{query_account}'"
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchone()

            sql2 = f"select DepartmentName from T_HR_Department where DepartmentID='{res[0]}'"
            self.curr_mssql.execute(sql2)
            res2 = self.curr_mssql.fetchone()

            sql3 = f"select EmployeeID , EmployeeName , Email , EnterDate from T_HR_Employee where loginID like '{query_account}%'"
            self.curr_mssql.execute(sql3)
            res3 = self.curr_mssql.fetchall()

            final_res = []

            for val in res3:
                final_res.append((val[0] , val[1] , val[2] , val[3] , res2[0]))
            
            return final_res
            
        except Exception as e:
            logging.info('\n<Error> bpm_account_list_by_dep_search : ' + str(e))
        finally:
            self.__disconnect_mssql__()
    
    ############################
    # bpm_account_list_by_dep
    ############################
    def bpm_account_list_by_dep(self):
        
        self.__connect_mssql__()

        try:
            ### by_upper_department
            sql = f"SELECT DepartmentID , DepartmentName FROM T_HR_Department WHERE DepartmentID IN (SELECT DISTINCT UpperDepartmentID FROM T_HR_Department)"
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()
            
            return res
            
        except Exception as e:
            logging.info('\n<Error> bpm_account_list_by_dep : ' + str(e))
        finally:
            self.__disconnect_mssql__()
    
    ###########################
    # bpm_account_search
    ###########################
    def bpm_account_search(self , item , login_id):
        
        self.__connect_mssql__()

        try:
            
            self.sql        = f"select {item} from T_HR_Employee where loginid='{login_id}'" 
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]

        except Exception as e:
            logging.info('\n<Error> bpm_account_search : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    #################################
    # search_bpm_account_item_list
    #################################
    def search_bpm_account_item_list(self , name):
        
        self.__connect_mssql__()

        try:
            
            sql = (
                    f"select Email from T_HR_Employee where EmployeeName=?"
            ) 
            self.curr_mssql.execute(sql , (name,))
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]

        except Exception as e:
            logging.info('\n<Error> search_bpm_account_item_list : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    ###########################
    # bpm_account_list
    ###########################
    def bpm_account_list(self):
        
        self.__connect_mssql__()

        try:
            
            self.sql        = f"select a.EmployeeID  , a.EmployeeName , a.loginID , a.DepartmentID , b.DepartmentName , b.UpperDepartmentID , a.Email , a.EnterDate  from T_HR_Employee a left join T_HR_Department b on a.DepartmentID = b.DepartmentID  order by a.DepartmentID asc" 
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchall()

            return self.res

        except Exception as e:
            logging.info('\n<Error> bpm_account_list : ' + str(e))
        finally:
            self.__disconnect_mssql__()
    
    ###########################
    # bpm_account_data
    ###########################
    def bpm_account_data(self , user , item):
        
        self.__connect_mssql__()

        try:
            
            sql        = f"SELECT {item} FROM T_HR_Employee the where loginID='{user}'" 
            self.curr_mssql.execute(sql)
            res        = self.curr_mssql.fetchone()

            return res[0]

        except Exception as e:
            logging.info('\n<Error> bpm_account_data : ' + str(e))
        finally:
            self.__disconnect_mssql__()
    
    ###########################
    # bpm_account_department
    ###########################
    def bpm_account_department(self , user , item):
        
        self.__connect_mssql__()

        try:
            
            self.sql        = f"select {item} from T_HR_Department a where DepartmentID = (select DepartmentID from T_HR_Employee where loginID='{user}')" 
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]

        except Exception as e:
            logging.info('\n<Error> bpm_account_department : ' + str(e))
        finally:
            self.__disconnect_mssql__()

    

    ########################
    # erp_hr_account_list
    ########################
    def erp_hr_account_list(self):
        try:
            ######################
            #
            # select from MSSQL
            #
            ######################
            if sys.platform.startswith('win'):
                conn_str = f"DRIVER={{SQL Server}};SERVER={parameter.otsuka_factory3['host']};DATABASE={parameter.otsuka_factory3['db']};UID={parameter.otsuka_factory3['user']};PWD={parameter.otsuka_factory3['pwd']}"  
            elif sys.platform.startswith('darwin'):
                conn_str = f"DRIVER={{/opt/homebrew/Cellar/msodbcsql17/17.10.5.1/lib/libmsodbcsql.17.dylib}};SERVER={parameter.otsuka_factory3['host']};DATABASE={parameter.otsuka_factory3['db']};UID={parameter.otsuka_factory3['user']};PWD={parameter.otsuka_factory3['pwd']}"  
            
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"SELECT a.EMPID AS EmployeeID, CASE (isnull(a.HECNAME, '')) WHEN '' THEN '' ELSE a.HECNAME END AS EmployeeName, CASE (isnull(a.HEENAME, '')) WHEN '' THEN '' ELSE a.HEENAME END AS EmployeeEnglishName, CASE (isnull(a.LOGIN_ID, '')) WHEN '' THEN '' ELSE a.LOGIN_ID END AS LoginID, a.CPNYID AS CompanyID, a.DEPT_NO AS DepartmentID, '' AS IdentityID, a.SEX, CASE (isnull(a.EMAIL, '')) WHEN '' THEN '' ELSE a.EMAIL END AS Email, CASE (isnull(a.MOBILNO, '')) WHEN '' THEN '' ELSE a.MOBILNO END AS Mobile, SUBSTRING(a.BIRTHDAY, 1, 4) + '/' + SUBSTRING(a.BIRTHDAY, 5, 2) + '/' + SUBSTRING(a.BIRTHDAY, 5, 2) AS Birthday, a.POSSIE AS JobTitleCode, CASE (isnull(b.POS_NAME, '')) WHEN '' THEN '' ELSE b.POS_NAME END AS JobTitleName, CASE (isnull(a.GRADE, '')) WHEN '' THEN '' ELSE a.GRADE END AS JobGrade, CASE (isnull(a.RANK, '')) WHEN '' THEN '' ELSE a.RANK END AS JobRank, '' AS JobCode, '' AS JobType, SUBSTRING(a.INADATE, 1, 4) + '/' + SUBSTRING(a.INADATE, 5, 2) + '/' + SUBSTRING(a.INADATE, 5, 2) AS EnterDate, CASE (isnull(a.PLACE, '')) WHEN '' THEN '' ELSE a.PLACE END AS WorkPlace, '' AS AreaCode, CASE (isnull(a.MOBILNO, '')) WHEN '' THEN '' ELSE a.MOBILNO END AS HomePhone, CASE (isnull(a.EXT, '')) WHEN '' THEN '' ELSE a.EXT END AS OfficePhone, CASE (isnull(a.COMADDR, '')) WHEN '' THEN '' ELSE a.COMADDR END AS Address, '' AS Synopsis FROM dbo.HRUSER AS a LEFT OUTER JOIN dbo.POSITION AS b ON a.POSSIE = b.POSSIE where a.STATE='A'" 
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchall()

            self.__connect__()
            for val in self.res:
                
                s_dep_code_sql = f"select DEP_CODE , DEP_SHORT_NAME from HRUSER_DEPT_BAS where DEP_NO='{val[5]}'"
                self.curr_mssql.execute(s_dep_code_sql)
                res_dep_code = self.curr_mssql.fetchall()

                for dep_val in res_dep_code:

                    ###########################
                    #
                    # check MsSQL hr account 
                    #
                    ###########################
                    s_sql = f"select employee_name from hr_a where employee_name='{val[1]}'"
                    self.curr.execute(s_sql)
                    s_r = self.curr.fetchone()

                    if s_r is None:
                        ######################
                        #
                        # insert into MySQL
                        #
                        ######################
                        sql  = f"insert into hr_a(employee_id , employee_name , employee_eng_name , login_id , company_id , department_id , identity_id , sex , email , mobile , birthday , job_title_code , job_title_name , job_grade , job_rank , job_code , job_type , end_date , work_place , area_code , home_phone , office_phone , addresses , department_code , department_name) "
                        sql += f"value('{val[0]}','{val[1]}','{val[2]}','{val[3]}','{val[4]}','{val[5]}','{val[6]}','{val[7]}','{val[8]}','{val[9]}','{val[10]}','{val[11]}','{val[12]}','{val[13]}','{val[14]}','{val[15]}','{val[16]}','{val[17]}','{val[18]}','{val[19]}','{val[20]}','{val[21]}','{val[22]}','{dep_val[0]}','{dep_val[1]}')"
                        self.curr.execute(sql)
                        self.conn.commit()
                    else:
                        pass
                        #logging.info(f"{s_r[0]} 已存在.")
            
            print('\n')
            logging.info(f"< Msg > HR account , 更新完成.")
            
            self.__disconnect__()

            return self.res
        
        except Exception as e:
            logging.info('\n<Error> erp_hr_account_list : ' + str(e))

        finally:
            self.curr_mssql.close()
            self.conn_mssql.close()

    ############################
    # department_account_list
    ############################
    def department_account_list(self):
        
        self.__connect__()

        try:
            
            sql  = f"SELECT department_name , department_code , department_id , count(*) FROM `hr_a` group by department_code order by department_name desc" 

            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
            
        except Exception as e:
            logging.info('\n<Error> department_account_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    ###########################
    # department_list_detail
    ###########################
    def department_list_detail(self , d_code):
        
        self.__connect__()
        
        try:
            ### connect mysql
            connect_sql = f"select department_name , employee_id , employee_name , department_code from hr_a where department_code='{d_code}' and login_id!='disabled' order by department_name asc"
            self.curr.execute(connect_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
            
        except Exception as e:
            logging.info('\n<Error> department_list_detail : ' + str(e))

        finally:
            self.__disconnect__()

    #############################
    # department_no_search_val
    #############################
    def department_no_search_val(self , employee_name):
        
        self.__connect__()
        
        try:
            ### connect mysql
            connect_sql = f"select department_name , department_code from hr_a where employee_name='{employee_name}'"
            self.curr.execute(connect_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res

            ### connect mssql
            '''
            conn_str        = f"DRIVER={{SQL Server}};SERVER={parameter.otsuka_factory3['host']};DATABASE={parameter.otsuka_factory3['db']};UID={parameter.otsuka_factory3['user']};PWD={parameter.otsuka_factory3['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"SELECT B.DEP_CODE FROM HR_Employee A , HRUSER_DEPT_BAS B WHERE A.DepartmentID = B.DEP_NO AND A.EmployeeName='{employee_name}'"
            self.curr_mssql.execute(self.sql)
            
            self.res        = self.curr_mssql.fetchone()
            self.curr_mssql.commit()
            
            return self.res[0]

            self.curr_mssql.close()
            self.conn_mssql.close()
            '''
            
        except Exception as e:
            logging.info('\n<Error> department_no_search_vals : ' + str(e))

        finally:
            self.__disconnect__()

    #############################
    # load_work_time_data_list
    #############################
    def load_work_time_data_list(self , e_id , e_name , b_date):
        
        self.__connect__()

        try:
            sql  = f"select normal_time , over_time , availability_time , total_time , b_date , " 
            sql += f"w_s_1 , w_s_1_product , w_s_1_num , w_s_1_normal_time , w_s_1_over_time , w_s_1_avail_time , w_s_1_remark ," 
            sql += f"w_s_2 , w_s_2_product , w_s_2_num , w_s_2_normal_time , w_s_2_over_time , w_s_2_avail_time , w_s_2_remark ," 
            sql += f"w_s_3 , w_s_3_product , w_s_3_num , w_s_3_normal_time , w_s_3_over_time , w_s_3_avail_time , w_s_3_remark ," 
            sql += f"w_s_4 , w_s_4_product , w_s_4_num , w_s_4_normal_time , w_s_4_over_time , w_s_4_avail_time , w_s_4_remark ," 
            sql += f"w_s_5 , w_s_5_product , w_s_5_num , w_s_5_normal_time , w_s_5_over_time , w_s_5_avail_time , w_s_5_remark ," 
            sql += f"w_s_6 , w_s_6_product , w_s_6_num , w_s_6_normal_time , w_s_6_over_time , w_s_6_avail_time , w_s_6_remark ," 
            sql += f"w_s_7 , w_s_7_product , w_s_7_num , w_s_7_normal_time , w_s_7_over_time , w_s_7_avail_time , w_s_7_remark ," 
            sql += f"w_s_8 , w_s_8_product , w_s_8_num , w_s_8_normal_time , w_s_8_over_time , w_s_8_avail_time , w_s_8_remark ," 
            sql += f"w_s_9 , w_s_9_product , w_s_9_num , w_s_9_normal_time , w_s_9_over_time , w_s_9_avail_time , w_s_9_remark ," 
            sql += f"w_s_10 , w_s_10_product , w_s_10_num , w_s_10_normal_time , w_s_10_over_time , w_s_10_avail_time , w_s_10_remark ," 
            sql += f"w_s_11 , w_s_11_product , w_s_11_num , w_s_11_normal_time , w_s_11_over_time , w_s_11_avail_time , w_s_11_remark ," 
            sql += f"w_s_12 , w_s_12_product , w_s_12_num , w_s_12_normal_time , w_s_12_over_time , w_s_12_avail_time , w_s_12_remark " 
            sql += f"from work_time where e_id='{e_id}' and e_name='{e_name}' and b_date='{b_date}'"

            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res

        except Exception as e:
            logging.info('\n<Error> load_work_time_data_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    ################################
    # load_check_member_data_list3
    ################################
    def load_check_member_data_list3(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 
                
            s_sql2 = f"select p_name , r_time , e_name from card_reader_{now_day} where e_name='{e_name}'"
            self.curr.execute(s_sql2)
            self.res2 = self.curr.fetchall()

            if self.res2 is not None:
                
                return self.res2
                
        except Exception as e:
            logging.info('\n<Error> load_check_member_data_list3 : ' + str(e))

        finally:
            self.__disconnect__()

    ########################################
    # load_card_reader_member_list_detail
    ########################################
    def load_card_reader_member_list_detail(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            e_name = str(e_name).strip()

            s_sql = f"select e_name , p_name , r_time from card_reader_{now_day} where e_name='{e_name}' order by r_time asc"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_list_detail : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # load_card_reader_member_list2
    #################################
    def load_card_reader_member_list2(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep  = str(dep).strip()
            dep2 = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'

            s_sql = f"select distinct e_name from card_reader_{now_day} where d_name='{dep2}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()
            
            for val in res:
                
                check_name = str(val[0]).strip()
                s_sql2 = f"select distinct e_name from factory_hr_a where d_name ='{dep}'"
                self.curr.execute(s_sql2)
                res2 = self.curr.fetchone()
                    
                if val[0] == res2[0]:
                    return False
                else:
                    return res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_list : ' + str(e))

        finally:
            self.__disconnect__()

    ###############################################
    # load_card_reader_member_check_status_list2
    ###############################################
    def load_card_reader_member_check_status_list2(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep  = str(dep).strip()
            dep2 = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select factory_hr_a.e_name from factory_hr_a inner join in_out_{now_day} on factory_hr_a.e_name!=in_out_{now_day}.e_name='{dep}' where factory_hr_a.d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchone()
            
            if res is not None:
                return res[0]  
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_list_real_total : ' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # load_card_reader_member_list_real_total
    ############################################
    def load_card_reader_member_list_real_total(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep  = str(dep).strip()
            dep2 = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select count(*) from in_out_{now_day} where d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchone()
            
            if res is not None:
                return res[0]  
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_list_real_total : ' + str(e))

        finally:
            self.__disconnect__()

    #######################################
    # load_card_reader_member_list_total
    #######################################
    def load_card_reader_member_list_total(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select count(*) from factory_hr_a where d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchone()
            
            if res is not None:
                return res[0]  
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_list_total : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##############################################
    # load_card_reader_member_check_status_list
    ##############################################
    def load_card_reader_member_check_status_list(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select distinct e_name from in_out_{now_day} where d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()
            
            return res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_check_status_list : ' + str(e))

        finally:
            self.__disconnect__()

    ##############################################
    # load_card_reader_door_list_by_every_month
    ##############################################
    def load_card_reader_door_list_by_every_month(self , position):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y%m%d" , time.localtime()) 

            position = str(position).strip()
            
            s_sql = f"show tables like 'card_reader_20%'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            return res
                
        except Exception as e:
            logging.info(f'\n<Error> load_card_reader_door_list_by_every_month : {str(e)}\n')

        finally:
            self.__disconnect__()

    ########################################
    # load_card_reader_door_list_by_month
    ########################################
    def load_card_reader_door_list_by_month(self , position):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y%m%d" , time.localtime()) 

            position = str(position).strip()
            
            s_sql = f"select r_date , count(*) from card_reader_{now_month} where p_name='{position}' group by r_date order by r_date asc"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            return res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_door_list_by_month : ' + str(e))

        finally:
            self.__disconnect__()

    ###################################################
    # create_bpm_expenditure_form_excel_pdf_csv_file
    ###################################################
    def create_bpm_expenditure_form_excel_pdf_csv_file(self , position  , day):
        
        self.__connect__()
        
        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y%m%d" , time.localtime()) 
            
            money_sql  = f"select r_time , e_name " 
            money_sql += f"from card_reader_{now_month} where p_name='{position}' and r_day='{day}' order by r_time asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/card_reader_'+ position + '_' + day + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"時間 , 人員"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/card_reader_'+ position + '_' + day + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('latin1')+' , '+str(val[1]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/card_reader_'+ position + '_' + day + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['時間','人員']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            '''
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)
            '''

            workbook.save(excel_file)

            #return money_res

        except Exception as e:
            logging.error(f'\n<Error> create_bpm_expenditure_form_excel_pdf_csv_file : {str(e)}\n')

        finally:
            self.__disconnect__()

    ##########################################
    # create_card_reader_excel_pdf_csv_file
    ##########################################
    def create_card_reader_excel_pdf_csv_file(self , position  , day):
        
        self.__connect__()
        
        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y%m%d" , time.localtime()) 
            
            money_sql  = f"select r_time , e_name " 
            money_sql += f"from card_reader_{now_month} where p_name='{position}' and r_day='{day}' order by r_time asc"
            self.curr.execute(money_sql)
            money_res = self.curr.fetchall()

            ###############
            #
            # export csv
            #
            ###############
            csv_file = 'csv/card_reader_'+ position + '_' + day + '.csv'
            #month    = '0' + month if int(month) < 10 else month
            
            with open(csv_file, mode='w', newline='' , encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='\t', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                
                title = f"時間 , 人員"
                
                writer.writerow(title)

                for row in money_res:
                    rows = f"{row}"
                    writer.writerow(row)
            
            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial',size=10)
            pdf_file = 'pdf/card_reader_'+ position + '_' + day + '.pdf'

            for val in money_res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('latin1')+' , '+str(val[1]).encode('utf8').decode('latin1') , ln=1 , align='left')

            pdf.output(pdf_file)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = 'excel/card_reader_'+ position + '_' + day + '.xlsx'

            ### title
            sheet.freeze_panes = 'A2'
            title = ['時間','人員']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(money_res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            '''
            total_sql  = f"select day_r_year , day_r_month , day_r_date , " 
            total_sql += f"format(sum(day_t_money1),0)  , format(sum(day_t_money2),0) , format(sum(day_t_money3),0) , format(sum(day_t_money4),0)  , format(sum(day_t_money5),0)  , format(sum(day_t_money6),0) , " 
            total_sql += f"format(sum(day_t_money7),0)  , format(sum(day_t_money8),0) , format(sum(day_t_money9),0) , format(sum(day_t_money10),0)  , format(sum(day_t_money11),0)  , format(sum(day_t_money12),0) , " 
            total_sql += f"format(sum(day_t_money13),0) , format(sum(day_t_money14),0) , format(sum(day_t_money15),0) , format(sum(day_t_money16),0)  , format(sum(day_t_money17),0)  , format(sum(day_t_money18),0) , " 
            total_sql += f"format(sum(day_t_money19),0) , format(sum(day_t_money20),0) , format(sum(day_t_money21),0) , format(sum(day_t_money22),0)  , format(sum(day_t_money23),0)  , format(sum(day_t_money24),0) , " 
            total_sql += f"format(sum(day_t_money25),0) , format(sum(day_t_money26),0) , format(sum(day_t_money27),0) , format(sum(day_t_money28),0)  , format(sum(day_t_money29),0)  , format(sum(day_t_money30),0) , " 
            total_sql += f"format(sum(day_t_money31),0) , format(sum(day_t_total),0) " 
            total_sql += f"from day_money_other where day_r_year='{year}' and day_r_month='{month}' "
            self.curr.execute(total_sql)
            total_res = self.curr.fetchall()
            
            last_row = sheet.max_row + 1

            for row_idx , row_data in enumerate(total_res , start=1):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=last_row , column=col_idx , value=cell_val)
            '''

            workbook.save(excel_file)

            #return money_res

        except Exception as e:
            logging.error('\n<Error> create_card_reader_excel_pdf_csv_file' + str(e))

        finally:
            self.__disconnect__()

    ############################################
    # card_reader_download_pdf_excel_by_month
    ############################################
    def card_reader_download_pdf_excel_by_month(self , position , month):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            first_day = time.strftime("%Y-%m" , time.localtime()) 
            first_day = first_day + '-01'
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            position  = str(position).strip()
            month     = str(month).strip()
            
            s_sql = f"select r_date , r_time , e_name from card_reader_{month} where p_name='{position}' and r_date >='{first_day}' and r_date <='{now_day}' order by r_date asc"
            
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('NotoSansCJK', '', '/home/otsuka/otsuka_platform/static/fonts/NotoSansTC-VariableFont_wght.ttf', uni=True)
            pdf.set_font('NotoSansCJK','',size=14)
            pdf_file = f'card_reader_{position}_{month}.pdf'
            pdf_destination = f"/home/otsuka/otsuka_platform/pdf/{pdf_file}"

            for val in res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('utf8')+' , '+str(val[1]).encode('utf8').decode('utf8')+' , '+str(val[2]).encode('utf8').decode('utf8') , ln=1 , align='left')
                
            pdf.output(pdf_file)
            shutil.move(pdf_file , pdf_destination)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = f'card_reader_{position}_{month}.xlsx'
            excel_destination = f"/home/otsuka/otsuka_platform/excel/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            title = ['日期','時間','人員']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)

            ################
            #
            # save to nas
            #
            ################
            '''
            cnopts = sftp.CnOpts()
            cnopts.hostkeys = None
            self.sftp = sftp.Connection(host=nas_para['host'] , username=nas_para['user'] , password=nas_para['pwd'] , port=nas_para['port'] , cnopts=cnopts)
            self.sftp.chdir(nas_para['nas_path_card_reader'])
            self.sftp.put(nas_para['linux_path_card_reader'] + excel_file)
            self.sftp.close()
            '''
            
            '''
            self.sftp.put(nas_para['linux_path_card_reader']+self.conver_pdf_day+'_S-7.pdf')
            if self.sftp.isfile(self.conver_pdf_day + '_WHFP-3-30-A.pdf'):
                self.sftp.remove(self.conver_pdf_day + '_WHFP-3-30-A.pdf')
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            else:
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            ### print msg
            print(self.r_time + ' , ' + self.conver_pdf_day + '_S-7(WHFP-3-30-A).pdf sftp put NAS successful.')
            
            '''
            
            #return res 
                
        except Exception as e:
            logging.info('\n<Error> card_reader_download_pdf_excel_by_month : ' + str(e))

        finally:
            self.__disconnect__()


    ###############################################
    # bpm_expenditure_form_download_excel_pdf
    ###############################################
    def bpm_expenditure_form_download_excel_pdf(self , q_s_date , q_e_date , q_b_e_dep , q_b_e_d_member , q_b_e_status , q_b_b_s_b_budget):
        
        self.__connect_mssql__()

        try:
            ### record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            q_s_date         = str(q_s_date).strip()
            q_e_date         = str(q_e_date).strip()
            q_b_e_dep        = str(q_b_e_dep).strip()
            q_b_e_d_member   = str(q_b_e_d_member).strip()
            q_b_e_status     = str(q_b_e_status).strip()
            q_b_b_s_b_budget = str(q_b_b_s_b_budget).strip()
            
            q_s_date = str(q_s_date).replace("-","/")
            q_e_date = str(q_e_date).replace("-","/")

            ## BPM expenditure MSSQL
            if q_b_e_status == '已結案':
                if q_b_e_d_member == '':
                    if q_b_b_s_b_budget == '':
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select (format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else:
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                else:
                    if q_b_b_s_b_budget == '':
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else: 
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='true' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"

            elif q_b_e_status == '未結案':
                if q_b_e_d_member == '':
                    if q_b_b_s_b_budget == '':
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else:
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"

                else:
                    if q_b_b_s_b_budget == '':
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
                    else:
                        s_sql  = f"select ITEM11 , ITEM12 , ITEM24 , ITEM19 , ITEM50 , format(ITEM15 , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false'  and ITEM24='{q_b_b_s_b_budget}'and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}' order by ITEM12 desc"
                        s2_sql = f"select format(sum(ITEM15) , 'N0' , 'en-US') from ART00691673700397107_Ins where ITEM56='{q_b_e_dep}' and ITEM50='{q_b_e_d_member}' and ITEM49='false' and ITEM24='{q_b_b_s_b_budget}'and ITEM12>='{q_s_date}' and ITEM12<='{q_e_date}'"
            
            #print(s_sql)

            self.curr_mssql.execute(s_sql)
            res = self.curr_mssql.fetchall()

            ###############
            #
            # export PDF
            #
            ###############
            '''
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('NotoSansCJK', '', '/home/otsuka/otsuka_platform/static/fonts/NotoSansTC-VariableFont_wght.ttf', uni=True)
            pdf.set_font('NotoSansCJK', '', size=14)
            
            if q_b_e_d_member is None:
                pdf_file = f"開支證明單_{q_b_e_dep}_{q_b_e_status}.pdf"
            elif q_b_e_d_member is not None:
                pdf_file = f"開支證明單_{q_b_e_dep}_{q_b_e_d_member}_{q_b_e_status}.pdf"
            
            pdf_destination = f"/home/otsuka/otsuka_platform/pdf/bpm/{pdf_file}"

            for val in res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('utf8')+' , '+str(val[1]).encode('utf8').decode('utf8')+' , '+str(val[2]).encode('utf8').decode('utf8') , ln=1 , align='left')
                
            pdf.output(pdf_file)
            shutil.move(pdf_file , pdf_destination)
            '''

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active

            excel_file = f"開支證明單_{q_b_e_dep}_{q_b_e_d_member}_{q_b_e_status}.xlsx"
            excel_destination = f"/home/otsuka/otsuka_platform/excel/bpm/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            title = ['表單單號','申請日期','預算來源','申請原因','申請人','金額']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF" , size=14 , name='Meiryo UI')
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            ### total
            self.curr_mssql.execute(s2_sql)
            total_res = self.curr_mssql.fetchone()
            
            
            # 找出最後一排和最後一列
            max_row = sheet.max_row
            max_column = sheet.max_column

            # 在最後一排的最後一列的下一個單元格中添加值
            sheet.cell(row=max_row, column=max_column + 1, value=total_res)
            
            
            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)

            ################
            #
            # save to nas
            #
            ################
            '''
            cnopts = sftp.CnOpts()
            cnopts.hostkeys = None
            self.sftp = sftp.Connection(host=nas_para['host'] , username=nas_para['user'] , password=nas_para['pwd'] , port=nas_para['port'] , cnopts=cnopts)
            self.sftp.chdir(nas_para['nas_path_card_reader'])
            self.sftp.put(nas_para['linux_path_card_reader'] + excel_file)
            self.sftp.close()
            '''
            
            '''
            self.sftp.put(nas_para['linux_path_card_reader']+self.conver_pdf_day+'_S-7.pdf')
            if self.sftp.isfile(self.conver_pdf_day + '_WHFP-3-30-A.pdf'):
                self.sftp.remove(self.conver_pdf_day + '_WHFP-3-30-A.pdf')
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            else:
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            ### print msg
            print(self.r_time + ' , ' + self.conver_pdf_day + '_S-7(WHFP-3-30-A).pdf sftp put NAS successful.')
            
            '''
            
            
            #return res 
                
        except Exception as e:
            logging.info(f'\n<Error> bpm_expenditure_form_download_excel_pdf : {str(e)}\n')

        finally:
            self.__disconnect_mssql__()

    ###############################################
    # card_reader_download_pdf_excel_every_month
    ###############################################
    def card_reader_download_pdf_excel_every_month(self , position , tb):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            position = str(position).strip()
            tb       = str(tb).strip()
            
            s_sql = f"select r_date , r_time , e_name from {tb} where p_name='{position}' order by r_date asc"
            
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('NotoSansCJK', '', '/home/otsuka/otsuka_platform/static/fonts/NotoSansTC-VariableFont_wght.ttf', uni=True)
            pdf.set_font('NotoSansCJK', '', size=14)
            pdf_file = f'{tb}_{position}.pdf'
            pdf_destination = f"/home/otsuka/otsuka_platform/pdf/{pdf_file}"

            for val in res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('utf8')+' , '+str(val[1]).encode('utf8').decode('utf8')+' , '+str(val[2]).encode('utf8').decode('utf8') , ln=1 , align='left')
                
            pdf.output(pdf_file)
            shutil.move(pdf_file , pdf_destination)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = f'{tb}_{position}.xlsx'
            excel_destination = f"/home/otsuka/otsuka_platform/excel/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            title = ['日期','時間','人員']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF" , size=14 , name='Andale mono')
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)

            ################
            #
            # save to nas
            #
            ################
            '''
            cnopts = sftp.CnOpts()
            cnopts.hostkeys = None
            self.sftp = sftp.Connection(host=nas_para['host'] , username=nas_para['user'] , password=nas_para['pwd'] , port=nas_para['port'] , cnopts=cnopts)
            self.sftp.chdir(nas_para['nas_path_card_reader'])
            self.sftp.put(nas_para['linux_path_card_reader'] + excel_file)
            self.sftp.close()
            '''
            
            '''
            self.sftp.put(nas_para['linux_path_card_reader']+self.conver_pdf_day+'_S-7.pdf')
            if self.sftp.isfile(self.conver_pdf_day + '_WHFP-3-30-A.pdf'):
                self.sftp.remove(self.conver_pdf_day + '_WHFP-3-30-A.pdf')
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            else:
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            ### print msg
            print(self.r_time + ' , ' + self.conver_pdf_day + '_S-7(WHFP-3-30-A).pdf sftp put NAS successful.')
            
            '''
            
            
            #return res 
                
        except Exception as e:
            logging.info(f'\n<Error> card_reader_download_pdf_excel_every_month : {str(e)}\n')

        finally:
            self.__disconnect__()

    #####################################################
    # e_board_list
    #####################################################
    def e_board_list(self):
        
        self.__connect7_1_38__()

        try:
            
            sql3 = f"select e_date , e_s_time , e_e_time , e_company , e_title , e_name , e_instructions from e_board order by e_date desc"
            self.curr_7_1_38.execute(sql3)
            e_b_res = self.curr_7_1_38.fetchall()

            return e_b_res

        except Exception as e:
            logging.error(f"\n<Error> e_board_list : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # submit_finish_e_board
    #####################################################
    def submit_finish_e_board(self , e_f_a_date , e_f_s_date , e_f_e_date , e_f_date , e_f_name):
        
        self.__connect7_1_38__()

        try:

            sql  = f"select * from e_board where e_date='{e_f_a_date}' and e_s_time='{e_f_s_date}' and e_e_time={e_f_e_date}"
            self.curr_7_1_38.execute(sql)
            res = self.curr_7_1_38.fetchone()

            if res is not None:

                sql2 = f"upadte e_board set e_c_finish_date='{e_f_date}' and e_c_name='{e_f_name}' where e_date='{e_f_a_date}' and e_s_time='{e_f_s_date}' and e_e_time={e_f_e_date}"
                self.curr_7_1_38.execute(sql2)

                sql3 = f"select e_date , e_s_time , e_e_time , e_company , e_title , e_name , e_instructions , e_c_finish_date , e_c_name from e_board order by e_date desc"
                self.curr_7_1_38.execute(sql3)
                e_b_res = self.curr_7_1_38.fetchall()

            return e_b_res

        except Exception as e:
            logging.error(f"\n<Error> submit_finish_e_board : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()
    
    #####################################################
    # submit_response_e_board
    #####################################################
    def submit_response_e_board(self , a_date , s_time , e_time):
        
        self.__connect7_1_38__()

        try:

            sql3 = '''
                    select e_date , e_s_time , e_e_time , e_company , e_title , e_name , e_instructions from e_board 
                    where e_date=%s and e_s_time=%s and e_e_time=%s
                    '''
                   
            self.curr_7_1_38.execute(sql3 , (a_date , s_time , e_time ,))
            e_b_res = self.curr_7_1_38.fetchall()

            return e_b_res

        except Exception as e:
            logging.error(f"\n<Error> submit_response_e_board : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    #####################################################
    # submit_e_board
    #####################################################
    def submit_e_board(self , e_a_date , e_s_date , e_e_date , e_c_name , e_title , e_name , e_other):
        
        self.__connect7_1_38__()

        try:

            sql2  = f"insert into e_board(e_date , e_s_time , e_e_time , e_company , e_title , e_name , e_instructions) "
            sql2 += f"value('{e_a_date}' , '{e_s_date}' , '{e_e_date}' , '{e_c_name}' , '{e_title}' , '{e_name}' , '{e_other}')"

            self.curr_7_1_38.execute(sql2)

            sql3 = f"select e_date , e_s_time , e_e_time , e_company , e_title , e_name , e_instructions from e_board order by e_date desc"
            self.curr_7_1_38.execute(sql3)
            e_b_res = self.curr_7_1_38.fetchall()

            return e_b_res

        except Exception as e:
            logging.error(f"\n<Error> submit_e_board : {str(e)}\n")

        finally:
            self.__disconnect7_1_38__()

    ######################################
    # factory_erp_ss2_download_pdf_json
    ######################################
    def custom_serializer(self,obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()  
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

    def factory_erp_ss2_download_pdf_json(self, q_s_date, q_e_date, q_c_name, q_p_name):
        
        self.__connect_mssql_ss2__()

        try:
            # Record the current time
            now_year = time.strftime("%Y", time.localtime())
            now_month = time.strftime("%Y%m", time.localtime())
            now_day = time.strftime("%Y-%m-%d", time.localtime())

            # Clean input parameters
            q_s_date = str(q_s_date).strip()
            q_e_date = str(q_e_date).strip()
            q_c_name = str(q_c_name).strip()
            q_p_name = str(q_p_name).strip()

            # Construct SQL query based on input conditions
            condition = ""
            if q_c_name == "" and q_p_name == "":
                condition = f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}'"
            elif q_c_name != "" and q_p_name == "":
                condition = f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}'"
            elif q_p_name != "" and q_c_name == "":
                condition = f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}'"
            elif q_c_name != "" and q_p_name != "":
                condition = f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}' AND p.PROD_NAME='{q_p_name}'"

            # Construct the full SQL query
            f_erp_sql = f"""
                            SELECT v.TRADE_DATE AS 交易日期, v.CUST_ID AS 客戶代號, c.FULLNAME AS 客戶名稱, c.ADDRESS AS 地址, 
                                v.OLD_ID, v.PROD_ID AS 產品代號, p.PROD_NAME AS 產品名稱, p.ERP_CODE AS ERP產品代碼,
                                v.LOT_NO AS 批號, v.NET_QTY AS 淨數量, v.NET_AMT AS 淨金額, v.RESELLER_ID AS 經銷商,
                                v.WH_ID AS 庫別, v.AMOUNT, v.DISCOUNT, v.QTY, v.TAX
                            FROM SALES2.dbo.vTRADEDET v
                            JOIN SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID
                            JOIN SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID
                            WHERE {condition}
                            ORDER BY v.TRADE_DATE DESC
                        """

            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            data_list = []
            for row in res:
                
                j_instruction  =  f"你是大塚製藥的產品經理，請根據大塚製藥的產品來進行相關回答"
                j_input        = f"{row[0]} {row[2]} 購買 {row[6]}"
                j_output       = f"客戶代號 : {row[1]} , 客戶名稱 : {row[2]} , 地址 : {row[3]} , OLD_ID : {row[4]} , " 
                j_output      += f"產品代號 : {row[5]} , 產品名稱 : {row[6]} , ERP產品代碼 : {row[7]} , 批號 : {row[8]} , 淨數量 : {row[9]} , " 
                j_output      += f"淨金額 : {row[10]} , 經銷商 : {row[11]} , 庫別 : {row[12]} , 數量 : {row[13]} , 折扣 : {row[14]} , 稅金 : {row[15]}"

                row_dict = {'instruction':j_instruction , 'input':j_input , 'output':j_output}
                data_list.append(row_dict)

            # Export data to a JSON file
            export_file = f'/home/otsuka/otsuka_platform/json/factory_erp_ss2/factory_erp_ss2_{q_s_date}_{q_e_date}.json'
            with open(export_file, 'w', encoding='utf-8') as json_file:
                json.dump(data_list, json_file, ensure_ascii=False, indent=4, default=self.custom_serializer)

            print(f"Data has been exported to {export_file}")

        except Exception as e:
            print(f"An error occurred: {e}")

    ###########################################
    # factory_erp_subform_download_pdf_excel
    ###########################################
    def factory_erp_subform_download_pdf_excel(self , q_s_date , q_e_date):
        
        self.__connect_mssql_erp__()

        try:
            # record time
            now_year  = time.strftime("%Y" , time.localtime()) 
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            s_date = datetime.strptime(q_s_date, "%Y-%m-%d")
            e_date = datetime.strptime(q_e_date, "%Y-%m-%d")
            s_d    = s_date.strftime("%Y%m%d")
            e_d    = e_date.strftime("%Y%m%d")

            s_erp_sql = '''
                            SELECT 
                                ACTTA.TA003 AS '傳票日期',
                                INVMB.MB008 AS '成品品類',
                                ACTXA.XA005 AS '品號',
                                INVMB.MB002 AS '品名', 
                                ACTXA.XA002 AS '傳票單號',
                                ACTTB.TB005 AS '科目編號',
                                ACTMA.MA003 AS '科目名稱',
                                ACTTB.TB006 AS '部門代號', 
                                (ACTTB.TB004 * ACTXA.XA010) AS Sub_Cost

                            FROM 
                                OtsukaDB.dbo.ACTTA ACTTA
                                INNER JOIN OtsukaDB.dbo.ACTTB ACTTB ON ACTTA.TA001 = ACTTB.TB001 AND ACTTA.TA002 = ACTTB.TB002
                                INNER JOIN OtsukaDB.dbo.ACTXA ACTXA ON ACTTB.TB001 = ACTXA.XA001 AND ACTTB.TB002 = ACTXA.XA002 AND ACTTB.TB003 = ACTXA.XA003
                                INNER JOIN OtsukaDB.dbo.ACTMA ACTMA ON ACTTB.TB005 = ACTMA.MA001
                                LEFT OUTER JOIN OtsukaDB.dbo.INVMB INVMB ON ACTXA.XA005 = INVMB.MB001

                            WHERE
                                ACTTA.TA003 BETWEEN ? AND ?
                                AND ACTTB.TB005 LIKE '62%' 

                            ORDER BY ACTTA.TA003;
                        '''
            
            self.curr_mssql_erp.execute(s_erp_sql , (s_d,e_d,))
            res = self.curr_mssql_erp.fetchall()

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
                
            excel_file        = f'factory_erp_subform_{s_d}_{e_d}.xlsx'
            excel_destination = f"/home/otsuka/otsuka_platform/excel/factory_erp_ss2/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            
            title = (['傳票日期','成品品類','品號','品名','傳票單號','科目編號','科目名稱','部門代號','Sub_Cost'])
            
            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)
                
        except Exception as e:
            logging.info('\n<Error> factory_erp_subform_download_pdf_excel : ' + str(e) + '\n')

        finally:
            self.__disconnect_mssql_erp__()

    ######################################
    # factory_erp_ss2_download_pdf_excel
    ######################################
    def factory_erp_ss2_download_pdf_excel(self , q_s_date , q_e_date , q_c_name , q_p_name):
        
        self.__connect_mssql_ss2__()

        try:
            # record time
            now_year  = time.strftime("%Y" , time.localtime()) 
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            q_s_date = str(q_s_date).strip()
            q_e_date = str(q_e_date).strip()
            q_c_name = str(q_c_name).strip()
            q_p_name = str(q_p_name).strip()
            
            if q_c_name == "" and q_p_name == "":
            
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.TAX "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' order by v.TRADE_DATE desc"
            
            elif q_c_name != "" and q_p_name == "":

                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.TAX "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}' order by v.TRADE_DATE desc"
            
            elif q_p_name != "" and q_c_name == "":

                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.TAX "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}' order by v.TRADE_DATE desc"
            
            elif q_c_name != "" and q_p_name != "":

                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.TAX "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}' AND p.PROD_NAME='{q_p_name}' order by v.TRADE_DATE desc"
            
            
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('NotoSansCJK', '', '/home/otsuka/otsuka_platform/static/fonts/NotoSansTC-VariableFont_wght.ttf', uni=True)
            pdf.set_font('NotoSansCJK','',size=14)
            
            if (q_c_name is "") and (q_p_name is ""):
                pdf_file = f'factory_erp_ss2_{q_s_date}_{q_e_date}.pdf'

            elif (q_c_name is "") and (q_p_name is not ""):
                pdf_file = f'factory_erp_ss2_{q_s_date}_{q_e_date}_{q_c_name}.pdf'

            elif (q_p_name is "") and (q_c_name is not ""):
                pdf_file = f'factory_erp_ss2_{q_s_date}_{q_e_date}_{q_p_name}.pdf'
            
            elif (q_c_name is not "") and (q_p_name is not ""):
                pdf_file = f'factory_erp_ss2_{q_s_date}_{q_e_date}_{q_c_name}_{q_p_name}.pdf'
            
            pdf_destination = f"/home/otsuka/otsuka_platform/pdf/factory_erp_ss2/{pdf_file}"

            for val in res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('utf8')+' , '+str(val[1]).encode('utf8').decode('utf8')+' , '+str(val[2]).encode('utf8').decode('utf8') , ln=1 , align='left')
                
            pdf.output(pdf_file)
            shutil.move(pdf_file , pdf_destination)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
                
            excel_file        = f'factory_erp_ss2_{q_s_date}_{q_e_date}.xlsx'
            excel_destination = f"/home/otsuka/otsuka_platform/excel/factory_erp_ss2/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            
            title = (['交易日期','客戶代號','客戶名稱','地址','舊產品代號','產品代號','產品名稱','ERP 產品代碼','批號','淨數量','淨金額','經銷商','庫別','數量','折扣','QTY','TAX'])
            
            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)

            ################
            #
            # save to nas
            #
            ################
            '''
            cnopts = sftp.CnOpts()
            cnopts.hostkeys = None
            self.sftp = sftp.Connection(host=nas_para['host'] , username=nas_para['user'] , password=nas_para['pwd'] , port=nas_para['port'] , cnopts=cnopts)
            self.sftp.chdir(nas_para['nas_path_card_reader'])
            self.sftp.put(nas_para['linux_path_card_reader'] + excel_file)
            self.sftp.close()
            '''
            
            '''
            self.sftp.put(nas_para['linux_path_card_reader']+self.conver_pdf_day+'_S-7.pdf')
            if self.sftp.isfile(self.conver_pdf_day + '_WHFP-3-30-A.pdf'):
                self.sftp.remove(self.conver_pdf_day + '_WHFP-3-30-A.pdf')
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            else:
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            ### print msg
            print(self.r_time + ' , ' + self.conver_pdf_day + '_S-7(WHFP-3-30-A).pdf sftp put NAS successful.')
            '''
            
            #return res 
                
        except Exception as e:
            logging.info('\n<Error> factory_erp_ss2_download_pdf_excel : ' + str(e) + '\n')

        finally:
            self.__disconnect_mssql_ss2__()

    ######################################
    # hr_360_download_pdf_excel
    ######################################
    def hr_360_download_pdf_excel(self , job):
        
        self.__connect7_1_38__()

        try:
            # record time
            now_year = time.strftime("%Y" , time.localtime()) 
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            job = str(job).strip()

            #########
            # 主管
            #########
            if job == 'manager':
                s_sql  = f"select " 
                s_sql += f"c_year , s_date , s_user , s_name , "
                s_sql += f"s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , "
                s_sql += f"s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , s_hr_360_2_6 , "
                s_sql += f"s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , "
                s_sql += f"s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , "
                s_sql += f"s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , "
                s_sql += f"s_hr_360_6_1 , s_hr_360_6_2 , s_hr_360_6_3 , s_hr_360_6_4 , s_hr_360_6_5 , "
                s_sql += f"s_hr_360_total_1 , s_hr_360_total_1_avg , s_hr_360_total_2 , s_hr_360_total_2_avg , "
                s_sql += f"s_hr_360_total_3 , s_hr_360_total_3_avg , s_hr_360_total_4 , s_hr_360_total_4_avg , "
                s_sql += f"s_hr_360_total_5 , s_hr_360_total_5_avg , s_hr_360_total_6 , s_hr_360_total_6_avg "
                s_sql += f"from hr_360_submit_manager_content order by s_name desc"
            #########
            # 員工
            #########
            elif job == 'member':
                s_sql  = f"select " 
                s_sql += f"c_year , s_date , s_user , s_name , "
                s_sql += f"s_hr_360_1_1 , s_hr_360_1_2 , s_hr_360_1_3 , s_hr_360_1_4 , s_hr_360_1_5 , "
                s_sql += f"s_hr_360_2_1 , s_hr_360_2_2 , s_hr_360_2_3 , s_hr_360_2_4 , s_hr_360_2_5 , "
                s_sql += f"s_hr_360_3_1 , s_hr_360_3_2 , s_hr_360_3_3 , s_hr_360_3_4 , s_hr_360_3_5 , "
                s_sql += f"s_hr_360_4_1 , s_hr_360_4_2 , s_hr_360_4_3 , s_hr_360_4_4 , s_hr_360_4_5 , "
                s_sql += f"s_hr_360_5_1 , s_hr_360_5_2 , s_hr_360_5_3 , s_hr_360_5_4 , s_hr_360_5_5 , s_hr_360_5_6 , "
                s_sql += f"s_hr_360_total_1 , s_hr_360_total_1_avg , s_hr_360_total_2 , s_hr_360_total_2_avg , "
                s_sql += f"s_hr_360_total_3 , s_hr_360_total_3_avg , s_hr_360_total_4 , s_hr_360_total_4_avg , "
                s_sql += f"s_hr_360_total_5 , s_hr_360_total_5_avg "
                s_sql += f"from hr_360_submit_member_content order by s_name desc"
            
            self.curr_7_1_38.execute(s_sql)
            res = self.curr_7_1_38.fetchall()

            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('NotoSansCJK', '', '/home/otsuka/otsuka_platform/static/fonts/NotoSansTC-VariableFont_wght.ttf', uni=True)
            pdf.set_font('NotoSansCJK','',size=14)
            pdf_file = f'hr_360_{job}.pdf'
            pdf_destination = f"/home/otsuka/otsuka_platform/pdf/hr_360/{pdf_file}"

            for val in res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('utf8')+' , '+str(val[1]).encode('utf8').decode('utf8')+' , '+str(val[2]).encode('utf8').decode('utf8') , ln=1 , align='left')
                
            pdf.output(pdf_file)
            shutil.move(pdf_file , pdf_destination)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = f'hr_360_{job}.xlsx'
            excel_destination = f"/home/otsuka/otsuka_platform/excel/hr_360/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            
            if job == 'manager':
                title = (['考評年度','考評時間','考評人員','受評人員','1-1','1-2','1-3','1-4','1-5','2-1','2-2','2-3','2-4','2-5','2-6',
                          '3-1','3-2','3-3','3-4','3-5','4-1','4-2','4-3','4-4','5-1','5-2','5-3','6-1','6-6','6-3','6-4','6-5',
                          'total_1','total_1_avg','total_2','total_2_avg','total_3','total_3_avg','total_4','total_4_avg','total_5','total_5_avg',
                          'total_6','total_6_avg'
                        ])
            elif job == 'member':
                title = (['考評年度','考評時間','考評人員','受評人員','1-1','1-2','1-3','1-4','1-5','2-1','2-2','2-3','2-4','2-5',
                          '3-1','3-2','3-3','3-4','3-5','4-1','4-2','4-3','4-4','4-5','5-1','5-2','5-3','5-4','5-5','5-6',
                          'total_1','total_1_avg','total_2','total_2_avg','total_3','total_3_avg','total_4','total_4_avg','total_5','total_5_avg'
                        ])
            

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)

            ################
            #
            # save to nas
            #
            ################
            '''
            cnopts = sftp.CnOpts()
            cnopts.hostkeys = None
            self.sftp = sftp.Connection(host=nas_para['host'] , username=nas_para['user'] , password=nas_para['pwd'] , port=nas_para['port'] , cnopts=cnopts)
            self.sftp.chdir(nas_para['nas_path_card_reader'])
            self.sftp.put(nas_para['linux_path_card_reader'] + excel_file)
            self.sftp.close()
            '''
            
            '''
            self.sftp.put(nas_para['linux_path_card_reader']+self.conver_pdf_day+'_S-7.pdf')
            if self.sftp.isfile(self.conver_pdf_day + '_WHFP-3-30-A.pdf'):
                self.sftp.remove(self.conver_pdf_day + '_WHFP-3-30-A.pdf')
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            else:
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            ### print msg
            print(self.r_time + ' , ' + self.conver_pdf_day + '_S-7(WHFP-3-30-A).pdf sftp put NAS successful.')
            
            '''
            
            
            #return res 
                
        except Exception as e:
            logging.info('\n<Error> hr_360_download_pdf_excel : ' + str(e) + '\n')

        finally:
            self.__disconnect7_1_38__()

    ######################################
    # card_reader_download_pdf_excel
    ######################################
    def card_reader_download_pdf_excel(self , position , day):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            position = str(position).strip()
            day      = str(day).strip()
            
            if day == '':
                s_sql = f"select r_date , r_time , e_name from card_reader_{now_month} where p_name='{position}' and r_date='{now_day}' order by r_time asc"
            else:
                s_sql = f"select r_date , r_time , e_name from card_reader_{now_month} where p_name='{position}' and r_date='{day}' order by r_time asc"
            
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            ###############
            #
            # export PDF
            #
            ###############
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('NotoSansCJK', '', '/home/otsuka/otsuka_platform/static/fonts/NotoSansTC-VariableFont_wght.ttf', uni=True)
            pdf.set_font('NotoSansCJK','',size=14)
            pdf_file = f'card_reader_{position}_{day}.pdf'
            pdf_destination = f"/home/otsuka/otsuka_platform/pdf/{pdf_file}"

            for val in res:
                pdf.cell(200 , 10 , txt=str(val[0]).encode('utf8').decode('utf8')+' , '+str(val[1]).encode('utf8').decode('utf8')+' , '+str(val[2]).encode('utf8').decode('utf8') , ln=1 , align='left')
                
            pdf.output(pdf_file)
            shutil.move(pdf_file , pdf_destination)

            ################
            #
            # export excel
            #
            ################
            workbook   = openpyxl.Workbook()
            sheet      = workbook.active
            excel_file = f'card_reader_{position}_{day}.xlsx'
            excel_destination = f"/home/otsuka/otsuka_platform/excel/{excel_file}"

            ### title
            sheet.freeze_panes = 'A2'
            title = ['日期','時間','人員']

            for col_num , header in enumerate(title , 1):
                cell            = sheet.cell(row=1 , column=col_num , value=header)
                cell.font       = Font(bold=True , color="FFFFFF")
                cell.alignment  = Alignment(horizontal="center")
                cell.fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 设置背景颜色为灰色

            ### content
            for row_idx , row_data in enumerate(res , start=2):
                for col_idx , cell_val in enumerate(row_data , start=1):
                    sheet.cell(row=row_idx , column=col_idx , value=cell_val)

            workbook.save(excel_file)
            shutil.move(excel_file , excel_destination)

            ################
            #
            # save to nas
            #
            ################
            '''
            cnopts = sftp.CnOpts()
            cnopts.hostkeys = None
            self.sftp = sftp.Connection(host=nas_para['host'] , username=nas_para['user'] , password=nas_para['pwd'] , port=nas_para['port'] , cnopts=cnopts)
            self.sftp.chdir(nas_para['nas_path_card_reader'])
            self.sftp.put(nas_para['linux_path_card_reader'] + excel_file)
            self.sftp.close()
            '''
            
            '''
            self.sftp.put(nas_para['linux_path_card_reader']+self.conver_pdf_day+'_S-7.pdf')
            if self.sftp.isfile(self.conver_pdf_day + '_WHFP-3-30-A.pdf'):
                self.sftp.remove(self.conver_pdf_day + '_WHFP-3-30-A.pdf')
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            else:
                self.sftp.rename(self.conver_pdf_day + '_S-7.pdf' , self.conver_pdf_day + '_WHFP-3-30-A.pdf')
            ### print msg
            print(self.r_time + ' , ' + self.conver_pdf_day + '_S-7(WHFP-3-30-A).pdf sftp put NAS successful.')
            
            '''
            
            
            #return res 
                
        except Exception as e:
            logging.info('\n<Error> card_reader_download_pdf_excel : ' + str(e))

        finally:
            self.__disconnect__()
    
    #####################################################
    # load_card_reader_door_list_by_every_month_detail
    #####################################################
    def load_card_reader_door_list_by_every_month_detail(self , position , tb):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day = time.strftime("%Y-%m-%d" , time.localtime()) 

            position = str(position).strip()
            tb       = str(tb).strip()
            
            s_sql = f"select r_date , r_time , e_name from {tb} where p_name='{position}' order by r_date asc"
            
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            return res
                
        except Exception as e:
            logging.info(f'\n<Error> load_card_reader_door_list_by_every_month_detail : {str(e)}\n')

        finally:
            self.__disconnect__()

    ######################################
    # load_card_reader_door_list_by_day
    ######################################
    def load_card_reader_door_list_by_day(self , position , day):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day = time.strftime("%Y-%m-%d" , time.localtime()) 

            position = str(position).strip()
            day      = str(day).strip()
            
            f_res = []

            if day == '':
                s_sql = f"select r_date , r_time , e_name , position from card_reader_{now_month} where p_name='{position}' and r_date='{now_day}' order by r_time asc"
            else:
                s_sql = f"select r_date , r_time , e_name , position from card_reader_{now_month} where p_name='{position}' and r_date='{day}' order by r_time asc"
            
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            for val in res:
                
                f_r_time = f"{val[1][:2]}:{val[1][2:]}" 

                f_res.append((val[0] , f_r_time , val[2], val[3]))


            return f_res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_door_list_by_day : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # load_card_reader_door_list
    #################################
    def load_card_reader_door_list(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
            now_day   = time.strftime("%Y-%m-%d" , time.localtime()) 

            position = str(dep).strip()
            
            s_sql = f"select r_date , r_time , e_name from card_reader_{now_month} where p_name='{position}' and r_date='{now_day}' order by r_time asc"
            
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            return res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_door_list : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # load_card_reader_member_list
    #################################
    def load_card_reader_member_list(self , dep):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 

            dep = str(dep).strip()

            if dep == '生產一部':
                dep = '生一部'
            elif dep == '生產二部':
                dep = '生二部'
            elif dep == '生產三部':
                dep = '生三部'
            
            s_sql = f"select distinct e_name from factory_hr_a where d_name='{dep}'"
            #s_sql = f"select distinct factory_hr_a.e_name from factory_hr_a inner join in_out_{now_day} on factory_hr_a.e_name = in_out_{now_day}.e_name where factory_hr_a.d_name='{dep}'"
            self.curr.execute(s_sql)
            res = self.curr.fetchall()

            for val in res:

                try:
                    a_sql =  f"create table in_out_{now_day}("
                    a_sql += f"no int not null primary key AUTO_INCREMENT,"
                    a_sql += f"r_date varchar(20) null,"
                    a_sql += f"r_time varchar(20) null,"
                    a_sql += f"d_id varchar(50) null,"
                    a_sql += f"d_name varchar(50) null,"
                    a_sql += f"e_id varchar(50) null,"
                    a_sql += f"e_name varchar(50) null,"
                    a_sql += f"p_id varchar(50) null,"
                    a_sql += f"p_name varchar(50) null,"
                    a_sql += f"c_id varchar(50) null"
                    a_sql += f")ENGINE=InnoDB DEFAULT CHARSET=utf8 COLLATE=utf8_unicode_ci;"

                    self.curr.execute(a_sql)

                except Exception as e:
                    
                    s_sql = f"select e_name from card_reader_{now_day} where e_name='{val[0]}' order by r_time desc limit 0,1"
                    self.curr.execute(s_sql)
                    res2 = self.curr.fetchone()

                    if res2 is not None:

                        s_sql3 = f"select e_name from in_out_{now_day} where d_name='{dep}' and e_name='{res2[0]}' order by no desc limit 0,1"
                        self.curr.execute(s_sql3)
                        res3 = self.curr.fetchone()

                        if res3 is None:

                            a_sql = f"insert into in_out_{now_day}(d_name , e_name) value('{dep}' , '{res2[0]}')"
                            self.curr.execute(a_sql)

                finally:
                    pass
            
            return res
                
        except Exception as e:
            logging.info('\n<Error> load_card_reader_member_list : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # load_group_member_list
    ###########################
    def load_group_member_list(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_day = time.strftime("%Y%m%d" , time.localtime()) 
                
            s_sql = f"SELECT TRIM(d_name) FROM `card_reader_{now_day}` where e_name='{e_name}' order by d_name desc limit 0,1"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchone()

            s_sql2 = f"select distinct e_name from card_reader_{now_day} where d_name='{self.res[0]}'"
            self.curr.execute(s_sql2)
            self.res2 = self.curr.fetchall()

            return self.res2
                
        except Exception as e:
            logging.info('\n<Error> load_group_member_list : ' + str(e))

        finally:
            self.__disconnect__()

    ################################
    # load_check_member_data_list2
    ################################
    def load_check_member_data_list2(self , e_name):
        
        self.__connect__()

        try:
            # record time
            now_month = time.strftime("%Y%m" , time.localtime()) 
                
            s_sql2 = f"SELECT p_name FROM `card_reader_{now_month}` group by p_name  order by p_id asc"
            self.curr.execute(s_sql2)
            self.res2 = self.curr.fetchall()

            if self.res2 is not None:
                return self.res2
                
        except Exception as e:
            logging.info('\n<Error> load_check_member_data_list2 : ' + str(e))

        finally:
            self.__disconnect__()

    ################################
    # load_check_member_data_list
    ################################
    def load_check_member_data_list(self , check_year , check_month , employee_name):
        
        self.__connect__()

        try:
            sql = f"select self_item_1_1 , self_item_1_s , self_item_1_3 , self_item_1_4 from check_member where check_year='{check_year}' and check_month='{check_month}' and employee_name='{employee_name}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res

        except Exception as e:
            logging.info('\n<Error> load_check_member_data_list : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # update_submit_check_member_2
    #################################
    def update_submit_check_member_2(self , employee_id , employee_name , check_year , check_month , sir_num1_1 , sir_num1_2 , sir_num1_3 , sir_num1_4 , sir_num2_1 , sir_num2_2 , sir_num2_3 , sir_num3_1 , sir_num3_2 , sir_num3_3 , sir_num4_1 , sir_num4_2 , sir_num4_3 , sir_num4_4 , sir_num5_1 , sir_num5_2 , sir_num5_3 , sir_num6_1 , sir_num6_2 , sir_num6_3 , sir_num7_1 , sir_num7_2 , sir_num7_3 , sir_num7_4 , sir_num8_1 , sir_num8_2 , sir_num8_3 , sir_num8_4 , sir_num8_5 , comment , other_total , sir_total , other_plus_total , final_total , final_comment):
        
        self.__connect__()

        try:
            sql = f"update check_member set sir_item_1_1='{sir_num1_1}' , sir_item_1_2='{sir_num1_2}' , sir_item_1_3='{sir_num1_3}' , sir_item_1_4='{sir_num1_4}' , sir_item_2_1='{sir_num2_1}' , sir_item_2_2='{sir_num2_2}' , sir_item_2_3='{sir_num2_3}' , sir_item_3_1='{sir_num3_1}' , sir_item_3_2='{sir_num3_2}' , sir_item_3_3='{sir_num3_3}' , sir_item_4_1='{sir_num4_1}' , sir_item_4_2='{sir_num4_2}' , sir_item_4_3='{sir_num4_3}' , sir_item_4_4='{sir_num4_4}' , sir_item_5_1='{sir_num5_1}' , sir_item_5_2='{sir_num5_2}' , sir_item_5_3='{sir_num5_3}' , sir_item_6_1='{sir_num6_1}' , sir_item_6_2='{sir_num6_2}' , sir_item_6_3='{sir_num6_3}' , sir_item_7_1='{sir_num7_1}' , sir_item_7_2='{sir_num7_2}' , sir_item_7_3='{sir_num7_3}' , sir_item_7_4='{sir_num7_4}' , sir_item_8_1='{sir_num8_1}' , sir_item_8_2='{sir_num8_2}' , sir_item_8_3='{sir_num8_3}' , sir_item_8_4='{sir_num8_4}' , sir_item_8_5='{sir_num8_5}' , comment='{comment}' , sir_total='{sir_total}' , other_total='{other_total}' , other_plus_total='{other_plus_total}' , final_total='{final_total}' , final_comment='{final_comment}' , sir_check='done' where employee_id='{employee_id}' and employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.res = self.curr.execute(sql)
            self.conn.commit()

            if self.res:
                b_res = 'ok'
                return b_res

        except Exception as e:
            logging.info('\n<Error> update_submit_check_member_2 : ' + str(e))

        finally:
            self.__disconnect__()
    
    #####################################
    # check_add_check_member_self_list
    #####################################
    def check_add_check_member_self_list(self):
        
        self.__connect__()

        try:
            
            sql = f"select check_year , check_month , employee_name from check_member where department_id='1BA' order by no"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('\n<Error> check_add_check_member_self_list : ' + str(e))

        finally:
            self.__disconnect__()

    #######################
    # search_member_item
    #######################
    def search_member_item(self , item , a_user):
        
        self.__connect__()

        try:
            
            '''
            conn_str        = f"DRIVER={{SQL Server}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"select {item} from T_HR_Employee where EmployeeName='{a_user}'"
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]
            '''
            
            sql = f"select {item} from check_member where employee_name='{a_user}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('\n<Error> search_member_item : ' + str(e))

        finally:
            self.__disconnect__()
            #self.curr_mssql.close()
            #self.conn_mssql.close()

    #############################
    # show_work_time_total_val
    #############################
    def show_work_time_total_val(self , e_name , e_id , item):
        
        self.__connect__()

        try:
            
            s_sql = f"select sum({item}) from work_time where e_name='{e_name}' and e_id='{e_id}'"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchone()

            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('\n<Error> show_work_time_total_val : ' + str(e))

        finally:
            self.__disconnect__()

    ######################## 
    # show_work_time_list
    ########################
    def show_work_time_list(self , e_name , e_id):
        
        self.__connect__()

        try:
            
            s_sql = f"select b_date from work_time where e_name='{e_name}' and e_id='{e_id}' order by b_date desc"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('\n<Error> show_work_time_list : ' + str(e))

        finally:
            self.__disconnect__()

    ##########################
    # submit_work_time_form
    ##########################
    def submit_work_time_form(self , a_work_no , a_name , dep_id , b_date , total_time , normal_time , over_time , availability_time , a_work_station_1 , a_production_1 , a_product_no_1 , a_work_normal_time_1 , a_work_over_time_1 , a_work_availability_time_1 , a_work_remark_1 , a_work_station_2 , a_production_2 , a_product_no_2 , a_work_normal_time_2 , a_work_over_time_2 , a_work_availability_time_2 , a_work_remark_2 , a_work_station_3 , a_production_3 , a_product_no_3 , a_work_normal_time_3 , a_work_over_time_3 , a_work_availability_time_3 , a_work_remark_3 , a_work_station_4 , a_production_4 , a_product_no_4 , a_work_normal_time_4 , a_work_over_time_4 , a_work_availability_time_4 , a_work_remark_4 , a_work_station_5 , a_production_5 , a_product_no_5 , a_work_normal_time_5 , a_work_over_time_5 , a_work_availability_time_5 , a_work_remark_5 , a_work_station_6 , a_production_6 , a_product_no_6 , a_work_normal_time_6 , a_work_over_time_6 , a_work_availability_time_6 , a_work_remark_6 , a_work_station_7 , a_production_7 , a_product_no_7 , a_work_normal_time_7 , a_work_over_time_7 , a_work_availability_time_7 , a_work_remark_7 , a_work_station_8 , a_production_8 , a_product_no_8 , a_work_normal_time_8 , a_work_over_time_8 , a_work_availability_time_8 , a_work_remark_8 , a_work_station_9 , a_production_9 , a_product_no_9 , a_work_normal_time_9 , a_work_over_time_9 , a_work_availability_time_9 , a_work_remark_9 , a_work_station_10 , a_production_10 , a_product_no_10 , a_work_normal_time_10 , a_work_over_time_10 , a_work_availability_time_10 , a_work_remark_10 , a_work_station_11 , a_production_11 , a_product_no_11 , a_work_normal_time_11 , a_work_over_time_11 , a_work_availability_time_11 , a_work_remark_11 , a_work_station_12 , a_production_12 , a_product_no_12 , a_work_normal_time_12 , a_work_over_time_12 , a_work_availability_time_12 , a_work_remark_12):
        
        self.__connect__()
        
        try:
        
            s_sql = f"select * from work_time where e_id='{a_work_no}' and e_name='{a_name}' and b_date='{b_date}' and dep_id='{dep_id}'"
            self.curr.execute(s_sql)
            self.res = self.curr.fetchone()

            data  = b_date.split('-')
            r_year  = data[0]
            r_month = data[1]
            r_day   = data[2]
            
            if self.res is None:
                a_sql  = f"insert into work_time("
                a_sql += f"e_id , e_name , dep_id , b_date , total_time , normal_time , over_time , availability_time , r_year , r_month , r_day ," 
                a_sql += f"w_s_1 , w_s_1_product , w_s_1_num , w_s_1_normal_time , w_s_1_over_time , w_s_1_avail_time , w_s_1_remark , " 
                a_sql += f"w_s_2 , w_s_2_product , w_s_2_num , w_s_2_normal_time , w_s_2_over_time , w_s_2_avail_time , w_s_2_remark , " 
                a_sql += f"w_s_3 , w_s_3_product , w_s_3_num , w_s_3_normal_time , w_s_3_over_time , w_s_3_avail_time , w_s_3_remark , " 
                a_sql += f"w_s_4 , w_s_4_product , w_s_4_num , w_s_4_normal_time , w_s_4_over_time , w_s_4_avail_time , w_s_4_remark , " 
                a_sql += f"w_s_5 , w_s_5_product , w_s_5_num , w_s_5_normal_time , w_s_5_over_time , w_s_5_avail_time , w_s_5_remark , " 
                a_sql += f"w_s_6 , w_s_6_product , w_s_6_num , w_s_6_normal_time , w_s_6_over_time , w_s_6_avail_time , w_s_6_remark , " 
                a_sql += f"w_s_7 , w_s_7_product , w_s_7_num , w_s_7_normal_time , w_s_7_over_time , w_s_7_avail_time , w_s_7_remark , " 
                a_sql += f"w_s_8 , w_s_8_product , w_s_8_num , w_s_8_normal_time , w_s_8_over_time , w_s_8_avail_time , w_s_8_remark , " 
                a_sql += f"w_s_9 , w_s_9_product , w_s_9_num , w_s_9_normal_time , w_s_9_over_time , w_s_9_avail_time , w_s_9_remark , " 
                a_sql += f"w_s_10 , w_s_10_product , w_s_10_num , w_s_10_normal_time , w_s_10_over_time , w_s_10_avail_time , w_s_10_remark , " 
                a_sql += f"w_s_11 , w_s_11_product , w_s_11_num , w_s_11_normal_time , w_s_11_over_time , w_s_11_avail_time , w_s_11_remark , " 
                a_sql += f"w_s_12 , w_s_12_product , w_s_12_num , w_s_12_normal_time , w_s_12_over_time , w_s_12_avail_time , w_s_12_remark" 
                a_sql += f") value("
                a_sql += f"'{a_work_no}' , '{a_name}' , '{dep_id}' ,'{b_date}', '{total_time}' , '{normal_time}' , '{over_time}' , '{availability_time}' , '{r_year}' , '{r_month}' , '{r_day}' ,"
                a_sql += f"'{a_work_station_1}' , '{a_production_1}' , '{a_product_no_1}' , '{a_work_normal_time_1}' , '{a_work_over_time_1}' , '{a_work_availability_time_1}' , '{a_work_remark_1}' , "
                a_sql += f"'{a_work_station_2}' , '{a_production_2}' , '{a_product_no_2}' , '{a_work_normal_time_2}' , '{a_work_over_time_2}' , '{a_work_availability_time_2}' , '{a_work_remark_2}' , "
                a_sql += f"'{a_work_station_3}' , '{a_production_3}' , '{a_product_no_3}' , '{a_work_normal_time_3}' , '{a_work_over_time_3}' , '{a_work_availability_time_3}' , '{a_work_remark_3}' , "
                a_sql += f"'{a_work_station_4}' , '{a_production_4}' , '{a_product_no_4}' , '{a_work_normal_time_4}' , '{a_work_over_time_4}' , '{a_work_availability_time_4}' , '{a_work_remark_4}' , "
                a_sql += f"'{a_work_station_5}' , '{a_production_5}' , '{a_product_no_5}' , '{a_work_normal_time_5}' , '{a_work_over_time_5}' , '{a_work_availability_time_5}' , '{a_work_remark_5}' , "
                a_sql += f"'{a_work_station_6}' , '{a_production_6}' , '{a_product_no_6}' , '{a_work_normal_time_6}' , '{a_work_over_time_6}' , '{a_work_availability_time_6}' , '{a_work_remark_6}' , "
                a_sql += f"'{a_work_station_7}' , '{a_production_7}' , '{a_product_no_7}' , '{a_work_normal_time_7}' , '{a_work_over_time_7}' , '{a_work_availability_time_7}' , '{a_work_remark_7}' , "
                a_sql += f"'{a_work_station_8}' , '{a_production_8}' , '{a_product_no_8}' , '{a_work_normal_time_8}' , '{a_work_over_time_8}' , '{a_work_availability_time_8}' , '{a_work_remark_8}' , "
                a_sql += f"'{a_work_station_9}' , '{a_production_9}' , '{a_product_no_9}' , '{a_work_normal_time_9}' , '{a_work_over_time_9}' , '{a_work_availability_time_9}' , '{a_work_remark_9}' , "
                a_sql += f"'{a_work_station_10}' , '{a_production_10}' , '{a_product_no_10}' , '{a_work_normal_time_10}' , '{a_work_over_time_10}' , '{a_work_availability_time_10}' , '{a_work_remark_10}' , "
                a_sql += f"'{a_work_station_11}' , '{a_production_11}' , '{a_product_no_11}' , '{a_work_normal_time_11}' , '{a_work_over_time_11}' , '{a_work_availability_time_11}' , '{a_work_remark_11}' , "
                a_sql += f"'{a_work_station_12}' , '{a_production_12}' , '{a_product_no_12}' , '{a_work_normal_time_12}' , '{a_work_over_time_12}' , '{a_work_availability_time_12}' , '{a_work_remark_12}'"
                a_sql += f")" 

                res = self.curr.execute(a_sql)

            else:
                r_val = 'no'
                return r_val
        
        except Exception as e:
            logging.info('\n<Error> submit_work_time_form : ' + str(e))

        finally:
            self.__disconnect__()
    
    ################
    # search_item
    ################
    def search_item(self , item , a_user):
        
        self.__connect__()
        
        try:
            
            '''
            conn_str        = f"DRIVER={{SQL Server}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql        = f"select {item} from T_HR_Employee where EmployeeName='{a_user}'"
            self.curr_mssql.execute(self.sql)
            self.res        = self.curr_mssql.fetchone()

            return self.res[0]
            '''
        
            sql = f"select {item} from hr_a where employee_name='{a_user}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()
            
            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('\n<Error> search_item : ' + str(e))

        finally:
            self.__disconnect__()
            #self.curr_mssql.close()
            #self.conn_mssql.close()

    ################################
    # check_add_check_member_list
    ################################
    def check_add_check_member_list(self , employee_name):
        
        self.__connect__()
        
        try:
            
            sql = f"select check_year , check_month , employee_name from check_member where employee_name='{employee_name}' order by b_date"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()

            if self.res is not None:
                return self.res
            
            else:
                self.res = '沒考核紀錄。'
                return self.res
            
        
        except Exception as e:
            logging.info('\n<Error> check_add_check_member_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    #################################
    # check_add_check_member_data
    #################################
    def check_add_check_member_data(self , employee_name , check_year , check_month):
        try:
            self.__connect__()
            
            sql = f"select employee_name from check_member where employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:

                res_a = 'ok'
                return res_a
                
            else: 

                res_a = 'no'
                return res_a
        
        except Exception as e:
            logging.info('\n<Error> check_add_check_member_data : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # submit_add_check_member_data
    #################################
    def submit_add_check_member_data(self , employee_id , employee_name , department_id , department_name , job_title , b_date , end_date , check_year , check_month , self_num1_1 , self_num1_2 , self_num1_3 , self_num1_4 , self_num2_1 , self_num2_2 , self_num2_3 , self_num3_1 , self_num3_2 , self_num3_3 , self_num4_1 , self_num4_2 , self_num4_3 , self_num4_4 , self_num5_1 , self_num5_2 , self_num5_3 , self_num6_1 , self_num6_2 , self_num6_3  , self_total):
        
        self.__connect__()
        
        try:
            
            sql = f"select employee_name from check_member where employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:
                
                sql2 = f"insert into check_member(employee_id , employee_name , department_id , department_name , b_date , end_date , check_year , check_month , self_item_1_1 , self_item_1_2 , self_item_1_3 , self_item_1_4 , self_item_2_1 , self_item_2_2 , self_item_2_3 , self_item_3_1 , self_item_3_2 , self_item_3_3 , self_item_4_1 ,  self_item_4_2 ,  self_item_4_3 ,  self_item_4_4 ,  self_item_5_1 , self_item_5_2 , self_item_5_3 , self_item_6_1 , self_item_6_2 , self_item_6_3 , self_total , self_check) value('{employee_id}' , '{employee_name}' , '{department_id}' , '{department_name}' , '{b_date}' , '{end_date}' , '{check_year}' , '{check_month}' , '{self_num1_1}' , '{self_num1_2}' , '{self_num1_3}' , '{self_num1_4}' , '{self_num2_1}' , '{self_num2_2}' , '{self_num2_3}' , '{self_num3_1}' , '{self_num3_2}' , '{self_num3_3}' , '{self_num4_1}' , '{self_num4_2}' , '{self_num4_3}' , '{self_num4_4}' , '{self_num5_1}' , '{self_num5_2}' , '{self_num5_3}' , '{self_num6_1}' , '{self_num6_2}' , '{self_num6_3}'  , '{self_total}' , 'done')" 
                self.curr.execute(sql2)
                self.conn.commit()

        except Exception as e:
            logging.info('\n<Error> submit_add_check_member_data : ' + str(e))

        finally:
            self.__disconnect__() 
    
    #############################
    # submit_add_check_account
    #############################
    def submit_add_check_account(self , employee_id , employee_name , login_id , mobile , department_name , department_code , company_id , end_date):
        try:
            self.__connect__()
            
            sql = f"select employee_name from hr_a where employee_name='{employee_name}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:
                
                sql3 = f"select employee_name from hr_a where login_id='{login_id}'"
                self.curr.execute(sql3)
                self.res3 = self.curr.fetchone()

                if self.res3 is None:
                
                    sql2 = f"insert into hr_a(employee_id , employee_name , login_id , mobile , department_name , department_code , company_id , end_date) value('{employee_id}' , '{employee_name}' , '{login_id}' , '{mobile}' , '{department_name}' , '{department_code}' , '{company_id}' , '{end_date}')" 
                    self.curr.execute(sql2)

                    res_a = 'ok'
                    return res_a
                
                else:

                    res_a = 'no_login_id'
                    return res_a    
            
            else:

                res_a = 'no'
                return res_a
        
        except Exception as e:
            logging.info('\n<Error> submit_add_check_account : ' + str(e))

        finally:
            self.__disconnect__()

    #####################################
    # load_account_data_form_self_item
    #####################################
    def load_account_data_form_self_item(self , employee_id , employee_name , check_year , check_month):
        
        self.__connect__()

        try:
            
            sql = f"select self_item_1_1 , self_item_1_2 , self_item_1_3 , self_item_1_4 , self_item_2_1 , self_item_2_1 , self_item_2_3 , self_item_3_1 , self_item_3_2 , self_item_3_3 , self_item_4_1 , self_item_4_2 , self_item_4_3 , self_item_4_4 , self_item_5_1 , self_item_5_2 , self_item_5_3 , self_item_6_1 , self_item_6_2 , self_item_6_3 , self_total from check_member where employee_id='{employee_id}' and employee_name='{employee_name}' and check_year='{check_year}' and check_month='{check_month}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('\n<Error> load_account_data_form_self_item : ' + str(e))

        finally:
            self.__disconnect__()
    
    ################################
    # load_account_data_form_item
    ################################
    def load_account_data_form_item(self , item , employee_id):
        try:
            self.__connect__()
            
            sql = f"select {item}  from check_member where employee_id='{employee_id}' order by check_year desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('\n<Error> load_account_data_form_item : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # load_account_data_item
    ###########################
    def load_account_data_item(self , user):
        try:
            self.__connect__()
            
            sql = f"select employee_id , employee_name , end_date from hr_a where employee_id='{user}' and department_code like '1B%' and job_title_name != '經理' "
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('\n<Error> load_account_data_item : ' + str(e))

        finally:
            self.__disconnect__()

    ############################
    # factory_check_form_item
    ############################
    def factory_check_form_item(self , user):
        
        self.__connect__()
        
        try:
            
            sql = f"select job_title_name from hr_a where employee_name='{user}'"
            self.curr.execute(sql)
            self.res = self.curr.fetchone()
            
            if self.res is not None:
                return self.res[0]
        
        except Exception as e:
            logging.info('\n<Error> factory_check_form_item : ' + str(e))

        finally:
            self.__disconnect__()

    ############################
    # factory_check_form_list
    ############################
    def factory_check_form_list(self):
        
        self.__connect__()
        
        try:
            
            sql = "select employee_name from hr_a where department_code like '1B%' and job_title_name != '經理' order by no desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('\n<Error> factory_check_form_list : ' + str(e))

        finally:
            self.__disconnect__()
    
    ##############################
    # factory_work_account_list
    ##############################
    def factory_work_account_list(self):
        try:
            self.__connect__()
            
            sql = "select a_user , a_name , a_pwd , a_status , a_work_no from account where a_lv='3' and a_position='生二部' order by no desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('\n<Error> factory_work_account_list : ' + str(e))

        finally:
            self.__disconnect__()

    #################################
    # alter_work_record_list_detail
    #################################
    def alter_work_record_list_detail(self , title , dep_id):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 

            sql  = f"select w_title , w_content , c_date , w_start , w_end , w_status , w_update , w_place from work_record_form where w_title='{title}' and w_dep_id='{dep_id}'"
 
            self.curr2.execute(sql)
            res = self.curr2.fetchall()

            return res
        
        except Exception as e:
            logging.info(f'\n<Error> alter_work_record_list_detail : {str(e)}\n')

        finally:
            self.__disconnect2__()
    
    #################################
    # del_work_record_list_detail
    #################################
    def del_work_record_list_detail(self , title):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            

            sql  = f"Delete from work_record_form where w_title='{title}'"
 
            res = self.curr2.execute(sql)

            return res
        
        except Exception as e:
            logging.info(f'\n<Error> del_work_record_list_detail : {str(e)}\n')

        finally:
            self.__disconnect2__()
    
    #################################
    # load_work_record_list_detail
    #################################
    def load_work_record_list_detail(self , title , dep_id):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            

            sql  = '''
                        SELECT w_content , w_start , w_end , w_status , w_update , w_title from work_record_form 
                        where w_title=%s and w_dep_id=%s
                   ''' 

            self.curr2.execute(sql , (title,dep_id,))
            res = self.curr2.fetchall()

            return res
        
        except Exception as e:
            logging.info(f'\n<Error> load_work_record_list_detail : {str(e)}\n')

        finally:
            self.__disconnect2__()

    ##########################
    # load_work_record_list
    ##########################
    def load_work_record_list(self , dep_id):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            if dep_id == '資訊部':
                sql  = f'SELECT w_dep_id, COUNT(*), GROUP_CONCAT(w_title ORDER BY w_start DESC) AS titles '
                sql += f'FROM work_record_form where w_dep="{dep_id}" or w_dep="資訊部(駐工廠)"'
                sql += f'GROUP BY w_dep_id'
            else:
                sql  = f'SELECT w_dep_id, COUNT(*), GROUP_CONCAT(w_title ORDER BY w_start DESC) AS titles '
                sql += f'FROM work_record_form where w_dep="{dep_id}" '
                sql += f'GROUP BY w_dep_id'
            
            self.curr2.execute(sql)
            res = self.curr2.fetchall()

            work_record_list = []

            for val in res:
                # Split the titles back into a list (if needed, or you can just keep them as a single string)
                titles = val[2].split(',') if val[2] else []
                work_record_list.append((val[0], val[1], titles))

            return work_record_list
        
        except Exception as e:
            logging.info(f'\n<Error> load_work_record_list : {str(e)}\n')

        finally:
            self.__disconnect2__()
    
    #######################################################
    # bpm_information_finish_by_dep_kind_total_cost_user
    #######################################################
    def bpm_information_finish_by_dep_kind_total_cost_user(self , dep , kind , user):
        
        self.__connect_mssql__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            kind = str(kind).replace(" - " , "_") + ','
            sql  = f"SELECT format(sum(ITEM98),'N0','en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM147='{kind}' and ITEM35='True'"
 
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchone()

            return res[0]
        
        except Exception as e:
            logging.info(f'\n<Error> bpm_information_finish_by_dep_kind_total_cost_user : {str(e)}\n')

        finally:
            self.__disconnect_mssql__()

    ##################################################
    # bpm_information_finish_by_dep_kind_total_cost
    ##################################################
    def bpm_information_finish_by_dep_kind_total_cost(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            kind = str(kind).replace(" - " , "_") + ','
            sql  = f"SELECT format(sum(ITEM98),'N0','en-US') FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM147='{kind}' and ITEM35='True'"
 
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchone()

            return res[0]
        
        except Exception as e:
            logging.info(f'\n<Error> bpm_information_finish_by_dep_kind_total_cost : {str(e)}\n')

        finally:
            self.__disconnect_mssql__()

     ###########################################
    # bpm_information_finish_by_dep_kind_user
    ###########################################
    def bpm_information_finish_by_dep_kind_user(self , dep , kind , user):
        
        self.__connect_mssql__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            kind = str(kind).replace(" - " , "_") + ','

            sql = f"SELECT ITEM11 , format(ITEM98 , 'N0' , 'en-US') , format(ITEM145,'N0','en-US') , ITEM75 FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM18='{user}' and ITEM147='{kind}' and ITEM35='True' order by ITEM11 asc"
 
            print(sql)
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()

            res_all = []

            for val in res:
                res_all.append(( val[0] , val[1] , val[2] , val[3]))

            return res_all
        
        except Exception as e:
            logging.info(f'\n<Error> bpm_information_finish_by_dep_kind_user : {str(e)}\n')

        finally:
            self.__disconnect_mssql__()
    
    #######################################
    # bpm_expenditure_finish_by_dep_kind
    #######################################
    def bpm_expenditure_finish_by_dep_kind(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            kind = str(kind).replace(" - " , "_") + ','

            sql = f"SELECT ITEM11 , format(ITEM98 , 'N0' , 'en-US') , format(ITEM145,'N0','en-US') , ITEM75 FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM147='{kind}' and ITEM35='True' order by ITEM11 asc"
 
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()

            res_all = []

            for val in res:
                res_all.append(( val[0] , val[1] , val[2] , val[3]))

            return res_all
        
        except Exception as e:
            logging.info(f'\n<Error> bpm_expenditure_finish_by_dep_kind : {str(e)}\n')

        finally:
            self.__disconnect_mssql__()
    
    #######################################
    # bpm_information_finish_by_dep_kind
    #######################################
    def bpm_information_finish_by_dep_kind(self , dep , kind):
        
        self.__connect_mssql__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            kind = str(kind).replace(" - " , "_") + ','

            sql = f"SELECT ITEM11 , format(ITEM98 , 'N0' , 'en-US') , format(ITEM145,'N0','en-US') , ITEM75 FROM ART01231708483412487_INS where ITEM19='{dep}' and ITEM147='{kind}' and ITEM35='True' order by ITEM11 asc"
 
            self.curr_mssql.execute(sql)
            res = self.curr_mssql.fetchall()

            res_all = []

            for val in res:
                res_all.append(( val[0] , val[1] , val[2] , val[3]))

            return res_all
        
        except Exception as e:
            logging.info(f'\n<Error> bpm_information_finish_by_dep_kind : {str(e)}\n')

        finally:
            self.__disconnect_mssql__()

    ##################################
    # work_record_final_list_by_dep
    ##################################
    def work_record_final_list_by_dep(self , dep_id):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            

            sql  = f"SELECT w_title , w_update , w_dep_id from work_record_form where w_dep='{dep_id}' order by w_update desc limit 0,10"
 
            self.curr2.execute(sql)
            res = self.curr2.fetchall()

            return res
        
        except Exception as e:
            logging.info(f'\n<Error> work_record_list_by_dep : {str(e)}\n')

        finally:
            self.__disconnect2__()

    ###########################
    # work_record_final_list
    ###########################
    def work_record_final_list(self , dep_id , user):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            

            sql  = f"SELECT w_title , w_update from work_record_form where w_dep='{dep_id}' and w_dep_id='{user}' order by w_update desc limit 0,6"
 
            self.curr2.execute(sql)
            res = self.curr2.fetchall()

            return res
        
        except Exception as e:
            logging.info(f'\n<Error> work_record_list : {str(e)}\n')

        finally:
            self.__disconnect2__()
    
    ####################
    # work_record_list
    ####################
    def work_record_list(self , dep_id):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year    = time.strftime("%Y" , time.localtime()) 
            r_month   = time.strftime("%m" , time.localtime()) 
            r_day     = time.strftime("%d" , time.localtime()) 
            r_time    = time.strftime("%H:%M:%S" , time.localtime()) 
            
            if dep_id == '資訊部':
                sql  = f'SELECT w_dep_id, COUNT(*), GROUP_CONCAT(w_title ORDER BY w_start DESC) AS titles '
                sql += f'FROM work_record_form where w_dep="{dep_id}" or w_dep="資訊部(駐工廠)"'
                sql += f'GROUP BY w_dep_id'
            else:
                sql  = f'SELECT w_dep_id, COUNT(*), GROUP_CONCAT(w_title ORDER BY w_start DESC) AS titles '
                sql += f'FROM work_record_form where w_dep="{dep_id}" '
                sql += f'GROUP BY w_dep_id'
 
            self.curr2.execute(sql)
            res = self.curr2.fetchall()

            work_record_list = []

            for val in res:
                # Split the titles back into a list (if needed, or you can just keep them as a single string)
                titles = val[2].split(',') if val[2] else []
                work_record_list.append((val[0], val[1], titles))

            return work_record_list
        
        except Exception as e:
            logging.info(f'\n<Error> work_record_list : {str(e)}\n')

        finally:
            self.__disconnect2__()

    ####################
    # alter_work_record
    ####################
    def alter_work_record(self , w_title , w_year , w_month , w_day , w_place , w_start , w_end , w_status , w_dep , w_user , w_new_work_record_content):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year   = time.strftime("%Y" , time.localtime()) 
            r_month  = time.strftime("%m" , time.localtime()) 
            r_day    = time.strftime("%d" , time.localtime()) 
            r_time   = time.strftime("%H:%M:%S" , time.localtime()) 

            alter_sql  = f"update work_record_form set w_update='{now_time}' , w_place='{w_place}' , w_end='{w_end}' , w_content='{w_new_work_record_content}' , w_status='{w_status}' " 
            alter_sql += f"where w_title='{w_title}' and w_dep='{w_dep}'"

            self.curr2.execute(alter_sql)
            self.conn2.commit()
        
        except Exception as e:
            logging.info(f'\n<Error> alter_work_record : {str(e)}\n')

        finally:
            self.__disconnect2__()
    
    #######################
    # search_work_record
    #######################
    def search_work_record(self , s_s_date , s_e_date , s_w_kind , dep_name):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year   = time.strftime("%Y" , time.localtime()) 
            r_month  = time.strftime("%m" , time.localtime()) 
            r_day    = time.strftime("%d" , time.localtime()) 
            r_time   = time.strftime("%H:%M:%S" , time.localtime()) 
            
            
            f_res = []

            sql = f"""
            SELECT wfr.w_dep_id, wfr.w_start, wfr.w_content
            FROM work_record_form wfr
            WHERE wfr.w_kind='{s_w_kind}' AND wfr.w_start>='{s_s_date}' AND wfr.w_start<='{s_e_date}' AND wfr.w_dep='{dep_name}'
            ORDER BY wfr.w_dep_id DESC, wfr.w_start DESC
            """
            self.curr2.execute(sql)
            res = self.curr2.fetchall()

            for val in res:
                dep_id = val[0]
                w_start = val[1]
                w_content = val[2]  

                f_res.append((dep_id, w_start, w_content))

            return f_res
        
        except Exception as e:
            logging.info(f'\n<Error> search_work_record : {str(e)}\n')

        finally:
            self.__disconnect2__()
    
    ####################
    # add_work_record
    ####################
    def add_work_record(self , w_title , w_year , w_month , w_day , w_place , w_start , w_end , w_status , w_dep , w_user , w_new_work_record_content , w_kind):
        
        self.__connect2__()
        
        try:
            
            ### time record
            now_time  = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            now_month = time.strftime("%Y-%m-%d" , time.localtime()) 
            r_year   = time.strftime("%Y" , time.localtime()) 
            r_month  = time.strftime("%m" , time.localtime()) 
            r_day    = time.strftime("%d" , time.localtime()) 
            r_time   = time.strftime("%H:%M:%S" , time.localtime()) 
            

            sql = f"select w_title from work_record_form where w_title='{w_title}' and w_kind='{w_kind}' and w_dep='{w_dep}'"
            self.curr2.execute(sql)
            self.res = self.curr2.fetchone()

            if self.res is None:

                add_sql  = f"insert into work_record_form(c_date , c_time , c_d_time , w_dep , w_dep_id , w_title , w_place , w_start , w_end , w_content , w_status , w_kind) " 
                add_sql += f"value('{now_month}','{r_time}','{now_time}','{w_dep}','{w_user}','{w_title}','{w_place}','{w_start}','{w_end}','{w_new_work_record_content}','{w_status}','{w_kind}')"

                self.curr2.execute(add_sql)
                self.conn2.commit()
        
        except Exception as e:
            logging.info(f'\n<Error> add_work_record : {str(e)}\n')

        finally:
            self.__disconnect2__()

    ################
    # add_account
    ################
    def add_account(self , a_date , a_name , a_work_no , a_position , a_status , a_user):
        try:
            self.__connect__()
            
            ### time record
            now_time = time.strftime("%Y-%m-%d %H:%M:%S" , time.localtime()) 
            r_year   = time.strftime("%Y" , time.localtime()) 
            r_month  = time.strftime("%m" , time.localtime()) 
            r_day    = time.strftime("%d" , time.localtime()) 
            r_time   = time.strftime("%H:%M:%S" , time.localtime()) 
            a_pwd    = 'm' + a_work_no
            a_lv     = 3

            sql = "select a_user from account where a_user='{0}'".format(a_user)
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res is None:
                
                if a_status == '使用':
                    a_status = 'run'
                else:
                    a_status = 'stop'

                add_sql  = "insert into account(r_year , r_month , r_day , r_time , a_name , a_pwd , a_work_no , a_position , a_status , a_lv , a_user)"
                add_sql += " value ('{0}' , '{1}' ,'{2}' ,'{3}' ,'{4}' ,'{5}' ,'{6}' ,'{7}' ,'{8}' ,'{9}' , '{10}')".format(r_year , r_month , r_day , r_time , a_name , a_pwd , a_work_no , a_position , a_status , a_lv , a_user)

                self.curr.execute(add_sql)
                return True
                
            else:
                return False
        
        except Exception as e:
            logging.info('\n<Error> add_account : ' + str(e))

        finally:
            self.__disconnect__()

    #########################
    # factory_work_position
    #########################
    def factory_work_position(self):
        try:
            self.__connect__()
            
            sql = "select distinct c_content from work_position order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('\n<Error> factory_work_position : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # factory_work_station_3
    ###########################
    def factory_work_station_3(self):
        
        self.__connect__()
        
        try:
            
            sql = "select distinct c_content from work_station_3 order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            if self.res is not None:
                return self.res
        
        except Exception as e:
            logging.info('\n<Error> factory_work_station_3 : ' + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # factory_work_station_1
    ###########################
    def factory_work_station_1(self):
        try:
            self.__connect__()
            
            sql = "select distinct c_content from work_station_1 order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('\n<Error> factory_work_statio_1 : ' + str(e))

        finally:
            self.__disconnect__()

    #########################
    # factory_work_station
    #########################
    def factory_work_station(self):
        try:
            self.__connect__()
            
            sql = "select distinct c_content from work_station order by e_name desc"
            self.curr.execute(sql)
            self.res = self.curr.fetchall()
            
            return self.res
        
        except Exception as e:
            logging.info('\n<Error> factory_work_station : ' + str(e))

        finally:
            self.__disconnect__()

    #####################
    # check_login_code
    #####################
    def check_login_code(self,user,login_code):
        
        try:
            self.user = user
            self.login_code = login_code

            self.__connect__()

            sql = "select login_code from login_out_record where a_user='{0}' order by no desc limit 0,1".format(self.user)
            self.curr.execute(sql)
            self.res = self.curr.fetchone()

            if self.res[0] == self.login_code:
                return 'ok'

        except Exception as e:
            logging.info("\n<Error> check login code : " + str(e))

        finally:
            self.__disconnect__()

    ###########
    # dep_id
    ###########
    def dep_id(self,user,pwd):
        
        self.__connect_mssql__()
        
        try:
            self.login_id = user
            self.mobile   = pwd
            
            #sql = f"select DepartmentID from T_HR_Employee where loginID='{self.login_id}' and Mobile='{self.mobile}'"
            sql = f"select DepartmentID from T_HR_Employee where loginID='{self.login_id}'"
            self.curr_mssql.execute(sql)
            dep_id = self.curr_mssql.fetchone()

            sql2 = f"select DepartmentName from T_HR_Department where DepartmentID='{dep_id[0]}'"
            self.curr_mssql.execute(sql2)
            dep_name = self.curr_mssql.fetchone()

            return dep_name[0]
            
            
            '''
            conn_str = f"DRIVER={{SQL Server}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']}"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            self.sql = f"select DepartmentID from T_HR_Employee where loginID='{self.login_id}' and Mobile='{self.mobile}'"
            self.curr_mssql.execute(self.sql)
            self.res = self.curr_mssql.fetchone()

            self.sql2 = f"select UpperDepartmentID from T_HR_Department where DepartmentID='{self.res[0]}'"
            self.curr_mssql.execute(self.sql2)
            self.res2 = self.curr_mssql.fetchone()

            return self.res2
            '''
            '''
            self.__connect__()

            self.sql = f"select department_code from hr_a where login_id='{self.login_id}' and mobile='{self.mobile}'"
            self.curr.execute(self.sql)
            self.res        = self.curr.fetchone()
            self.dep_code   = self.res[0]
            self.r_dep_code = self.dep_code[0:2]
            '''
            '''
            ###########
            # 生一部
            ###########
            if self.r_dep_code == '1A':
                return self.r_dep_code
            ###########
            # 生二部
            ###########
            elif self.r_dep_code == '1B':
                return self.r_dep_code
            ###########
            # 生三部
            ###########
            elif self.r_dep_code == '1K':
                return self.r_dep_code
            ##############################
            # 剩下都依照原本部門代號顯示
            ##############################
            else:
                return self.dep_code
            '''

        except Exception as e:
            logging.error("\n<Error> dep_id : " + str(e))

        finally:
            self.__disconnect_mssql__()

    ##################
    # login_ad_user
    ##################
    def login_ad_user(self , user):

        # link WMS AD Server
        self.__connect_mssql_wms_account__()
        
        try:
            sql = '''
                    select u_name from wms_ad_account where u_name=?;
                  '''
            self.curr_mssql_wms_account.execute(sql , (user,))
            res = self.curr_mssql_wms_account.fetchone()

            return res[0]

        except Exception as e:
                logging.info(f"\n<Error>login_ad_user : {str(e)}\n")
        finally:
            #conn_mssql_wms.close()
            self.__connect_mssql_wms_account__()
    
    ##############
    # wms_login
    ##############
    def wms_login(self,user,pwd):
        
        try:
            ### variables
            login_id  = user
            login_pwd = pwd

            #####################################
            #
            # connect AD Server 
            # IP : 192.168.1.10:389
            #
            #####################################
            ad_server = parameter.ad_server['host']
            ad_domain = parameter.ad_server['domain']
            username  = login_id
            password  = login_pwd
            
            server = Server(ad_server, get_info=ALL, use_ssl=False)  
            
            try:
                conn = Connection(server, user=f'{ad_domain}\\{username}', password=password, authentication='SIMPLE', auto_bind=True)
                
                print("\n------------------------------------------------------------------------------")
                print(f"WMS {username} login , AD : Authentication successful.")

                search_base   = 'dc=otsukatw,dc=corp'
                search_filter = f'(sAMAccountName={username})'

                ### 查询
                conn.search(search_base, search_filter, attributes=['cn', 'mail' , 'userAccountControl', 'lockoutTime'])

                ### 检查是否找到了用户
                if conn.entries:
                    username = conn.entries[0]
                    print(f"Found user : {username.cn}")
                    print(f"Email : {username.mail}")
                    
                    uac = username.userAccountControl.value

                    ### 检查用户是否被啟用
                    if uac and (uac & 2) == 0:  # 2 是 'ACCOUNTDISABLE' 標誌
                        print(f'{username.cn} , 帳號已啟用。')
                        print("------------------------------------------------------------------------------\n")
                        
                        self.check_res = 1
                        return self.check_res
                    else:
                        print(f'{username.cn} , 帳號被停用。')
                        print("------------------------------------------------------------------------------\n")
                        self.check_res = 0
                        return self.check_res

            except LDAPException as e:  
                logging.info(f"Authentication failed : {str(e)}\n")
            except Exception as e:
                logging.info(f"An error occurred: {str(e)}\n")
            finally:
                if 'conn' in locals() and conn.bound:
                    conn.unbind()

        except Exception as e:
            logging.info(f"\n<Error> login : {str(e)}\n")

        finally:
            pass

    ##########
    # login
    ##########
    def login(self,user,pwd):
        
        try:
            ### variables
            login_id  = user
            login_pwd = pwd

            #####################################
            #
            # connect AD Server 
            # IP : 192.168.1.10:389
            #
            #####################################
            ad_server = parameter.ad_server['host']
            ad_domain = parameter.ad_server['domain']
            username  = login_id
            password  = login_pwd
            
            server = Server(ad_server, get_info=ALL, use_ssl=False)  
            
            try:
                conn = Connection(server, user=f'{ad_domain}\\{username}', password=password, authentication='SIMPLE', auto_bind=True)
                
                print("\n------------------------------------------------------------------------------")
                print(f"Otsuka platfrom {username} login , AD : Authentication successful.")

                search_base   = 'dc=otsukatw,dc=corp'
                search_filter = f'(sAMAccountName={username})'

                ### 查询
                conn.search(search_base, search_filter, attributes=['cn', 'mail' , 'userAccountControl', 'lockoutTime'])

                ### 检查是否找到了用户
                if conn.entries:
                    username = conn.entries[0]
                    print(f"Found user : {username.cn}")
                    print(f"Email : {username.mail}")
                    
                    uac = username.userAccountControl.value

                    ### 检查用户是否被啟用
                    if uac and (uac & 2) == 0:  # 2 是 'ACCOUNTDISABLE' 標誌
                        print(f'{username.cn} , 帳號已啟用。')
                        print("------------------------------------------------------------------------------\n")
                        
                        self.check_res = 1
                        return self.check_res
                    else:
                        print(f'{username.cn} , 帳號被停用。')
                        print("------------------------------------------------------------------------------\n")
                        self.check_res = 0
                        return self.check_res

            except LDAPException as e:  
                logging.info(f"Authentication failed : {str(e)}\n")
            except Exception as e:
                logging.info(f"An error occurred: {str(e)}\n")
            finally:
                if 'conn' in locals() and conn.bound:
                    conn.unbind()

        except Exception as e:
            logging.info(f"\n<Error> login : {str(e)}\n")

        finally:
            pass
        
    #################
    # login_record   
    ################# 
    def login_record(self,user,login_code,r_time,ip):
        
        try:
            self.user       = user
            self.login_code = login_code
            self.r_time     = r_time
            self.ip         = ip

            self.__connect__()

            self.sql2 = "insert into login_out_record(a_user,login_code,login_time,login_ip) value('{0}','{1}','{2}','{3}')".format(self.user , self.login_code , self.r_time , self.ip)
            self.curr.execute(self.sql2)

        except Exception as e:
            logging.info("\n<Error> login record : " + str(e))

        finally:
            self.__disconnect__()
    
    ###########################
    # otsuka_contract_by_kind
    ###########################
    def otsuka_contract_by_kind(self):
        
        self.__connect__()
        
        try:
            
            contract_sql = f"SELECT * FROM `otsuka_contract` order by kind desc"
            self.curr.execute(contract_sql)
            res = self.curr.fetchall()

            return res

        except Exception as e:
            logging.info("\n<Error> otsuka_contract_by_date : " + str(e))

        finally:
            self.__disconnect__()

    ###########################
    # otsuka_contract_summary
    ###########################
    def otsuka_contract_summary(self):
        
        self.__connect__()
        
        try:
            
            contract_sql = f"SELECT kind , count(*) , round(sum(cost)) FROM `otsuka_contract` group by kind order by b_date desc"
            self.curr.execute(contract_sql)
            res = self.curr.fetchall()

            return res

        except Exception as e:
            logging.info("\n<Error> otsuka_contract_summary : " + str(e))

        finally:
            self.__disconnect__()

    #############################
    # it_annual_budget_by_date
    #############################
    def it_annual_budget_by_date(self):
        
        self.__connect2__()
        
        try:
            
            contract_sql = f"SELECT * FROM `it_annual_budget` order by b_date desc"
            self.curr.execute(contract_sql)
            res = self.curr.fetchall()

            return res

        except Exception as e:
            logging.info("\n<Error> it_annual_budget_by_date : " + str(e))

        finally:
            self.__disconnect2__()

    ###########################
    # otsuka_contract_by_date
    ###########################
    def otsuka_contract_by_date(self):
        
        self.__connect__()
        
        try:
            
            contract_sql = f"SELECT * FROM `otsuka_contract` order by b_date desc"
            self.curr.execute(contract_sql)
            self.res = self.curr.fetchall()

            return self.res

        except Exception as e:
            logging.info("\n<Error> otsuka_contract_by_date : " + str(e))

        finally:
            self.__disconnect__()

    #######################################################
    # it_annual_budget_year_total_by_kind_remaining_cost
    #######################################################
    def it_annual_budget_year_total_by_kind_remaining_cost(self , year):
        
        self.__connect2__()
        
        try:
            
            year_total_sql = f"SELECT kind , format(REPLACE(total_cost,',','')-REPLACE(use_cost,',',''),0) FROM `it_annual_budget_kind_remaining_cost` WHERE c_year='{year}'"
            self.curr2.execute(year_total_sql)
            self.res = self.curr2.fetchall()

            if self.res is not None:
                
                return self.res

        except Exception as e:
            logging.info(f"\n<Error> it_annual_budget_year_total_by_kind_remaining_cost : {str(e)}\n")

        finally:
            self.__disconnect2__()


    #######################################################
    # it_annual_budget_year_total_by_kind_use_cost
    #######################################################
    def it_annual_budget_year_total_by_kind_use_cost(self , year):
        
        self.__connect2__()
        
        try:
            
            year_total_sql = f"SELECT kind , format(sum(use_cost),0) , count(*) FROM `it_annual_budget_kind_use_cost` where c_year={year}  group by kind"
            self.curr2.execute(year_total_sql)
            self.res = self.curr2.fetchall()

            if self.res is not None:

                for val in self.res:
                    
                    check_kind_sql = f"select * from it_annual_budget_kind_remaining_cost where c_year='{year}' and kind='{val[0]}'"
                    self.curr2.execute(check_kind_sql)
                    self.check_res2 = self.curr2.fetchone()

                    if self.check_res2 is not None:
                        save_check_sql = f"update it_annual_budget_kind_remaining_cost set use_cost='{val[1]}' where c_year='{year}' and kind='{val[0]}'"
                        self.curr2.execute(save_check_sql)

                return self.res

        except Exception as e:
            logging.info(f"\n<Error> it_annual_budget_year_total_by_kind_use_cost : {str(e)}\n")

        finally:
            self.__disconnect2__()

    ########################################
    # it_annual_budget_year_total_by_kind
    ########################################
    def it_annual_budget_year_total_by_kind(self , year):
        
        self.__connect2__()
        
        try:
            
            year_total_sql = f"SELECT kind , count(*) , format(sum(annual_budget_year),0) FROM `it_annual_budget` where build_year='{year}' group by kind"
            self.curr2.execute(year_total_sql)
            self.res = self.curr2.fetchall()

            if self.res is not None:

                for val in self.res:
                    
                    check_kind_sql = f"select * from it_annual_budget_kind_remaining_cost where c_year='{year}' and kind='{val[0]}'"
                    self.curr2.execute(check_kind_sql)
                    self.check_res2 = self.curr2.fetchone()

                    if self.check_res2 is None:
                        save_check_sql = f"insert into it_annual_budget_kind_remaining_cost(c_year,kind,total_cost) value('{year}' , '{val[0]}' , '{val[2]}')"
                        self.curr2.execute(save_check_sql)
                
                return self.res

        except Exception as e:
            logging.info(f"\n<Error> it_annual_budget_year_total_by_kind : {str(e)}\n")

        finally:
            self.__disconnect2__()

    ################################
    # it_annual_budget_year_total
    ################################
    def it_annual_budget_year_total(self):
        
        self.__connect2__()
        
        try:
            
            year_total_sql = "SELECT build_year , format(sum(annual_budget_year),0) FROM `it_annual_budget` group by build_year order by build_year desc"
            self.curr2.execute(year_total_sql)
            self.res = self.curr2.fetchall()

            if self.res is not None:
                
                return self.res

        except Exception as e:
            logging.info(f"\n<Error> it_annual_budget_year_total : {str(e)}\n")

        finally:
            self.__disconnect2__()

    ##########################################
    # it_annual_budget_otsuka_holdings_list
    ##########################################
    def it_annual_budget_otsuka_holdings_list(self):
        
        self.__connect2__()
        
        try:
            
            otsuka_holdings_sql = "SELECT otsuka_holdings FROM `it_annual_budget` group by otsuka_holdings order by c_date desc"
            self.curr2.execute(otsuka_holdings_sql)
            self.res = self.curr2.fetchall()

            if self.res is not None:
                return self.res

        except Exception as e:
            logging.info(f"\n<Error> it_annual_budget_otsuka_holdings_list : {str(e)}\n")

        finally:
            self.__disconnect2__()
    
    ###############################
    # it_annual_budget_kind_list
    ###############################
    def it_annual_budget_kind_list(self):
        
        self.__connect2__()
        
        try:
            
            kind_sql = "SELECT kind FROM `it_annual_budget` group by kind order by c_date desc"
            self.curr2.execute(kind_sql)
            self.res = self.curr2.fetchall()

            if self.res is not None:
                return self.res

        except Exception as e:
            logging.info("\n<Error> it_annual_budget_kind_list : " + str(e))

        finally:
            self.__disconnect2__()

    ############################
    # submit_it_annual_budget
    ############################
    def submit_it_annual_budget(self , it_annual_budget_date , it_annual_budget_build_year , it_annual_budget_kind , it_annual_budget_title , it_annual_budget_otsuka_holdings , it_annual_budget_year , it_annual_budget_remaining_now , it_annual_budget_comment):
        
        self.__connect2__()
        
        try:
            ### time record
            now_time = time.strftime("%H:%M:%S" , time.localtime()) 
            
            b_date          = it_annual_budget_date
            b_build_year    = it_annual_budget_build_year
            kind            = it_annual_budget_kind
            title           = it_annual_budget_title
            otsuka_holdings = it_annual_budget_otsuka_holdings
            year            = it_annual_budget_year
            remaining_now   = it_annual_budget_remaining_now
            comment         = it_annual_budget_comment

            c_d_time        = b_date + " " + now_time 
            data            = str(b_date).split('-')
            c_year          = data[0]
            c_month         = data[1]
            c_day           = data[2]

            ### check kind and title 
            check_it_annual_budget_sql = f"select * from it_annual_budget where kind='{kind}' and title='{title}' and build_year='{b_build_year}'"
            self.curr2.execute(check_it_annual_budget_sql)
            check_it_annual_budget_res = self.curr2.fetchone()

            if check_it_annual_budget_res is None:
                it_annual_budget_sql  = f"insert into it_annual_budget(c_date , c_year , c_month , c_day , c_time , c_d_time , kind , title , otsuka_holdings , annual_budget_year , remaining_now_annual_budget , comment , build_year) value( "
                it_annual_budget_sql += f"'{b_date}' , '{c_year}' , '{c_month}' , '{c_day}' , '{now_time}' , '{c_d_time}' , '{kind}' , '{title}' , '{otsuka_holdings}' , '{year}' , '{remaining_now}' , '{comment}' , '{b_build_year}') "
                self.curr2.execute(it_annual_budget_sql)
                self.conn2.commit()

        except Exception as e:
            logging.info("\n<Error> submit_it_annual_budget : " + str(e))

        finally:
            pass
            #self.__disconnect2__()
    
    #######################################
    # convert_decimal
    #######################################
    def convert_decimal(self , obj):
        if isinstance(obj, list):
            return [self.convert_decimal(item) for item in obj]
        elif isinstance(obj, dict):
            return {key: self.convert_decimal(value) for key, value in obj.items()}
        elif isinstance(obj, self.Decimal):
            return float(obj)  
        else:
            return obj

    #######################################
    # convert_to_numeric
    #######################################
    def convert_to_numeric(self,value):
        try:
            return float(value) if value not in (None, '') else None
        except ValueError:
            print(f"無法轉換為數字類型：{value}")
            return None

    #######################################
    # wms_add_erp_picking_list_data_1
    #######################################
    def wms_add_erp_picking_list_data_1(self,i1,i2,i3,i4):
        
        self.__connect_mssql_erp__()
        
        try:
            
            ##################
            # 調撥單 & 退料單
            ##################
            db_h = f"Leader.dbo.INVTA"
            db_b = f"Leader.dbo.INVTB"

            sql1 = f" INSERT INTO {db_h} (TA001, TA002, TA003, TA009) VALUES (?,?,?,?)"
            self.curr_mssql_erp.execute(sql1 , (i1,i2,i3,i4,))
            self.curr_mssql_erp.commit()

            return 'ok'

            ##########
            # 調撥單
            ##########
                # -- INVTA調撥單頭
                #sql1 = f'''
                #        INSERT INTO {db_h} (COMPANY, CREATOR, USR_GROUP, CREATE_DATE, MODIFIER, 
                #                            MODI_DATE, FLAG, TA001, TA002, TA003, 
                #                            TA004, TA005, TA006, TA008, TA009, 
                #                            TA010, TA011, TA012, TA014, TA016,
                #                            TA017, TA018, TA019, TA021)
                #                    VALUES (?,?,?,?,?,
                #                            ?,?,?,?,?,
                #                            ?,?,?,?,?,
                #                            ?,?,?,?,?,
                #                           ?,?,?,?)
                #'''
                #self.curr_mssql_erp.execute(sql1, (i1,i2,i3,i4,i5,i6,i7,i8,i9,i10,i11,i12,i13,i15,i16,i17,i18,i19,i21,i22,i23,i24,i25,i26,))
                #self.curr_mssql_erp.commit()
                
                # -- 調撥單身
                #sql2 = f'''
                #    INSERT INTO {db_b} (COMPANY, CREATOR, USR_GROUP, CREATE_DATE, MODIFIER, 
                #                        MODI_DATE, FLAG, TB001, TB002, TB003, 
                #                        TB004, TB005, TB006, TB007, TB008, 
                #                        TB009, TB012, TB013, TB014, TB017, 
                #                        TB018)
                #                VALUES (?,?,?,?,?,
                #                        ?,?,?,?,?,
                #                        ?,?,?,?,?,
                #                        ?,?,?,?,?,
                #                        ?)
                #'''
                #self.curr_mssql_erp.execute(sql2, (i1,i2,i3,i4,i5,i6,i7,i8,i9,i10,i11,i12,i13,i14,i15,i16,i19,i20,i21,i23,i24,))
                #self.curr_mssql_erp.commit()

                #return 'ok'
            
            ##########
            # 領料單  
            ##########   
                # --領料單頭
                #INSERT INTO Leader.dbo.MOCTC (COMPANY, CREATOR, USR_GROUP, CREATE_DATE, MODIFIER, MODI_DATE, FLAG, TC001, TC002, TC003, TC004, TC005, TC008, TC009, TC010, TC014, TC015, TC016)
                #     VALUES ('{picking.COMPANY}', '{picking.CREATOR}', '{picking.USR_GROUP}', '{picking.CREATE_DATE}', '{picking.MODIFIER}', '{picking.MODI_DATE}', '{picking.FLAG}', '{picking.TC001}', '{picking.TC002}', '{picking.TC003}', '{picking.TC004}', '{picking.TC005}', '{picking.TC008}', '{picking.TC009}', '{picking.TC010}', '{picking.TC014}', '{picking.TC015}', '{picking.TC016}')
                
                # --領料單身
                #INSERT 
                #     INTO Leader.dbo.MOCTE (COMPANY, CREATOR, USR_GROUP, CREATE_DATE, MODIFIER, MODI_DATE, FLAG, TE001, TE002, TE003, TE004, TE005, TE006, TE008, TE009, TE010, TE011, TE012, TE014, TE016, TE017, TE018, TE019, TE021)
                #     VALUES ('{picking.COMPANY}', '{picking.CREATOR}', '{picking.USR_GROUP}', '{picking.CREATE_DATE}', '{picking.MODIFIER}', '{picking.MODI_DATE}', '{picking.FLAG}', '{picking.TE001}', '{picking.TE002}', '{picking.TE003}', '{picking.TE004}', '{picking.TE005}', '{picking.TE006}', '{picking.TE008}', '{picking.TE009}', '{picking.TE010}', '{picking.TE011}', '{picking.TE012}', '{picking.TE014}', '{picking.TE016}', '{picking.TE017}', '{picking.TE018}', '{picking.TE019}', '{picking.TE021}')
                
                # --領料製令單
                #INSERT INTO Leader.dbo.MOCTD (COMPANY, CREATOR, USR_GROUP, CREATE_DATE, MODIFIER, MODI_DATE, FLAG, TD001, TD002, TD003, TD004, TD005, TD006, TD007, TD008, TD014)
                #     VALUES ('{picking.COMPANY}', '{picking.CREATOR}', '{picking.USR_GROUP}', '{picking.CREATE_DATE}', '{picking.MODIFIER}', '{picking.MODI_DATE}', '{picking.FLAG}', '{picking.TD001}', '{picking.TD002}', '{picking.TD003}', '{picking.TD004}', '{picking.TD005}', '{picking.TD006}', '{picking.TD007}', '{picking.TD008}', '{picking.TD014}')
                
            ##########
            # 製令單
            ##########
                # --製令單頭
                # UPDATE Leader.dbo.MOCTA
                #     SET TA011 ='{picking.MOCTA011}', TA016 ='{picking.MOCTA016} '
                #     WHERE TA001 ='{picking.MOCTA001}'AND TA002 ='{picking.MOCTA002} '
            
                # --製令單身
                # UPDATE Leader.dbo.MOCTB
                #     SET TB005 ='{picking.MOCTB005} '
                #     WHERE TB001 ='{picking.MOCTB001}'AND TB002 ='{picking.MOCTB002}'AND TB003 ='{picking.MOCTB003} '
            
            ##########
            # 退料單
            ##########
                # --製令單頭
                # UPDATE Leader.dbo.MOCTA
                #     SET TA011 ='{picking.MOCTA011}', TA016 ='{picking.MOCTA016} '
                #     WHERE TA001 ='{picking.MOCTA001}'AND TA002 ='{picking.MOCTA002} '
            
                # --製令單身
                # UPDATE Leader.dbo.MOCTB
                #     SET TB005 ='{picking.MOCTB005} '
                #     WHERE TB001 ='{picking.MOCTB001}'AND TB002 ='{picking.MOCTB002}'AND TB003 ='{picking.MOCTB003} '

        except Exception as e:
            logging.info(f"\n<Error> wms_add_erp_picking_list_data_1 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #######################################
    # wms_query_erp_picking_list_query
    #######################################
    def wms_query_erp_picking_list_query(self , f_h , f_b):
        
        self.__connect_mssql_erp__()
        
        try:

            r_f_h = f_h + " "
            f_num = f"{r_f_h}-{f_b}"

            sql = '''
                SELECT TOP 10
                ISNULL(TA001, '') + '-' + ISNULL(TA002, '') AS ManufacturingOrderNo,
                TA063 AS BatchNo,
                TA009 AS ManufacturingDateFrom,
                TA010 AS ManufacturingDateTo,
                TA009 AS PackagingDateFrom,
                TA010 AS PackagingDateTo,
                TA006 AS FinishedProduct,
                TA015 AS BatchQuantity,
                TB003 AS ProductCode,
                TB012 AS ProductName,
                TB004 AS RequiredQuantity,
                '轉出庫' AS TransferOutWarehouse,
                '轉入庫' AS TransferInWarehouse
                FROM Leader.dbo.MOCTA
                LEFT JOIN Leader.dbo.MOCTB ON (MOCTA.TA001 = MOCTB.TB001) AND (MOCTA.TA002 = MOCTB.TB002)
                WHERE TA013 = 'Y' AND TA001 + '-' + TA002 = ?
                ORDER BY TA001 DESC, TA002 DESC
            '''

            self.curr_mssql_erp.execute(sql, (f_num,))
            res = self.curr_mssql_erp.fetchall()

            column_names = [column[0] for column in self.curr_mssql_erp.description]
            
            msg = []
            for row in res:
                row_dict = {col: (float(val) if isinstance(val, Decimal) else val) for col, val in zip(column_names, row)}
                msg.append(row_dict)
            
            return jsonify({"Otsuak 領料單查詢": msg})

        except Exception as e:
            logging.info(f"\n<Error> wms_query_erp_picking_list_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #######################################
    # factory_erp_query_import_3
    #######################################
    def factory_erp_query_import_3(self):
        
        self.__connect_mssql_erp__()
        
        try:

            sql  = f"SELECT DISTINCT "
            sql += f"B.TA006 AS '產品品號' "
            sql += f"FROM "
            sql += f"OtsukaDB.dbo.MOCTE A "
            sql += f"JOIN " 
            sql += f"OtsukaDB.dbo.MOCTA B ON A.TE011 = B.TA001 AND A.TE012 = B.TA002 "
            sql += f"JOIN " 
            sql += f"OtsukaDB.dbo.INVMB C ON B.TA006 = C.MB001 "
            sql += f"LEFT JOIN " 
            sql += f"OtsukaDB.dbo.INVME D ON A.TE004 = D.ME001 AND A.TE010 = D.ME002 "
            sql += f"LEFT JOIN " 
            sql += f"OtsukaDB.dbo.INVMB E ON D.ME001 = E.MB001 "
            sql += f"LEFT JOIN " 
            sql += f"OtsukaDB.dbo.PURMA F ON E.MB032 = F.MA001 "
            sql += f"WHERE " 
            sql += f"B.TA009 >= '20240101' "
            sql += f"AND D.ME003 >= '20200101' "
            sql += f"ORDER BY B.TA006"
            
            self.curr_mssql_erp.execute(sql)
            res = self.curr_mssql_erp.fetchall()
            
            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_query_import_3 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()


    #######################################
    # factory_erp_query_import_year_month
    #######################################
    def factory_erp_query_import_year_month(self):
        
        self.__connect_mssql_erp__()
        
        try:

            sql  = f"SELECT DISTINCT "
            sql += f"LEFT(F.TF003,4) + SUBSTRING(F.TF003,5,2) AS '入庫年月' "
            sql += f"FROM " 
            sql += f"OtsukaDB.dbo.MOCTF F "
            sql += f"JOIN "
            sql += f"OtsukaDB.dbo.MOCTG G ON F.TF001 = G.TG001 AND F.TF002 = G.TG002 "
            sql += f"JOIN " 
            sql += f"OtsukaDB.dbo.CMSMC C ON G.TG010 = C.MC001 "
            sql += f"WHERE " 
            sql += f"F.TF003 >= '2024-01-01' "
            sql += f"AND F.TF001 IN ('580', '583', '584', '585', '590') "
            sql += f"AND G.TG018 > CONVERT(VARCHAR, Getdate(), 112) "
            sql += f"AND LEFT(F.TF003,4) + SUBSTRING(F.TF003,5,2) BETWEEN FORMAT(DATEADD(MONTH, -3, GETDATE()), 'yyyyMM') AND FORMAT(DATEADD(MONTH, 0, GETDATE()), 'yyyyMM') "
            sql += f"ORDER BY '入庫年月'"
            
            self.curr_mssql_erp.execute(sql)
            res = self.curr_mssql_erp.fetchall()
            
            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_query_import_year_month : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    ##################################
    # login_out_record
    ##################################
    def login_out_record(self):
        
        self.__connect__()
        
        try:
            sql = f"SELECT a_user , count(*) FROM `login_out_record` group by a_user"

            self.curr.execute(sql)
            res = self.curr.fetchall()
            
            return res

        except Exception as e:
            logging.info(f"\n<Error> login_out_record : {str(e)}\n")

        finally:
            self.__disconnect__()

    ##################################
    # factory_erp_query_product_num
    ##################################
    def factory_erp_query_product_num(self , item):
        
        self.__connect_mssql_erp__()
        
        try:
            
            if item == '產品品號':
                
                f_erp_sql  = f"SELECT DISTINCT "
                f_erp_sql += f"A.MF001 AS '產品品號', C.MB002 AS '產品名稱' "
                f_erp_sql += f"FROM OtsukaDB.dbo.INVMF A "
                f_erp_sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
                f_erp_sql += f"WHERE A.MF003 >= '20190101' "
                f_erp_sql += f"AND A.MF001 LIKE 'X%' "
                f_erp_sql += f"ORDER BY A.MF001 "

                self.curr_mssql_erp.execute(f_erp_sql)
                res = self.curr_mssql_erp.fetchall()

            elif item == '產品名稱':
                pass
            elif item == '批號':
                pass
            elif item == '有效日期':
                pass
            
            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_query_product_num : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()
    
    #############################################
    # factory_erp_ss2_product_name_query_total2
    #############################################
    def factory_erp_ss2_product_name_total2(self , q_s_date , q_e_date , q_c_name , q_p_name):
        
        self.__connect_mssql_ss2__()
        
        try:
            
            if q_c_name != "" and q_p_name != "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"c.FULLNAME , "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}' AND c.FULLNAME='{q_c_name}' group by c.FULLNAME"

            elif q_c_name == "" and q_p_name != "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"c.FULLNAME , "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}' group by c.FULLNAME"
            
            elif q_c_name != "" and q_p_name == "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"c.FULLNAME , "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}' group by c.FULLNAME"
            
            elif q_c_name == "" and q_p_name == "":
            
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"c.FULLNAME , "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' group by c.FULLNAME"
            
                
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_ss2_product_name_total : {str(e)}\n")

        finally:
            self.__disconnect_mssql_ss2__()
    
    #############################################
    # factory_erp_ss2_product_name_query_total
    #############################################
    def factory_erp_ss2_product_name_total(self , q_s_date , q_e_date , q_c_name , q_p_name):
        
        self.__connect_mssql_ss2__()
        
        try:
            
            if q_c_name != "" and q_p_name != "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}' AND c.FULLNAME='{q_c_name}'"

            elif q_c_name == "" and q_p_name != "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}'"
            
            elif q_c_name != "" and q_p_name == "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}'"
            
            elif q_c_name == "" and q_p_name == "":
            
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"FORMAT(count(*) , 'N0') "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}'"
            
                
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchone()

            return res[0]

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_ss2_product_name_total : {str(e)}\n")

        finally:
            self.__disconnect_mssql_ss2__()


    #######################################
    # factory_erp_bom_query2
    #######################################
    def factory_erp_bom_query2(self , q_c_name):
        
        self.__connect_mssql_erp__()

        print(f"{q_c_name}")
        
        try:
            
            if q_c_name != "":
                
                f_erp_sql1  =  '''
                                SELECT 
                                MC.MC001 AS '主件品號', 
                                (SELECT MB002 FROM INVMB WHERE MB001 = MC.MC001) AS '主件品名',  
                                MC.MC002 AS '主件單位', 
                                CAST(MC.MC004 AS INT) AS '標準批量', 
                                MC.MC005 AS '製令單別',  
                                MD.MD003 AS '元件品號', 
                                (SELECT MB002 FROM INVMB WHERE MB001 = MD.MD003) AS '元件品名',  
                                MD.MD004 AS '元件單位', 
                                MD.MD006 AS '組成用量', 
                                MD.MD017 AS '材料型態', 
                                ME.ME002 AS '元件批號', 
                                ME.ME003 AS '最早入庫日',
                                ME.ME009 AS '有效日期', 
                                MF_DATA.MF007 AS '庫別', 
                                MF_DATA.MF_SUM AS '庫存數量'  

                                FROM BOMMC MC  
                                JOIN BOMMD MD ON MC.MC001 = MD.MD001  
                                LEFT JOIN INVME ME ON MD.MD003 = ME.ME001  
                                OUTER APPLY ( SELECT MF007, SUM(MF008 * MF010) AS MF_SUM  
                                    FROM INVMF MF  
                                    WHERE MF.MF001 = ME.ME001 AND MF.MF002 = ME.ME002  
                                    GROUP BY MF007  
                                    HAVING SUM(MF008 * MF010) > 0  
                                ) AS MF_DATA  

                                WHERE MC.MC016 = 'Y' and MC.MC001 = ? AND MF_DATA.MF_SUM IS NOT NULL;
                               '''
                
                self.curr_mssql_erp.execute(f_erp_sql1 , ( q_c_name ,))
                res = self.curr_mssql_erp.fetchall()

                return res
            
            else:
                f_erp_sql1  =  '''
                                SELECT 
                                MC.MC001 AS '主件品號', 
                                (SELECT MB002 FROM INVMB WHERE MB001 = MC.MC001) AS '主件品名',  
                                MC.MC002 AS '主件單位', 
                                CAST(MC.MC004 AS INT) AS '標準批量', 
                                MC.MC005 AS '製令單別',  
                                MD.MD003 AS '元件品號', 
                                (SELECT MB002 FROM INVMB WHERE MB001 = MD.MD003) AS '元件品名',  
                                MD.MD004 AS '元件單位', 
                                MD.MD006 AS '組成用量', 
                                MD.MD017 AS '材料型態', 
                                ME.ME002 AS '元件批號', 
                                ME.ME003 AS '最早入庫日',
                                ME.ME009 AS '有效日期', 
                                MF_DATA.MF007 AS '庫別', 
                                MF_DATA.MF_SUM AS '庫存數量'  

                                FROM BOMMC MC  
                                JOIN BOMMD MD ON MC.MC001 = MD.MD001  
                                LEFT JOIN INVME ME ON MD.MD003 = ME.ME001  
                                OUTER APPLY ( SELECT MF007, SUM(MF008 * MF010) AS MF_SUM  
                                    FROM INVMF MF  
                                    WHERE MF.MF001 = ME.ME001 AND MF.MF002 = ME.ME002  
                                    GROUP BY MF007  
                                    HAVING SUM(MF008 * MF010) > 0  
                                ) AS MF_DATA  

                                WHERE MC.MC016 = 'Y' AND MF_DATA.MF_SUM IS NOT NULL;
                                '''
                
                self.curr_mssql_erp.execute(f_erp_sql1)
                res = self.curr_mssql_erp.fetchall()

                return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_bom_query1 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #######################################
    # factory_erp_subform_query1
    #######################################
    def factory_erp_subform_query1(self , q_s_date , q_e_date):
        
        self.__connect_mssql_erp__()

        try:

            s_date = datetime.strptime(q_s_date, "%Y-%m-%d")
            e_date = datetime.strptime(q_e_date, "%Y-%m-%d")
            s_d    = s_date.strftime("%Y%m%d")
            e_d    = e_date.strftime("%Y%m%d")
                
            f_erp_sql1  =  '''
                            SELECT 
                                ACTTA.TA003 AS '傳票日期',
                                INVMB.MB008 AS '成品品類',
                                ACTXA.XA005 AS '品號',
                                INVMB.MB002 AS '品名', 
                                ACTXA.XA002 AS '傳票單號',
                                ACTTB.TB005 AS '科目編號',
                                ACTMA.MA003 AS '科目名稱',
                                ACTTB.TB006 AS '部門代號', 
                                (ACTTB.TB004 * ACTXA.XA010) AS Sub_Cost

                            FROM 
                                OtsukaDB.dbo.ACTTA ACTTA
                                INNER JOIN OtsukaDB.dbo.ACTTB ACTTB ON ACTTA.TA001 = ACTTB.TB001 AND ACTTA.TA002 = ACTTB.TB002
                                INNER JOIN OtsukaDB.dbo.ACTXA ACTXA ON ACTTB.TB001 = ACTXA.XA001 AND ACTTB.TB002 = ACTXA.XA002 AND ACTTB.TB003 = ACTXA.XA003
                                INNER JOIN OtsukaDB.dbo.ACTMA ACTMA ON ACTTB.TB005 = ACTMA.MA001
                                LEFT OUTER JOIN OtsukaDB.dbo.INVMB INVMB ON ACTXA.XA005 = INVMB.MB001

                            WHERE
                                ACTTA.TA003 BETWEEN ? AND ?
                                AND ACTTB.TB005 LIKE '62%' 

                            ORDER BY ACTTA.TA003;
                            '''
            
            self.curr_mssql_erp.execute(f_erp_sql1 , (s_d,e_d,))
            res = self.curr_mssql_erp.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_subform_query1 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #######################################
    # factory_erp_bom_query1
    #######################################
    def factory_erp_bom_query1(self , q_c_name):
        
        self.__connect_mssql_erp__()

        print(f"{q_c_name}")
        
        try:
            
            if q_c_name != "":
                
                f_erp_sql1  =  '''
                                SELECT
                                MC.MC001 AS '主件品號',
                                (SELECT MB002 FROM INVMB WHERE MB001 = MC.MC001) AS  '主件品名',  
                                CAST(MC.MC004 AS INT) AS '標準批量',
                                MC.MC002 AS '主件單位', 
                                MD.MD002 AS '序號', 
                                MD.MD003 AS '元件品號',  
                                (SELECT MB002 FROM INVMB WHERE MB001 = MD.MD003)  AS '元件品名',  
                                MD.MD004 AS '元件單位', 
                                MD.MD017 AS '材料型態', 
                                MD.MD006 AS '組成用量'
                                FROM BOMMC MC JOIN BOMMD MD ON MC.MC001 = MD.MD001 WHERE MC.MC016 = 'Y' and MC.MC001 = ?;
                               '''
                
                self.curr_mssql_erp.execute(f_erp_sql1 , ( q_c_name ,))
                res = self.curr_mssql_erp.fetchall()

                return res
            
            else:
                f_erp_sql1  = '''
                                SELECT
                                MC.MC001 AS '主件品號',
                                (SELECT MB002 FROM INVMB WHERE MB001 = MC.MC001) AS  '主件品名',  
                                CAST(MC.MC004 AS INT) AS '標準批量',
                                MC.MC002 AS '主件單位', 
                                MD.MD002 AS '序號', 
                                MD.MD003 AS '元件品號',  
                                (SELECT MB002 FROM INVMB WHERE MB001 = MD.MD003)  AS '元件品名',  
                                MD.MD004 AS '元件單位', 
                                MD.MD017 AS '材料型態', 
                                MD.MD006 AS '組成用量'
                                FROM BOMMC MC JOIN BOMMD MD ON MC.MC001 = MD.MD001 WHERE MC.MC016 = 'Y';
                                '''
                
                self.curr_mssql_erp.execute(f_erp_sql1)
                res = self.curr_mssql_erp.fetchall()

                return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_bom_query1 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    #######################################
    # factory_erp_ss2_product_name_query2
    #######################################
    def factory_erp_ss2_product_name_query2(self , q_s_date , q_e_date , q_c_name , q_p_name):
        
        self.__connect_mssql_ss2__()
        
        try:
            
            if q_c_name != "" and q_p_name != "":
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.YY AS 年, "
                f_erp_sql += f"v.MM AS 月, "
                f_erp_sql += f"v.YYMM AS 年月, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "
                f_erp_sql += f"v.DISCOUNT2, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.SAMPLE, "
                f_erp_sql += f"v.TAX, "
                f_erp_sql += f"v.TEMP, "
                f_erp_sql += f"v.TEMP2 "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}' AND c.FULLNAME='{q_c_name}' order by v.TRADE_DATE desc"

            elif q_c_name == "" and q_p_name != "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.YY AS 年, "
                f_erp_sql += f"v.MM AS 月, "
                f_erp_sql += f"v.YYMM AS 年月, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "
                f_erp_sql += f"v.DISCOUNT2, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.SAMPLE, "
                f_erp_sql += f"v.TAX, "
                f_erp_sql += f"v.TEMP, "
                f_erp_sql += f"v.TEMP2 "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND p.PROD_NAME='{q_p_name}' order by v.TRADE_DATE desc"
            
            elif q_c_name != "" and q_p_name == "":
                
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.YY AS 年, "
                f_erp_sql += f"v.MM AS 月, "
                f_erp_sql += f"v.YYMM AS 年月, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "
                f_erp_sql += f"v.DISCOUNT2, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.SAMPLE, "
                f_erp_sql += f"v.TAX, "
                f_erp_sql += f"v.TEMP, "
                f_erp_sql += f"v.TEMP2 "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' AND c.FULLNAME='{q_c_name}' order by v.TRADE_DATE desc"
            
            elif q_c_name == "" and q_p_name == "":
            
                f_erp_sql  = f"SELECT "
                f_erp_sql += f"v.TRADE_DATE AS 交易日期, "
                f_erp_sql += f"v.CUST_ID AS 客戶代號, "
                f_erp_sql += f"c.FULLNAME AS 客戶名稱, "
                f_erp_sql += f"c.ADDRESS AS 地址, "
                f_erp_sql += f"v.OLD_ID, "
                f_erp_sql += f"v.PROD_ID AS 產品代號, "
                f_erp_sql += f"p.PROD_NAME AS 產品名稱, "
                f_erp_sql += f"p.ERP_CODE AS ERP產品代碼, "
                f_erp_sql += f"v.LOT_NO AS 批號, "
                f_erp_sql += f"v.NET_QTY AS 淨數量, "
                f_erp_sql += f"v.NET_AMT AS 淨金額, "
                f_erp_sql += f"v.RESELLER_ID AS 經銷商, "
                f_erp_sql += f"v.YY AS 年, "
                f_erp_sql += f"v.MM AS 月, "
                f_erp_sql += f"v.YYMM AS 年月, "
                f_erp_sql += f"v.WH_ID AS 庫別, "
                f_erp_sql += f"v.AMOUNT, "
                f_erp_sql += f"v.DISCOUNT, "
                f_erp_sql += f"v.DISCOUNT2, "  
                f_erp_sql += f"v.QTY, "
                f_erp_sql += f"v.SAMPLE, "
                f_erp_sql += f"v.TAX, "
                f_erp_sql += f"v.TEMP, "
                f_erp_sql += f"v.TEMP2 "
                f_erp_sql += f"FROM "
                f_erp_sql += f"SALES2.dbo.vTRADEDET v "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.CUSTOMER c ON v.CUST_ID = c.CUST_ID "
                f_erp_sql += f"JOIN "
                f_erp_sql += f"SALES2.dbo.PRODUCT p ON v.PROD_ID = p.PROD_ID "
                f_erp_sql += f"WHERE "
                f_erp_sql += f"v.TRADE_DATE >= '{q_s_date}' AND v.TRADE_DATE <= '{q_e_date}' order by v.TRADE_DATE desc"
            
                
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            f_final_res = []

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_ss2_product_name_query2 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_ss2__()

    #######################################
    # factory_erp_ss2_customer_name_query
    #######################################
    def factory_erp_ss2_customer_name_query(self):
        
        self.__connect_mssql_ss2__()
        
        try:
            f_erp_sql  = '''SELECT
                            c.FULLNAME AS '客戶名稱'
                            FROM
                                SALES2.dbo.vTRADEDET v
                            JOIN
                                SALES2.dbo.CUSTOMER c ON v."CUST_ID" = c."CUST_ID"
                            JOIN
                                SALES2.dbo.PRODUCT p ON v."PROD_ID" = p."PROD_ID"
                            WHERE
                                v."TRADE_DATE" >= '2020-01-01' AND v."TRADE_DATE" <= '2100-01-01' group by c."FULLNAME";'''
            
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            f_final_res = []

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_ss2_customer_name_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_ss2__()

    #######################################
    # factory_erp_product_num_query
    #######################################
    def factory_erp_product_num_query(self):
        
        self.__connect_mssql_erp__()
        
        try:
            f_erp_sql  = '''
                            SELECT 
                            MC.MC001 AS '產品品號' ,
                            MB.MB002 AS '產品名稱'
                            FROM OtsukaDB.dbo.BOMMC MC
                            JOIN OtsukaDB.dbo.INVMB MB ON MC.MC001 = MB.MB001
                         '''
            
            self.curr_mssql_erp.execute(f_erp_sql)
            res = self.curr_mssql_erp.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_product_num_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()


    #######################################
    # factory_erp_ss2_product_name_query
    #######################################
    def factory_erp_ss2_product_name_query(self):
        
        self.__connect_mssql_ss2__()
        
        try:
            f_erp_sql  = '''SELECT
                            p."PROD_NAME" AS '產品名稱'
                            FROM
                                SALES2.dbo.vTRADEDET v
                            JOIN
                                SALES2.dbo.CUSTOMER c ON v."CUST_ID" = c."CUST_ID"
                            JOIN
                                SALES2.dbo.PRODUCT p ON v."PROD_ID" = p."PROD_ID"
                            WHERE
                                v."TRADE_DATE" >= '2020-01-01' AND v."TRADE_DATE" <= '2100-01-01' group by p."PROD_NAME";'''
            
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            f_final_res = []

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_ss2_product_name_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_ss2__()

    ####################################
    # purchase_sales_erp_query_by_date
    ####################################
    def purchase_sales_erp_query_by_date(self):
        
        self.__connect_mssql_erp__()
        
        try:
            
            ### ERP 進銷項彙總表
            f_erp_sql1  = '''
                            SELECT MC002 , count(*) FROM OtsukaDB.dbo.TAXMC group by MC002 order by MC002 DESC;
            '''
            
            self.curr_mssql_erp.execute(f_erp_sql1)
            res = self.curr_mssql_erp.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> purchase_sales_erp_query_by_date : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    ####################################
    # purchase_sales_erp_person_query
    ####################################
    def purchase_sales_erp_person_query(self):
        
        self.__connect_mssql_erp__()
        
        try:
            
            ### ERP 進銷項彙總表
            f_erp_sql1  = '''
                            SELECT CREATOR AS '建立者' FROM OtsukaDB.dbo.TAXMC group by CREATOR
            '''
            
            self.curr_mssql_erp.execute(f_erp_sql1)
            res = self.curr_mssql_erp.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> purchase_sales_erp_person_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    ###########################
    # purchase_sales_erp_query2
    ###########################
    def purchase_sales_erp_query2(self):
        
        self.__connect_mssql_erp__()
        
        try:
            
            ### ERP 進銷項彙總表
            f_erp_sql1  = '''
                            SELECT 
                            CREATOR AS '建立者',
                            MC001,  
                            MC002 AS '申報年月', 
                            LEFT(MC002, 4) AS '申報年',
                            MC003, 
                            MC004 AS '格式代號',
                            CASE 
                                WHEN MC004 IN ('21', '22', '25', '28', '23') THEN '進項'
                                WHEN MC004 IN ('31', '32', '35', '33', '36') THEN '銷項'
                            END AS '進銷項區分',
                            CASE
                                WHEN MC004 IN ('21', '22', '25', '28') THEN '進項(加)'
                                WHEN MC004 = '23' THEN '進項(減)'
                                WHEN MC004 IN ('31', '32', '35') THEN '銷項(加)'
                                WHEN MC004 = '33' THEN '銷項(減)'
                                WHEN MC004 = '36' THEN '零稅率' 	
                            END AS '進銷項',
                            CASE
                                WHEN MC004 IN ('21', '22', '25', '28') THEN 1
                                WHEN MC004 = '23' THEN 2
                                WHEN MC004 IN ('31', '32', '35') THEN 3
                                WHEN MC004 = '33' THEN 4
                                WHEN MC004 = '36' THEN 5
                            END AS '進銷項排序',
                            CASE 
                                WHEN MC004 IN ('21', '22', '25', '28') THEN 1
                                WHEN MC004 = '23' THEN -1
                                WHEN MC004 IN ('31', '32', '35') THEN 1
                                WHEN MC004 = '33' THEN -1
                                ELSE 1
                            END AS '進銷項加減',
                            MC005 AS '稅籍編號',
                            MC006 AS '流水號',
                            MC007 AS '開立日期',
                            MC008 AS '買方統一編號',
                            MC009 AS '賣方統一編號',
                            MC010 AS '發票號碼',
                            CASE
                                WHEN MC004 = '32' THEN CAST(MC011 / 1.05 AS INT)  
                                ELSE CAST(MC011 AS INT)  
                            END AS '發票金額',
                            CASE
                                WHEN MC012 = '0' THEN '應稅內含'
                                WHEN MC012 = '1' THEN '應稅外加'
                                WHEN MC012 = '2' THEN '零稅率'
                                WHEN MC012 = '3' THEN '免稅'
                                WHEN MC012 = 'D' THEN '作廢'
                            END AS '課稅別',
                            MC013 AS '營業稅額',
                            CASE
                                WHEN MC014 = '1' THEN '可扣抵之進貨及費用'
                                WHEN MC014 = '2' THEN '可扣抵之固定資產'
                                WHEN MC014 = '3' THEN '不可扣抵之進貨及費用'
                                WHEN MC014 = '4' THEN '不可扣抵之固定資產'
                            END AS '扣抵別',
                            MC015, 
                            MC016, 
                            MC017, 
                            MC018 AS '備註',
                            MC019 AS '來源方式'
                        FROM OtsukaDB.dbo.TAXMC 
                        WHERE MC002 >= '202410' AND MC002 <= '202410' 
                        ORDER BY MC007 DESC;
            '''
            
            self.curr_mssql_erp.execute(f_erp_sql1)
            res = self.curr_mssql_erp.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> purchase_sales_erp_query2 : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()


    ###########################
    # purchase_sales_erp_query
    ###########################
    def purchase_sales_erp_query(self):
        
        self.__connect_mssql_erp__()
        
        try:
            
            ### ERP 進銷項明細表
            f_erp_sql1  = '''
                            SELECT TOP 10
                            CREATOR AS '建立者',
                            MC001,
                            MC002 AS '申報年月',
                            LEFT(MC002,4)	AS '申報年',
                            MC003,
                            MC004 AS '格式代號',
                            CASE
                                WHEN MC004 = '21' OR MC004 = '22' OR MC004 = '25' OR MC004 = '28' THEN '進項(加)'
                                WHEN MC004 = '23' THEN '進項(減)'
                                WHEN MC004 = '31' OR MC004 = '32' OR MC004 = '35' THEN '銷項(加)'
                                WHEN MC004 = '33' THEN '銷項(減)'
                                WHEN MC004 = '36' THEN '零稅率' 
                            END	AS	'進銷項',
                            MC005 AS '稅籍編號',
                            MC006 AS '流水號',
                            MC007 AS '開立日期',
                            MC008 AS '買方統一編號',
                            MC009 AS '賣方統一編號',
                            MC010 AS '發票號碼',
                            CASE
                                WHEN MC004='32' THEN MC011/1.05
                                ELSE MC011
                            END  AS	'銷售金額',
                            CASE
                                WHEN MC012 = '0' THEN '應稅內含'
                                WHEN MC012 = '1' THEN '應稅外加'
                                WHEN MC012 = '2' THEN '零稅率'
                                WHEN MC012 = '3' THEN '免稅'
                                WHEN MC012 = 'D' THEN '作廢' 
                            END AS '課稅別',
                            MC013 AS '營業稅額',
                            CASE
                            WHEN MC014 = 1 THEN '可扣抵之進貨及費用'
                            WHEN MC014 = 2 THEN '可扣抵之固定資產'
                            WHEN MC014 = 3 THEN '不可扣抵之進貨及費用'
                            WHEN MC014 = 4 THEN '不可扣抵之固定資產' 
                            END AS '扣抵別',
                            MC015,
                            MC016,
                            MC017,
                            MC018 AS '備註',
                            MC019 AS '來源方式',
                            MC020,
                            MC021,
                            MC022,
                            MC023,
                            MC024,
                            MC025,
                            MC026,
                            MC027,
                            MC028,
                            MC029,
                            MC030,
                            MC031

                        FROM OtsukaDB.dbo.TAXMC WHERE MC002>= '2024010' and MC002 <= '202410';
            '''
            
            self.curr_mssql_erp.execute(f_erp_sql1)
            res = self.curr_mssql_erp.fetchall()

            return res

        except Exception as e:
            logging.info(f"\n<Error> purchase_sales_erp_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    ###########################
    # factory_erp_ss2_query
    ###########################
    def factory_erp_ss2_query(self):
        
        self.__connect_mssql_ss2__()
        
        try:
            f_erp_sql  = '''SELECT TOP 50
                            v."TRADE_DATE" AS '交易日期',
                            v."CUST_ID" AS '客戶代號',
                            c."FULLNAME" AS '客戶名稱',
                            c."ADDRESS" AS '地址',
                            v."OLD_ID",
                            v."PROD_ID" AS '產品代號',
                            p."PROD_NAME" AS '產品名稱',
                            p.ERP_CODE AS 'ERP產品代碼',
                            v."LOT_NO" AS '批號',
                            v."NET_QTY" AS '淨數量',
                            v."NET_AMT" AS '淨金額',
                            v."RESELLER_ID" AS '經銷商',
                            v.YY AS '年',
                            v.MM AS '月',
                            v.YYMM AS '年月',
                            v."WH_ID" AS '庫別',
                            v.AMOUNT,
                            v.DISCOUNT,
                            v.DISCOUNT2,   
                            v.QTY,
                            v.SAMPLE,
                            v.TAX,
                            v."TEMP",
                            v.TEMP2
                        FROM
                            SALES2.dbo.vTRADEDET v
                        JOIN
                            SALES2.dbo.CUSTOMER c ON v."CUST_ID" = c."CUST_ID"
                        JOIN
                            SALES2.dbo.PRODUCT p ON v."PROD_ID" = p."PROD_ID"
                        WHERE
                            v."TRADE_DATE" >= '2020-01-01' AND v."TRADE_DATE" <= '2100-01-01'  order by v.TRADE_DATE desc'''
            
            self.curr_mssql_ss2.execute(f_erp_sql)
            res = self.curr_mssql_ss2.fetchall()

            f_final_res = []

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_ss2__()

    ###########################
    # factory_erp_query
    ###########################
    def factory_erp_query(self):
        
        self.__connect_mssql_erp__()
        
        try:
            f_erp_sql  = f"SELECT "
            f_erp_sql += f"A.MF001 AS '產品品號', "
            f_erp_sql += f"C.MB002 AS '產品名稱', "
            f_erp_sql += f"A.MF002 AS '批號', "
            f_erp_sql += f"E.ME009 AS '有效日期' "
            f_erp_sql += f"FROM OtsukaDB.dbo.INVMF A "
            f_erp_sql += f"JOIN OtsukaDB.dbo.CMSMQ B ON A.MF004 = B.MQ001 "
            f_erp_sql += f"JOIN OtsukaDB.dbo.INVMB C ON A.MF001 = C.MB001 "
            f_erp_sql += f"JOIN OtsukaDB.dbo.CMSMC D ON A.MF007 = D.MC001 "
            f_erp_sql += f"JOIN OtsukaDB.dbo.INVME E ON A.MF001 = E.ME001 AND A.MF002 = E.ME002 "
            f_erp_sql += f"WHERE A.MF003 >= '20190101' "
            f_erp_sql += f"AND A.MF001 LIKE 'X%' "
            f_erp_sql += f"AND E.ME009 >= '20240101' "
            f_erp_sql += f"AND E.ME003 >= '20200101' "
            f_erp_sql += f"GROUP BY A.MF001, C.MB002 , A.MF002 , E.ME009 "
            f_erp_sql += f"ORDER BY A.MF001"
            
            self.curr_mssql_erp.execute(f_erp_sql)
            res = self.curr_mssql_erp.fetchall()

            f_final_res = []

            return res

        except Exception as e:
            logging.info(f"\n<Error> factory_erp_query : {str(e)}\n")

        finally:
            self.__disconnect_mssql_erp__()

    ###########################
    # submit_otsuka_contract
    ###########################
    def submit_otsuka_contract(self , o_c_date , o_c_kind , o_c_title , o_c_cost , o_c_time , o_c_company , o_c_name , o_c_telephone , o_c_phone , o_c_comment):
        
        self.__connect__()
        
        try:
            b_date    = o_c_date
            kind      = o_c_kind
            title     = o_c_title
            cost      = o_c_cost
            t_s_e     = o_c_time
            company   = o_c_company
            name      = o_c_name
            telephone = o_c_telephone
            phone     = o_c_phone
            comment   = o_c_comment
            
            self.sql = f"insert into otsuka_contract(b_date , kind , title , cost , t_s_e , company , contract_name , contract_telephone , contract_phone , comment) value('{b_date}','{kind}','{title}','{cost}','{t_s_e}','{company}' , '{name}'  , '{telephone}' , '{phone}' , '{comment}')"
            self.curr.execute(self.sql)

        except Exception as e:
            logging.info("\n<Error> submit_otsuka_contract : " + str(e))

        finally:
            self.__disconnect__()
    
    ####################
    # add_vmware_data
    ####################
    def add_vmware_data(self , date , position , name , power , os , os_state , state , boot_time , ip , cpu , ram , hdd1 , hdd2 , hdd3 , hdd4 , hdd5 , hdd6 , hdd7):
        
        self.__connect__()
        
        try:
            self.sql  = f"insert into otsuka_vmware " 
            self.sql += f"(c_date , vm_position , vm_name , vm_power , vm_os , vm_os_state , vm_state , vm_boot_time , vm_ip , vm_cpu , vm_ram , vm_hdd1 , vm_hdd2 , vm_hdd3 , vm_hdd4 , vm_hdd5 , vm_hdd6 , vm_hdd7) "
            self.sql += f"value('{date}' , '{position}' , '{name}' , '{power}' , '{os}' , '{os_state}' , '{state}' , '{boot_time}' , '{ip}' , '{cpu}' , '{ram}' , '{hdd1}' , '{hdd2}' , '{hdd3}' , '{hdd4}' , '{hdd5}' , '{hdd6}' , '{hdd7}')"
            self.curr.execute(self.sql)

        except Exception as e:
            logging.info("\n<Error> add_vmware_data : " + str(e))

        finally:
            self.__disconnect__()


    #####################
    # operation_record
    #####################
    def operation_record(self,r_time,user,login_code,item):
        
        self.__connect__()
        
        try:
            self.r_time     = r_time
            self.user       = user
            self.item       = item
            self.login_code = login_code
            
            self.sql = "insert into operation_record(r_time,a_user,item,login_code) value('{0}','{1}','{2}','{3}')".format(self.r_time , self.user , self.item , self.login_code)
            self.curr.execute(self.sql)

        except Exception as e:
            logging.info("\n<Error> operation record : " + str(e))

        finally:
            self.__disconnect__()
    
    ##################
    # logout_record
    ##################
    def logout_record(self,user,login_code,r_time):
        
        try:
            self.user = user
            self.login_code = login_code
            self.r_time = r_time

            self.__connect__()    

            self.sql = "update login_out_record set logout_time='{0}' where login_code='{1}' and a_user='{2}'".format(self.r_time , self.login_code , self.user)
            self.curr.execute(self.sql)

        except Exception as e:
            logging.info("\n<Error> logout record : " + str(e))

        finally:
            self.__disconnect__()

    ###################################
    # __connect_mssql_wms_account__ 
    ###################################
    def __connect_mssql_wms_account__(self):
        
        try:
            ############
            # FreeTDS 
            ############
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};port=1433;DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #self.conn_mssql = pyodbc.connect(conn_str)
            #self.curr_mssql = self.conn_mssql.cursor()
            
            ##################################
            # ODBC Driver 18 for SQL Server 
            ##################################
            conn_str_wms_account = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={parameter.otsuka_factory9_wms['host']};DATABASE={parameter.otsuka_factory9_wms['db']};UID={parameter.otsuka_factory9_wms['user']};PWD={parameter.otsuka_factory9_wms['password']};TrustServerCertificate=yes;"  
            self.conn_mssql_wms_account = pyodbc.connect(conn_str_wms_account)
            self.curr_mssql_wms_account = self.conn_mssql_wms_account.cursor()
            
            
        except Exception as e:
            logging.info("\n<Error> __connect_mssql_wms_account__ " + str(e))

        finally:
            pass
            #self.curr_mssql.close()
            #self.conn_mssql.close()
    
    ####################################
    # __disconnect_mssql_wms_account__ 
    ####################################
    def __disconnect_mssql_wms_account__(self):
        
        try:
            self.curr_mssql_wms_account.close()
            self.conn_mssql_wms_account.close()
            
        except Exception as e:
            logging.info("\n<Error> __disconnect_mssql_wms_account__ " + str(e))

        finally:
            pass

    ##########################
    # __connect_mssql_wms__ 
    ##########################
    def __connect_mssql_wms__(self):
        
        try:
            ############
            # FreeTDS 
            ############
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};port=1433;DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #self.conn_mssql = pyodbc.connect(conn_str)
            #self.curr_mssql = self.conn_mssql.cursor()
            
            ##################################
            # ODBC Driver 18 for SQL Server 
            ##################################
            conn_str_wms = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={parameter.otsuka_wms_1_61['host']};DATABASE={parameter.otsuka_wms_1_61['db']};UID={parameter.otsuka_wms_1_61['user']};PWD={parameter.otsuka_wms_1_61['password']};TrustServerCertificate=yes;"  
            self.conn_mssql_wms = pyodbc.connect(conn_str_wms)
            self.curr_mssql_wms = self.conn_mssql_wms.cursor()
            
        except Exception as e:
            logging.info("\n<Error> __connect_mssql_wms__ " + str(e))

        finally:
            pass
            #self.curr_mssql.close()
            #self.conn_mssql.close()
    
    #############################
    # __disconnect_mssql_wms__ 
    #############################
    def __disconnect_mssql_wms__(self):
        
        try:
            self.curr_mssql_wms.close()
            self.conn_mssql_wms.close()
            
        except Exception as e:
            logging.info("\n<Error> __disconnect_mssql_wms__ " + str(e))

        finally:
            pass


    ##########################
    # __connect_mssql_erp__ 
    ##########################
    def __connect_mssql_erp__(self):
        
        try:
            ############
            # FreeTDS 
            ############
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};port=1433;DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #self.conn_mssql = pyodbc.connect(conn_str)
            #self.curr_mssql = self.conn_mssql.cursor()
            
            ##################################
            # ODBC Driver 18 for SQL Server 
            ##################################
            conn_str_erp = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={parameter.otsuka_factory5['host']};DATABASE={parameter.otsuka_factory5['db']};UID={parameter.otsuka_factory5['user']};PWD={parameter.otsuka_factory5['pwd']};TrustServerCertificate=yes;"  
            self.conn_mssql_erp = pyodbc.connect(conn_str_erp)
            self.curr_mssql_erp = self.conn_mssql_erp.cursor()
            
        except Exception as e:
            logging.info("\n<Error> __connect_mssql_erp__ " + str(e))

        finally:
            pass
            #self.curr_mssql.close()
            #self.conn_mssql.close()
    
    #############################
    # __disconnect_mssql_erp__ 
    #############################
    def __disconnect_mssql_erp__(self):
        
        try:
            self.curr_mssql_erp.close()
            self.conn_mssql_erp.close()
            
        except Exception as e:
            logging.info("\n<Error> __disconnect_mssql_erp__ " + str(e))

        finally:
            pass

    ##########################
    # __connect_mssql_ss2__ 
    ##########################
    def __connect_mssql_ss2__(self):
        
        try:
            ############
            # FreeTDS 
            ############
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};port=1433;DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #self.conn_mssql = pyodbc.connect(conn_str)
            #self.curr_mssql = self.conn_mssql.cursor()
            
            ##################################
            # ODBC Driver 18 for SQL Server 
            ##################################
            conn_str = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={parameter.otsuka_factory6['host']};DATABASE={parameter.otsuka_factory6['db']};UID={parameter.otsuka_factory6['user']};PWD={parameter.otsuka_factory6['pwd']};TrustServerCertificate=yes;"  
            self.conn_mssql_ss2 = pyodbc.connect(conn_str)
            self.curr_mssql_ss2 = self.conn_mssql_ss2.cursor()
            
        except Exception as e:
            logging.info("\n<Error> __connect_mssql_ss2__ " + str(e))

        finally:
            pass
            #self.curr_mssql.close()
            #self.conn_mssql.close()
    
    #############################
    # __disconnect_mssql_ss2__ 
    #############################
    def __disconnect_mssql_ss2__(self):
        
        try:
            self.conn_mssql_ss2.close()
            
        except Exception as e:
            logging.info("\n<Error> __disconnect_mssql_ss2__ " + str(e))

        finally:
            pass

    ######################
    # __connect_mssql__ 
    ######################
    def __connect_mssql__(self):
        
        try:
            ############
            # FreeTDS 
            ############
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};port=1433;DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #conn_str = f"DRIVER={{FreeTDS}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            #self.conn_mssql = pyodbc.connect(conn_str)
            #self.curr_mssql = self.conn_mssql.cursor()
            
            ##################################
            # ODBC Driver 18 for SQL Server 
            ##################################
            conn_str = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={parameter.otsuka_factory2['host']};DATABASE={parameter.otsuka_factory2['db']};UID={parameter.otsuka_factory2['user']};PWD={parameter.otsuka_factory2['pwd']};TrustServerCertificate=yes;"  
            self.conn_mssql = pyodbc.connect(conn_str)
            self.curr_mssql = self.conn_mssql.cursor()
            
        except Exception as e:
            logging.info("\n<Error> __connect_mssql__ " + str(e))

        finally:
            pass
            #self.curr_mssql.close()
            #self.conn_mssql.close()
    
    #########################
    # __disconnect_mssql__ 
    #########################
    def __disconnect_mssql__(self):
        
        try:
            self.conn_mssql.close()
            
        except Exception as e:
            logging.info("\n<Error> __disconnect_mssql__ " + str(e))

        finally:
            pass

    ################
    # __connect2__ 
    ################
    def __connect2__(self):
        
        try:
            self.conn2 = pymysql.connect(host=parameter.otsuka_factory7['host'],port=parameter.otsuka_factory7['port'],user=parameter.otsuka_factory7['user'],password=parameter.otsuka_factory7['pwd'],database=parameter.otsuka_factory7['db'],charset=parameter.otsuka_factory7['charset'])
            self.curr2 = self.conn2.cursor()

        except Exception as e:
            logging.info("\n<Error> __connect2__ " + str(e))

        finally:
            pass
    
    ###################
    # __disconnect2__
    ###################
    def __disconnect2__(self):
        
        try:
            self.conn2.commit()
            self.conn2.close()

        except Exception as e:
            logging.info("\n<Error> __disconnect2__ : " + str(e))

        finally:
            pass

    ################
    # __connect__ 
    ################
    def __connect__(self):
        
        try:
            self.conn = pymysql.connect(host=parameter.otsuka_factory['host'],port=parameter.otsuka_factory['port'],user=parameter.otsuka_factory['user'],password=parameter.otsuka_factory['pwd'],database=parameter.otsuka_factory['db'],charset=parameter.otsuka_factory['charset'])
            self.curr = self.conn.cursor()

        except Exception as e:
            logging.info("\n<Error> __connect__ " + str(e))

        finally:
            pass

    ###################
    # __disconnect__
    ###################
    def __disconnect__(self):
        
        try:
            self.conn.commit()
            self.conn.close()

        except Exception as e:
            logging.info("\n<Error> __disconnect__ : " + str(e))

        finally:
            pass

    #############################
    # __connect_taipei_1_38__ 
    #############################
    def __connect_taipei_1_38__ (self):
        
        try:
            self.conn = pymysql.connect(host=parameter.otsuka_taipei_1_38['host'],port=parameter.otsuka_taipei_1_38['port'],user=parameter.otsuka_taipei_1_38['user'],password=parameter.otsuka_taipei_1_38['password'],database=parameter.otsuka_taipei_1_38['db'],charset=parameter.otsuka_taipei_1_38['charset'])
            self.curr = self.conn.cursor()

        except Exception as e:
            logging.info(f"\n<Error> __connect_taipei_1_38__ : {str(e)}\n")

        finally:
            pass

    ################################
    # __disconnect_taipei_1_38__ 
    ################################
    def __disconnect_taipei_1_38__ (self):
        
        try:
            self.conn.commit()
            self.conn.close()

        except Exception as e:
            logging.info(f"\n<Error> __disconnect_taipei_1_38__  : {str(e)}\n")

        finally:
            pass



    #############################
    # __connect_factory_1_38__ 
    #############################
    def __connect_factory_1_38__ (self):
        
        try:
            self.conn = pymysql.connect(host=parameter.otsuka_factory1_38['host'],port=parameter.otsuka_factory_1_38['port'],user=parameter.otsuka_factory_1_38['user'],password=parameter.otsuka_factory_1_38['pwd'],database=parameter.otsuka_factory_1_38['db'],charset=parameter.otsuka_factory_1_38['charset'])
            self.curr = self.conn.cursor()

        except Exception as e:
            logging.info(f"\n<Error> __connect_factory_1_38__ : {str(e)}\n")

        finally:
            pass

    ################################
    # __disconnect_factory_1_38__ 
    ################################
    def __disconnect_factory_1_38__ (self):
        
        try:
            self.conn.commit()
            self.conn.close()

        except Exception as e:
            logging.info(f"\n<Error> __disconnect_factory_1_38__  : {str(e)}\n")

        finally:
            pass

    ######################
    # __connect7_1_38__ 
    ######################
    def __connect7_1_38__(self):
        
        ####################
        # mysql.connectot
        ####################
        '''
        try:
            self.conn_7_1_38 = mysql.connector.connect(**parameter.otsuka_factory7_1_38)
            self.curr_7_1_38 = self.conn_7_1_38.cursor()
        except mysql.connector.Error as e:
            logging.info(f"<Error> __connect7_1_38__ : {str(e)}\n")
            
        finally:
            pass
        '''
        ############
        # pymysql
        ############
        try:
            self.conn_7_1_38 = pymysql.connect(host=parameter.otsuka_factory7['host'],port=parameter.otsuka_factory7['port'],user=parameter.otsuka_factory7['user'],password=parameter.otsuka_factory7['pwd'],database=parameter.otsuka_factory7['db'],charset=parameter.otsuka_factory7['charset'])
            self.curr_7_1_38 = self.conn_7_1_38.cursor()

        except Exception as e:
            logging.info("\n<Error> __connect7_1_38__ " + str(e))

        finally:
            pass

    #######################
    # __disconnect7_1_38__
    #######################
    def __disconnect7_1_38__(self):
        
        try:
            self.conn_7_1_38.commit()
            self.conn_7_1_38.close()

        except Exception as e:
            logging.info(f"\n<Error> __disconnect7_1_38__ : {str(e)}\n")

        finally:
            pass
    
    
    ################
    # __connect4__ 
    ################
    def __connect4__(self):
        
        try:
            self.conn = pymysql.connect(host=parameter.otsuka_factory4['host'],port=parameter.otsuka_factory4['port'],user=parameter.otsuka_factory4['user'],password=parameter.otsuka_factory4['pwd'],database=parameter.otsuka_factory4['db'],charset=parameter.otsuka_factory4['charset'])
            self.curr = self.conn.cursor()

        except Exception as e:
            logging.info("\n<Error> __connect4__ " + str(e))

        finally:
            pass

    ###################
    # __disconnect4__
    ###################
    def __disconnect4__(self):
        
        try:
            self.conn.commit()
            self.conn.close()

        except Exception as e:
            logging.info("\n<Error> __disconnect4__ : " + str(e))

        finally:
            pass

